from django.core.paginator import Paginator
from django.http import HttpResponse, StreamingHttpResponse, JsonResponse
from . import models
import datetime
import json
from login import models as loginmoddel
from school import models as schoolmoddel
from common.queSubTable import query_sub_table,CqSchoolvalue
from common import extractors
from common.utils import Code, parse_arguments, FormattedView, require_arguments, decode_id
from common.files import save_img, save_maskfile, read_mask, foll_save_img, write_zippl, ExcelFile
from .height_percent import get_percent, get_height
from django.db import transaction, connection
from wjwsjk import settings
from django.db.models import Count
from django.db import transaction
import django.utils.timezone as timezone
import os
from login import views as loginView



class CaseListView(FormattedView):
    extractor = extractors.AllExtractor()

    # 查询主列表
    @parse_arguments('url')
    def get(self, request, *args, **kwargs):
        filters = self.get_filters(request.user, kwargs)
        if filters is None:
            return self.make_response(None, Code.PERMISSION_DENIED)
        else:
            patient = models.Patient.objects.filter(**filters).all()
            if request.user.is_superuser == 1:
                if 'upmec' in kwargs and kwargs['upmec']:
                    unitlist = loginmoddel.Unit.objects.filter(unit_name__contains = kwargs['upmec']).values_list('id')
                    patient = patient.filter(up_mec__in=unitlist)
            else:
                patient = patient.filter(up_mec=request.user.unit)
            if 'sortby' in kwargs and kwargs['sortby']:
                sortby_map = {
                    'disClass': 'dis_class',
                    'caseNum': 'case_num',
                    'sex': 'sex',
                    'cTime': 'c_time',
                    'name': 'name',
                }
                if 'order' in kwargs and kwargs['order'] == 'desc':
                    patient = patient.order_by('-' + sortby_map[kwargs['sortby']])
                else:
                    patient = patient.order_by(sortby_map[kwargs['sortby']])
            else:
                patient = patient.order_by('-modify_time')
            limit = kwargs['limit']
            paginator = Paginator(patient, limit)  # 每页显示10条
            page = kwargs['currPage']
            if page == '0':
                page = '1'
            pagedata = {}  # 获取分页信息
            pagedata['count'] = paginator.count
            pagedata['num_pages'] = paginator.num_pages
            pagedata['per_page'] = limit
            pagedata['current'] = page
            context = {}
            try:
                list = paginator.page(page).object_list
            except:
                list = paginator.page('1').object_list
            for item in list:
                if item.modify_per and len(item.modify_per) > 0:
                    item.modify_per = loginView.getNameById(item.modify_per)
            contacts = self.extractor.extract(list)
            context['contacts'] = contacts
            context['pagedata'] = pagedata
            return self.make_response(context)

    def get_filters(self, user, source):
        """
        获取查询条件（将请求参数转换为数据库表字段）
        为安全考虑，请求传过来的参数名称尽量不要和数据库中的字段同名
        """

        filters = {}

        if 'caseNum' in source and source['caseNum']:
            filters['case_num__contains'] = source['caseNum']

        if 'gender' in source and source['gender']:
            filters['sex__contains'] = source['gender']

        if 'disclass' in source and source['disclass']:
            filters['dis_class'] = source['disclass']

        if 'name' in source and source['name']:
            filters['name__contains'] = source['name']

        if 'userNum' in source and source['userNum']:
            filters['user_num__contains'] = source['userNum']

        if 'createDateRange' in source and ',' in source['createDateRange']:
            items = source['createDateRange'].split(',')
            filters['c_time__gte'] = items[0]
            filters['c_time__lte'] = items[1]

        filters['del_flg'] = '1'

        return filters

class PatientView(FormattedView):
    extractor = extractors.AllExtractor()
    caseextr = extractors.AddPatExtractor()
    # 查询患者详细数据
    @require_arguments(['queryId'], 'url')
    def get(self, request, *args, **kwargs):
        # 主键id
        if len(kwargs['queryId']) > 0:
            patient_id = decode_id(kwargs['queryId'])
        else:
            return self.make_response(None, Code.MAIN_NULL)
        try:
            patient = models.Patient.objects.get(pk=patient_id)
            return self.make_response(patient)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

    # 添加患者详细数据
    @require_arguments(['disClass'], 'body')
    def put(self, request, *args, **kwargs):
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            patient = models.Patient.objects.get(pk=decode_id(kwargs['queryId']))
            addModifylod(request, patient.id, "10050001")
        else:
            patient = models.Patient()
        # 疾病分类
        patient.dis_class = kwargs['disClass']
        # 患者编号
        if 'userNum' in kwargs:
            patient.user_num = kwargs['userNum']
        # 患者姓名
        if 'name' in kwargs:
            patient.name = kwargs['name']
        # 病历号
        if 'medrecNum' in kwargs:
            patient.medrec_num = kwargs['medrecNum']
        # 社会性别
        if 'sex' in kwargs:
            patient.sex = kwargs['sex']
        # 患者身份证号码
        if 'card' in kwargs:
            patient.card = kwargs['card']
        # 家庭住址
        if 'familyAdress' in kwargs:
            patient.fam_adr = kwargs['familyAdress']
        # 联系人姓名
        if 'contactsName' in kwargs:
            patient.contacts_name = kwargs['contactsName']
        # 与患者关系
        if 'relation' in kwargs:
            patient.relation = kwargs['relation']
        # 联系电话
        if 'contactsNum' in kwargs:
            patient.contacts_num = kwargs['contactsNum']
        # 性腺性别
        if 'gonadalSex' in kwargs:
            patient.gonadal_sex = kwargs['gonadalSex']
        # 初诊时间
        if 'firVisTime' in kwargs and kwargs['firVisTime'] is not None and len(kwargs['firVisTime']):
            time = kwargs['firVisTime'][0:10]
            patient.fir_vis_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        # AGEy
        if 'AGEy' in kwargs:
            patient.AGEy = kwargs['AGEy']
        # AGEm
        if 'AGEm' in kwargs:
            patient.AGEm = kwargs['AGEm']
        # 出生日期
        if 'birthTime' in kwargs and kwargs['birthTime'] is not None and len(kwargs['birthTime']):
            time = kwargs['birthTime'][0:10]
            patient.birth_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        # 主诉
        if 'chiCom' in kwargs:
            patient.chi_com = kwargs['chiCom']
        # 籍贯  细化到市
        if 'natPla' in kwargs:
            patient.nat_pla = kwargs['natPla']
        # FHt cm  父亲身高
        if 'FHt' in kwargs:
            patient.FHt = kwargs['FHt']
        # MHt cm  母亲身高
        if 'MHt' in kwargs:
            patient.MHt = kwargs['MHt']
        # 父亲体重
        if 'FHw' in kwargs:
            patient.FHw = kwargs['FHw']
        # 母亲体重
        if 'MHw' in kwargs:
            patient.MHw = kwargs['MHw']
        # 家族史
        if 'familyHis' in kwargs:
            patient.family_his = kwargs['familyHis']
        # 初潮年龄
        if 'menAge' in kwargs:
            patient.men_age = kwargs['menAge']
        # 有无兄弟姐妹
        if 'isBot' in kwargs:
            patient.is_bot = kwargs['isBot']
        # 胎龄周
        if 'gesWeek' in kwargs:
            patient.ges_week = kwargs['gesWeek']
        # BWtkg  出生体重
        if 'BWt' in kwargs:
            patient.BWt = kwargs['BWt']
        # BLcm   出生身长
        if 'BL' in kwargs:
            patient.BL = kwargs['BL']
        # 剖宫产   剖宫产=1、自然=0
        if 'cesaSec' in kwargs:
            patient.cesa_sec = kwargs['cesaSec']
        # 保胎史   1=无，2=有
        if 'fetProHis' in kwargs:
            patient.fet_pro_his = kwargs['fetProHis']
        # 既往史
        if 'oldHis' in kwargs:
            patient.past_his = kwargs['oldHis']
        # 窒息抢救史
        if 'cesaAsphyxia' in kwargs:
            patient.cesa_asphyxia = kwargs['cesaAsphyxia']
        # 联系人姓名
        if 'contactsName' in kwargs:
            patient.contacts_name = kwargs['contactsName']
        # 与患者关系
        if 'relation' in kwargs:
            patient.relation = kwargs['relation']
        # 联系电话
        if 'contactsNum' in kwargs:
            patient.contacts_num = kwargs['contactsNum']
        # 家庭住址
        if 'famAdr' in kwargs:
            patient.fam_adr = kwargs['famAdr']
        # 患者身份证号码
        if 'card' in kwargs:
            patient.card = kwargs['card']
        # 入组序号
        if 'enNum' in kwargs:
            patient.enrollment_num = kwargs['enNum']
        # 入组时间
        if 'enTime' in kwargs and kwargs['enTime'] is not None and len(kwargs['enTime']):
            time = kwargs['enTime'][0:10]
            patient.enrollment_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        # 所在中心
        if 'hosName' in kwargs:
            patient.hospital_name = kwargs['hosName']
        # 确诊年龄
        if 'firVisAge' in kwargs:
            patient.fir_vis_age = kwargs['firVisAge']
        # 胎次
        if 'parity' in kwargs:
            patient.parity = kwargs['parity']
        # 产次
        if 'proNum' in kwargs:
            patient.pronum = kwargs['proNum']
        # 孕期感染
        if 'preInf' in kwargs:
            patient.pregnancy_infection = kwargs['preInf']
        # IDC(国际疾病分类)
        if 'ICD' in kwargs:
            patient.ICD = kwargs['ICD']
        # 首次提交时间
        if 'oneTime' in kwargs and kwargs['oneTime'] is not None and len(kwargs['oneTime']):
            time = kwargs['oneTime'][0:10]
            patient.one_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        # 是否达终身高
        if 'isfinalhei' in kwargs:
            patient.is_finalhei = kwargs['isfinalhei']
        # 疾病描述
        if 'category_describe' in kwargs:
            patient.category_describe = kwargs['category_describe']
        # 身高
        if 'height' in kwargs:
            patient.height = kwargs['height']
        # 体重
        if 'weight' in kwargs:
            patient.weight = kwargs['weight']
        # bmi值
        if 'Bmi' in kwargs:
            patient.bmi = kwargs['Bmi']
        # 民族
        if 'ethnic' in kwargs:
            patient.ethnic = kwargs['ethnic']
        patient.up_mec = request.user.unit
        patient.check_hospital = request.user.unit
        patient.imp_per = request.user.pk
        patient.modify_per = request.user.pk
        patient.modify_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        patient.del_flg = '1'
        try:
            with transaction.atomic():
                if patient.case_num and len(patient.case_num) > 0:
                    patient.save()
                else:
                    patient.c_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    beiwei = {
                        '10000001': 'Case',
                        '10000002': 'Short',
                        '10000003': 'Sexprecocity',
                        '10000004': 'MAS',
                        '10000005': 'SGA',
                        '10000006': 'JzxShort',
                        '10000007': 'Eltm',
                    }
                    qianzui = 'US-'+beiwei[kwargs['disClass']]+str(timezone.now().year*10000+timezone.now().month*100+timezone.now().day)
                    num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(num=Count('case_num'))
                    nums = num['num']
                    caseNum = qianzui+str(nums+1)
                    case_num = caseNum
                    patient.case_num = case_num
                    patient.up_mec = request.user.unit
                    patient.check_hospital = request.user.unit
                    patient.imp_per = request.user.pk
                    patient.save()
                # 判断检查部位，新增附表
                if kwargs['disClass'] == "10000001":
                    result = modifyorAddCase(patient.pk, kwargs)
                    result.save()
                elif kwargs['disClass'] == "10000002":
                    result = modifyorAddShort(patient.pk, kwargs)
                    result.save()
                elif kwargs['disClass'] == "10000003":
                    result = modifyorAddSexpre(patient.pk, kwargs)
                    result.save()
                elif kwargs['disClass'] == "10000005":
                    result = modifyorAddSGA(patient.pk, kwargs)
                    result.save()
                elif kwargs['disClass'] == "10000004":
                    result = modifyorAddMas(patient.pk, kwargs)
                    result.save()
                    result2 = modifyorAddMasFoll(result.pk, kwargs)
                    result2.save()
                elif kwargs['disClass'] == "10000006":
                    result = modifyorAddjzxShort(patient.pk, kwargs)
                    result.save()
                elif kwargs['disClass'] == "10000007":
                    result = modifyorAddEltm(patient.pk, kwargs)
                    result.save()

            return self.make_response(patient, extractor=self.caseextr)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

    # 删除病例详细数据
    @require_arguments(['queryId'], 'url')
    def delete(self, request, *args, **kwargs):
        # 主键id
        if len(kwargs['queryId']) > 0:
            case_id = decode_id(kwargs['queryId'])
        else:
            return self.make_response(None, Code.MAIN_NULL)
        try:
            patient = models.Patient.objects.get(pk=case_id, del_flg='1')
            patient.del_flg = '0'
            patient.save()
            return self.make_response(None)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

class CaseView(FormattedView):
    extractor = extractors.AllExtractor()
    # 根据主表id查询病例详细数据
    @require_arguments(['queryId'], 'url')
    def get(self, request, *args, **kwargs):
        # 主键id
        if len(kwargs['queryId']) > 0:
            case_id = decode_id(kwargs['queryId'])
        else:
            return self.make_response(None, Code.MAIN_NULL)
        try:
            dis_class = models.Patient.objects.get(pk=case_id).dis_class
            # 根据疾病分类确定分表
            result = query_sub_table(dis_class, case_id)
            return self.make_response(result)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)


# 修改或添加性发育异常
def modifyorAddCase(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            case = models.Case.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            case = models.Case()
        # 病人编号
        if 'userNum' in kwargs:
            case.user_num = kwargs['userNum']
        case.patient_id = casePk
        #  Htcm  现身高
        if 'Ht' in kwargs:
            case.Ht = kwargs['Ht']
        #  变异类型
        if 'mutKind' in kwargs:
            case.mut_kind = kwargs['mutKind']
        #  促肾上腺皮质激素
        if 'ACTH' in kwargs:
            case.ACTH = kwargs['ACTH']
        #  皮质醇
        if 'Hyd' in kwargs:
            case.Hyd = kwargs['Hyd']
        #  17-OHP
        if 'OHP' in kwargs:
            case.OHP = kwargs['OHP']
        #  硫酸脱氢表雄酮
        if 'DHEA' in kwargs:
            case.DHEAS = kwargs['DHEA']
        #  是否有生物样本库
        if 'biolog' in kwargs:
            case.biolog = kwargs['biolog']
        #  生物样本库名称
        if 'biologBank' in kwargs:
            case.biolog_bank = kwargs['biologBank']
        #  诊断
        if 'diagnosis' in kwargs:
            case.diagnosis = kwargs['diagnosis']
        # HSDS  现身高标准差（单位sds）
        if 'HSDS' in kwargs:
            case.HSDS = kwargs['HSDS']
        # Wtkg  现体重
        if 'Wt' in kwargs:
            case.Wt = kwargs['Wt']
        # WSDS  现体重标准差
        if 'WSDS' in kwargs:
            case.WSDS = kwargs['WSDS']
        # 阴茎长cm
        if 'penileLength' in kwargs:
            case.penile_length = kwargs['penileLength']
        # 阴茎直径cm
        if 'penileDia' in kwargs:
            case.penile_dia = kwargs['penileDia']
        # 睾丸容积ml
        if 'tesVolume' in kwargs:
            case.tes_volume = kwargs['tesVolume']
        # Prader分期
        if 'prader' in kwargs:
            case.prader = kwargs['prader']
        # 尿道口位置
        if 'locaUreOri' in kwargs:
            case.loca_ure_ori = kwargs['locaUreOri']
        # 右睾丸位置
        if 'rigTesPos' in kwargs:
            case.rig_tes_pos = kwargs['rigTesPos']
        # 左睾丸位置
        if 'lefTesPos' in kwargs:
            case.lef_tes_pos = kwargs['lefTesPos']
        # BA岁骨龄
        if 'boneAge' in kwargs:
            case.bone_age = kwargs['boneAge']
        # LH
        if 'LH' in kwargs:
            case.LH = kwargs['LH']
        # FSH
        if 'FSH' in kwargs:
            case.FSH = kwargs['FSH']
        # LHmax
        if 'LHmax' in kwargs:
            case.LHmax = kwargs['LHmax']
        # FSHmax
        if 'FSHmax' in kwargs:
            case.FSHmax = kwargs['FSHmax']
        # T
        if 'T' in kwargs:
            case.T = kwargs['T']
        # E2
        if 'E2' in kwargs:
            case.E2 = kwargs['E2']
        # DHT
        if 'DHT' in kwargs:
            case.DHT = kwargs['DHT']
        # 游离睾酮
        if 'FT' in kwargs:
            case.FT = kwargs['FT']
        # SHBG
        if 'SHBG' in kwargs:
            case.SHBG = kwargs['SHBG']
        # IGF1
        if 'IGF1' in kwargs:
            case.IGF1 = kwargs['IGF1']
        # IGFBP-3（ug/ml）
        if 'IGFBP3' in kwargs:
            case.IGFBP3 = kwargs['IGFBP3']
        # AMH 抗缪勒管激素
        if 'AMH' in kwargs:
            case.AMH = kwargs['AMH']
        # INHB 抑制素B
        if 'INHB' in kwargs:
            case.INHB = kwargs['INHB']
        # HCG激发试验
        if 'HCG' in kwargs:
            case.HCG = kwargs['HCG']
        # 标准HCG激发T
        if 'HCGT' in kwargs:
            case.HCGT = kwargs['HCGT']
        # 标准HCG激发激发DHT
        if 'HCGDHT' in kwargs:
            case.HCGDHT = kwargs['HCGDHT']
        # 标准HCG激发激发AD
        if 'HCGAD' in kwargs:
            case.HCGAD = kwargs['HCGAD']
        # 延长HCG激发T
        if 'HCGT_ext' in kwargs:
            case.HCGT_ext = kwargs['HCGT_ext']
        # 延长HCG激发激发DHT
        if 'HCGDHT_ext' in kwargs:
            case.HCGDHT_ext = kwargs['HCGDHT_ext']
        # 延长HCG激发激发AD
        if 'HCGAD_ext' in kwargs:
            case.HCGAD_ext = kwargs['HCGAD_ext']
        # 特殊核型
        if 'speKar' in kwargs:
            case.spe_kar = kwargs['speKar']
        # SRY基因
        if 'SRY' in kwargs:
            case.SRY = kwargs['SRY']
        # 致病基因名称
        if 'genData' in kwargs:
            case.gen_mut_name = json.dumps(kwargs['genData'])
        # 突变来源
        if 'sourMut' in kwargs:
            case.sour_mut = kwargs['sourMut']
        # 核酸变异
        if 'baseMut' in kwargs:
            case.base_mut = kwargs['baseMut']
        # 氨基酸变异
        if 'amiAciMut' in kwargs:
            case.ami_aci_mut = kwargs['amiAciMut']
        # 其他
        if 'other' in kwargs:
            case.other = kwargs['other']
        # 手术情况
        if 'operation' in kwargs:
            case.operation = kwargs['operation']
        # 病理结果
        if 'patRes' in kwargs:
            case.pat_res = kwargs['patRes']
        # 处理意见
        if 'hanOpi' in kwargs:
            case.han_opi = kwargs['hanOpi']
        # 雄烯二酮
        if 'AD' in kwargs:
            case.AD = kwargs['AD']
        # 图像说明
        if 'uterusOne' in kwargs:
            txsm = '{"uterusOne":"' + kwargs['uterusOne'] + '"'
        else:
            txsm = '{"uterusOne":"null"'
        if 'uterusTwo' in kwargs:
            txsm += ',"uterusTwo":"'+kwargs['uterusTwo']+'"'
        else:
            txsm += ',"uterusTwo":"null"'
        if 'uterusThr' in kwargs:
            txsm += ',"uterusThr":"'+kwargs['uterusThr']+'"'
        else:
            txsm += ',"uterusThr":"null"'
        if 'intima' in kwargs:
            txsm += ',"intima":"'+kwargs['intima']+'"'
        else:
            txsm += ',"intima":"null"'
        if 'ovaLeftOne' in kwargs:
            txsm += ',"ovaLeftOne":"'+kwargs['ovaLeftOne']+'"'
        else:
            txsm += ',"ovaLeftOne":"null"'
        if 'ovaLeftTwo' in kwargs:
            txsm += ',"ovaLeftTwo":"'+kwargs['ovaLeftTwo']+'"'
        else:
            txsm += ',"ovaLeftTwo":"null"'
        if 'ovaLeftThr' in kwargs:
            txsm += ',"ovaLeftThr":"'+kwargs['ovaLeftThr']+'"'
        else:
            txsm += ',"ovaLeftThr":"null"'
        if 'ovaRightOne' in kwargs:
            txsm += ',"ovaRightOne":"'+kwargs['ovaRightOne']+'"'
        else:
            txsm += ',"ovaRightOne":"null"'
        if 'ovaRightTwo' in kwargs:
            txsm += ',"ovaRightTwo":"'+kwargs['ovaRightTwo']+'"'
        else:
            txsm += ',"ovaRightTwo":"null"'
        if 'ovaRightThr' in kwargs:
            txsm += ',"ovaRightThr":"'+kwargs['ovaRightThr']+'"'
        else:
            txsm += ',"ovaRightThr":"null"'
        if 'follDiameter' in kwargs:
            txsm += ',"follDiameter":"'+kwargs['follDiameter']+'"'
        else:
            txsm += ',"follDiameter":"null"'
        if 'testisLeftOne' in kwargs:
            txsm += ',"testisLeftOne":"'+kwargs['testisLeftOne']+'"'
        else:
            txsm += ',"testisLeftOne":"null"'
        if 'testisLeftTwo' in kwargs:
            txsm += ',"testisLeftTwo":"'+kwargs['testisLeftTwo']+'"'
        else:
            txsm += ',"testisLeftTwo":"null"'
        if 'testisLeftThr' in kwargs:
            txsm += ',"testisLeftThr":"'+kwargs['testisLeftThr']+'"'
        else:
            txsm += ',"testisLeftThr":"null"'
        if 'testisRightOne' in kwargs:
            txsm += ',"testisRightOne":"'+kwargs['testisRightOne']+'"'
        else:
            txsm += ',"testisRightOne":"null"'
        if 'testisRightTwo' in kwargs:
            txsm += ',"testisRightTwo":"'+kwargs['testisRightTwo']+'"'
        else:
            txsm += ',"testisRightTwo":"null"'
        if 'testisRightThr' in kwargs:
            txsm += ',"testisRightThr":"' + kwargs['testisRightThr'] + '"}'
        else:
            txsm += ',"testisRightThr":"null"}'
        case.bscanExplain = txsm
        # 生殖器评估
        if 'genitals' in kwargs:
            case.genitals = kwargs['genitals']
        # 双乳发育分期
        be = {}
        if 'breastDev' in kwargs:
            be["breastDev"] = kwargs['breastDev']
        if 'breastDevRight' in kwargs:
            be["breastDevRight"] = kwargs['breastDevRight']
        case.breast_dev = json.dumps(be)
        # 外生殖器分期
        if 'exGenitalia' in kwargs:
            case.ex_genitalia = kwargs['exGenitalia']
        # 阴毛分期
        if 'pubicHair' in kwargs:
            case.pubic_hair = kwargs['pubicHair']
        # 其他
        if 'bodyOther' in kwargs:
            case.body_other = kwargs['bodyOther']
        # 磁共振
        if 'MRI' in kwargs:
            case.MRI = kwargs['MRI']
        # 其他
        if 'supOther' in kwargs:
            case.sup_other = kwargs['supOther']
        case.c_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result = case
    except:
        result = False
    return result

# 修改或添加矮小症
def modifyorAddShort(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            short = models.Short.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            short = models.Short()
        # 病人编号
        if 'userNum' in kwargs:
            short.user_num = kwargs['userNum']
        short.patient_id = casePk
        #  家族史
        if 'famHis' in kwargs:
            short.fam_his = json.dumps(kwargs['famHis'])
        #  运动发育落后
        if 'motDevBack' in kwargs:
            ydfy = '{"motDevBack":"' + kwargs['motDevBack'] + '",'
        else:
            ydfy = '{"motDevBack":"null",'
        if 'sport' in kwargs:
            ydfy += '"sport":"' + kwargs['sport'] + '"}'
        else:
            ydfy += '"sport":"null"}'
        short.mot_dev_back = ydfy
        #  语言发育落后
        if 'lanDevBack' in kwargs:
            yafy = '{"lanDevBack":"' + kwargs['lanDevBack'] + '",'
        else:
            yafy = '{"lanDevBack":"null",'
        if 'language' in kwargs:
            yafy += '"language":"' + kwargs['language'] + '"}'
        else:
            yafy += '"language":"null"}'
        short.lan_dev_back = yafy
        #  智力发育落后
        if 'intDevBack' in kwargs:
            zlfy = '{"intDevBack":"' + kwargs['intDevBack'] + '",'
        else:
            zlfy = '{"intDevBack":"null",'
        if 'intelligence' in kwargs:
            zlfy += '"intelligence":"' + kwargs['intelligence'] + '"}'
        else:
            zlfy += '"intelligence":"null"}'
        short.int_dev_back = zlfy
        #  听力异常
        if 'abnHear' in kwargs:
            tlyc = '{"abnHear":"' + kwargs['abnHear'] + '",'
        else:
            tlyc = '{"abnHear":"null",'
        if 'hear' in kwargs:
            tlyc += '"hear":"' + kwargs['hear'] + '"}'
        else:
            tlyc += '"hear":"null"}'
        short.abn_hear = tlyc
        #  反复感染史
        if 'recInfHis' in kwargs:
            ffgr = '{"recInfHis":"' + kwargs['recInfHis'] + '",'
        else:
            ffgr = '{"recInfHis":"null",'
        if 'infection' in kwargs:
            ffgr += '"infection":"' + kwargs['infection'] + '"}'
        else:
            ffgr += '"infection":"null"}'
        short.rec_inf_his = ffgr
        #  抽搐史
        if 'conHis' in kwargs:
            short.con_his = kwargs['conHis']
        #  其他
        if 'pastOther' in kwargs:
            short.past_other = kwargs['pastOther']
        #  病史
        if 'firVisTime' in kwargs:
            bs = '{"firVisTime":"' + kwargs['firVisTime'] + '",'
        else:
            bs = '{"firVisTime":"null",'
        if 'morbidAge' in kwargs:
            bs += '"morbidAge":"'+kwargs['morbidAge']+'",'
        else:
            bs += '"morbidAge":"null",'
        if 'chiefCom' in kwargs:
            bs += '"chiefCom":"'+kwargs['chiefCom']+'",'
        else:
            bs += '"chiefCom":"null",'
        if 'growRate' in kwargs:
            bs += '"growRate":"'+kwargs['growRate']+'",'
        else:
            bs += '"growRate":"null",'
        if 'rate' in kwargs:
            bs += '"rate":"'+kwargs['rate']+'",'
        else:
            bs += '"rate":"null",'
        if 'menarchyTime' in kwargs:
            bs += '"menarchyTime":"'+kwargs['menarchyTime']+'",'
        else:
            bs += '"menarchyTime":"null",'
        if 'menarchy' in kwargs:
            bs += '"menarchy":"'+kwargs['menarchy']+'"}'
        else:
            bs += '"menarchy":"null"}'
        short.med_his = bs
        #  体格检查
        if 'height' in kwargs:
            tgjc = '{"height":"' + kwargs['height'] + '",'
        else:
            tgjc = '{"height":"null",'
        if 'weight' in kwargs:
            tgjc += '"weight":"' + kwargs['weight'] + '",'
        else:
            tgjc += '"weight":"null",'
        if 'Bmi' in kwargs:
            tgjc += '"Bmi":"' + kwargs['Bmi'] + '",'
        else:
            tgjc += '"Bmi":"null",'
        if 'breastDev' in kwargs:
            tgjc += '"breastDev":"' + kwargs['breastDev'] + '",'
        else:
            tgjc += '"breastDev":"null",'
        if 'breastDevRight' in kwargs:
            tgjc += '"breastDevRight":"' + kwargs['breastDevRight'] + '",'
        else:
            tgjc += '"breastDevRight":"null",'
        if 'exGenitalia' in kwargs:
            tgjc += '"exGenitalia":"' + kwargs['exGenitalia'] + '",'
        else:
            tgjc += '"exGenitalia":"null",'
        if 'pubicHair' in kwargs:
            tgjc += '"pubicHair":"' + kwargs['pubicHair'] + '",'
        else:
            tgjc += '"pubicHair":"null",'
        if 'armLength' in kwargs:
            tgjc += '"armLength":"' + kwargs['armLength'] + '",'
        else:
            tgjc += '"armLength":"null",'
        if 'specialFace' in kwargs:
            tgjc += '"specialFace":"' + kwargs['specialFace'] + '",'
        else:
            tgjc += '"specialFace":"null",'
        if 'specialFaceDesc' in kwargs:
            tgjc += '"specialFaceDesc":"' + kwargs['specialFaceDesc'] + '",'
        else:
            tgjc += '"specialFaceDesc":"null",'
        if 'scoliosis' in kwargs:
            tgjc += '"scoliosis":"' + kwargs['scoliosis'] + '",'
        else:
            tgjc += '"scoliosis":"null",'
        if 'scoliosisDegree' in kwargs:
            tgjc += '"scoliosisDegree":"' + kwargs['scoliosisDegree'] + '",'
        else:
            tgjc += '"scoliosisDegree":"null",'
        if 'rashDescribe' in kwargs:
            tgjc += '"rashDescribe":"' + kwargs['rashDescribe'] + '",'
        else:
            tgjc += '"rashDescribe":"null",'
        if 'rash' in kwargs:
            tgjc += '"rash":"' + kwargs['rash'] + '"}'
        else:
            tgjc += '"rash":"null"}'
        short.phy_exa = tgjc
        #  实验室检查
        if 'LH' in kwargs:
            sysbg = '{"LH":"' + kwargs['LH'] + '",'
        else:
            sysbg = '{"LH":"null",'
        if 'FSH' in kwargs:
            sysbg += '"FSH":"' + kwargs['FSH'] + '",'
        else:
            sysbg += '"FSH":"null",'
        if 'E2' in kwargs:
            sysbg += '"E2":"' + kwargs['E2'] + '",'
        else:
            sysbg += '"E2":"null",'
        if 'T' in kwargs:
            sysbg += '"T":"' + kwargs['T'] + '",'
        else:
            sysbg += '"T":"null",'
        if 'PRL' in kwargs:
            sysbg += '"PRL":"' + kwargs['PRL'] + '",'
        else:
            sysbg += '"PRL":"null",'
        if 'IGF' in kwargs:
            sysbg += '"IGF":"' + kwargs['IGF'] + '",'
        else:
            sysbg += '"IGF":"null",'
        if 'IGFBP3' in kwargs:
            sysbg += '"IGFBP3":"' + kwargs['IGFBP3'] + '",'
        else:
            sysbg += '"IGFBP3":"null",'
        if 'thyroid' in kwargs:
            sysbg += '"thyroid":"' + kwargs['thyroid'] + '",'
        else:
            sysbg += '"thyroid":"null",'
        if 'thyroidDescribe' in kwargs:
            sysbg += '"thyroidDescribe":"' + kwargs['thyroidDescribe'] + '",'
        else:
            sysbg += '"thyroidDescribe":"null",'
        if 'ACTH' in kwargs:
            sysbg += '"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            sysbg += '"ACTH":"null",'
        if 'cortisol' in kwargs:
            sysbg += '"cortisol":"' + kwargs['cortisol'] + '",'
        else:
            sysbg += '"cortisol":"null",'
        if 'DHEAS' in kwargs:
            sysbg += '"DHEAS":"' + kwargs['DHEAS'] + '",'
        else:
            sysbg += '"DHEAS":"null",'
        if 'OHP' in kwargs:
            sysbg += '"OHP":"' + kwargs['OHP'] + '",'
        else:
            sysbg += '"OHP":"null",'
        if 'blood' in kwargs:
            sysbg += '"blood":"' + kwargs['blood'] + '",'
        else:
            sysbg += '"blood":"null",'
        if 'bloodDescribe' in kwargs:
            sysbg += '"bloodDescribe":"' + kwargs['bloodDescribe'] + '",'
        else:
            sysbg += '"bloodDescribe":"null",'
        if 'urinalysis' in kwargs:
            sysbg += '"urinalysis":"' + kwargs['urinalysis'] + '",'
        else:
            sysbg += '"urinalysis":"null",'
        if 'urinalysisDescribe' in kwargs:
            sysbg += '"urinalysisDescribe":"' + kwargs['urinalysisDescribe'] + '",'
        else:
            sysbg += '"urinalysisDescribe":"null",'
        if 'LAKLGE' in kwargs:
            sysbg += '"LAKLGE":"' + kwargs['LAKLGE'] + '",'
        else:
            sysbg += '"LAKLGE":"null",'
        if 'laklgeDescribe' in kwargs:
            sysbg += '"laklgeDescribe":"' + kwargs['laklgeDescribe'] + '",'
        else:
            sysbg += '"laklgeDescribe":"null",'
        if 'HBs' in kwargs:
            sysbg += '"HBs":"' + kwargs['HBs'] + '",'
        else:
            sysbg += '"HBs":"null",'
        if 'HBsDescribe' in kwargs:
            sysbg += '"HBsDescribe":"' + kwargs['HBsDescribe'] + '",'
        else:
            sysbg += '"HBsDescribe":"null",'
        if 'LHFSHTime' in kwargs:
            sysbg += '"LHFSHTime":"' + kwargs['LHFSHTime'] + '",'
        else:
            sysbg += '"LHFSHTime":"null",'
        if 'E2Time' in kwargs:
            sysbg += '"E2Time":"' + kwargs['E2Time'] + '",'
        else:
            sysbg += '"E2Time":"null",'
        if 'TTime' in kwargs:
            sysbg += '"TTime":"' + kwargs['TTime'] + '",'
        else:
            sysbg += '"TTime":"null",'
        if 'PRLTime' in kwargs:
            sysbg += '"PRLTime":"' + kwargs['PRLTime'] + '",'
        else:
            sysbg += '"PRLTime":"null",'
        if 'IGFBPTime' in kwargs:
            sysbg += '"IGFBPTime":"' + kwargs['IGFBPTime'] + '",'
        else:
            sysbg += '"IGFBPTime":"null",'
        if 'thyroidTime' in kwargs:
            sysbg += '"thyroidTime":"' + kwargs['thyroidTime'] + '",'
        else:
            sysbg += '"thyroidTime":"null",'
        if 'ACTHTime' in kwargs:
            sysbg += '"ACTHTime":"' + kwargs['ACTHTime'] + '",'
        else:
            sysbg += '"ACTHTime":"null",'
        if 'cortisolTime' in kwargs:
            sysbg += '"cortisolTime":"' + kwargs['cortisolTime'] + '",'
        else:
            sysbg += '"cortisolTime":"null",'
        if 'DHEATime' in kwargs:
            sysbg += '"DHEATime":"' + kwargs['DHEATime'] + '",'
        else:
            sysbg += '"DHEATime":"null",'
        if 'OHPTime' in kwargs:
            sysbg += '"OHPTime":"' + kwargs['OHPTime'] + '",'
        else:
            sysbg += '"OHPTime":"null",'
        if 'bloodTime' in kwargs:
            sysbg += '"bloodTime":"' + kwargs['bloodTime'] + '",'
        else:
            sysbg += '"bloodTime":"null",'
        if 'urinalysisTime' in kwargs:
            sysbg += '"urinalysisTime":"' + kwargs['urinalysisTime'] + '",'
        else:
            sysbg += '"urinalysisTime":"null",'
        if 'LAKLGETime' in kwargs:
            sysbg += '"LAKLGETime":"' + kwargs['LAKLGETime'] + '",'
        else:
            sysbg += '"LAKLGETime":"null",'
        if 'HBsTime' in kwargs:
            sysbg += '"HBsTime":"' + kwargs['HBsTime'] + '",'
        else:
            sysbg += '"HBsTime":"null",'
        if 'ghTime' in kwargs:
            sysbg += '"ghTime":"' + kwargs['ghTime'] + '",'
        else:
            sysbg += '"ghTime":"null",'
        if 'glyHemA' in kwargs:
            sysbg += '"glyHemA":"' + kwargs['glyHemA'] + '",'
        else:
            sysbg += '"glyHemA":"null",'
        if 'glyHemATime' in kwargs:
            sysbg += '"glyHemATime":"' + kwargs['glyHemATime'] + '",'
        else:
            sysbg += '"glyHemATime":"null",'
        if 'fasBloodGlu' in kwargs:
            sysbg += '"fasBloodGlu":"' + kwargs['fasBloodGlu'] + '",'
        else:
            sysbg += '"fasBloodGlu":"null",'
        if 'fasBloodGluTime' in kwargs:
            sysbg += '"fasBloodGluTime":"' + kwargs['fasBloodGluTime'] + '",'
        else:
            sysbg += '"fasBloodGluTime":"null",'
        if 'fasInsulin' in kwargs:
            sysbg += '"fasInsulin":"' + kwargs['fasInsulin'] + '",'
        else:
            sysbg += '"fasInsulin":"null",'
        if 'fasInsulinTime' in kwargs:
            sysbg += '"fasInsulinTime":"' + kwargs['fasInsulinTime'] + '",'
        else:
            sysbg += '"fasInsulinTime":"null",'
        if 'glyHem' in kwargs:
            sysbg += '"glyHem":"' + kwargs['glyHem'] + '",'
        else:
            sysbg += '"glyHem":"null",'
        if 'glyHemTime' in kwargs:
            sysbg += '"glyHemTime":"' + kwargs['glyHemTime'] + '",'
        else:
            sysbg += '"glyHemTime":"null",'
        if 'gh' in kwargs:
            sysbg += '"gh":"' + kwargs['gh'] + '"}'
        else:
            sysbg += '"gh":"null"}'
        short.lab_exa = sysbg
        #  心电图
        if 'electdiogram' in kwargs:
            short.electr = kwargs['electdiogram']
        #  性腺B超
        if 'uterusOne' in kwargs:
            xxbc = '{"uterusOne":"' + kwargs['uterusOne'] + '",'
        else:
            xxbc = '{"uterusOne":"null",'
        if 'uterusTwo' in kwargs:
            xxbc += '"uterusTwo":"' + kwargs['uterusTwo'] + '",'
        else:
            xxbc += '"uterusTwo":"null",'
        if 'uterusThr' in kwargs:
            xxbc += '"uterusThr":"' + kwargs['uterusThr'] + '",'
        else:
            xxbc += '"uterusThr":"null",'
        if 'cervixLong' in kwargs:
            xxbc += '"cervixLong":"' + kwargs['cervixLong'] + '",'
        else:
            xxbc += '"cervixLong":"null",'
        if 'intima' in kwargs:
            xxbc += '"intima":"' + kwargs['intima'] + '",'
        else:
            xxbc += '"intima":"null",'
        if 'ovaLeftOne' in kwargs:
            xxbc += '"ovaLeftOne":"' + kwargs['ovaLeftOne'] + '",'
        else:
            xxbc += '"ovaLeftOne":"null",'
        if 'ovaLeftTwo' in kwargs:
            xxbc += '"ovaLeftTwo":"' + kwargs['ovaLeftTwo'] + '",'
        else:
            xxbc += '"ovaLeftTwo":"null",'
        if 'ovaLeftThr' in kwargs:
            xxbc += '"ovaLeftThr":"' + kwargs['ovaLeftThr'] + '",'
        else:
            xxbc += '"ovaLeftThr":"null",'
        if 'ovaRightOne' in kwargs:
            xxbc += '"ovaRightOne":"' + kwargs['ovaRightOne'] + '",'
        else:
            xxbc += '"ovaRightOne":"null",'
        if 'ovaRightTwo' in kwargs:
            xxbc += '"ovaRightTwo":"' + kwargs['ovaRightTwo'] + '",'
        else:
            xxbc += '"ovaRightTwo":"null",'
        if 'ovaRightThr' in kwargs:
            xxbc += '"ovaRightThr":"' + kwargs['ovaRightThr'] + '",'
        else:
            xxbc += '"ovaRightThr":"null",'
        if 'follDiameter' in kwargs:
            xxbc += '"follDiameter":"' + kwargs['follDiameter'] + '",'
        else:
            xxbc += '"follDiameter":"null",'
        if 'isCyst' in kwargs:
            xxbc += '"isCyst":"' + kwargs['isCyst'] + '",'
        else:
            xxbc += '"isCyst":"null",'
        if 'cyst' in kwargs:
            xxbc += '"cyst":"' + kwargs['cyst'] + '",'
        else:
            xxbc += '"cyst":"null",'
        if 'cystOne' in kwargs:
            xxbc += '"cystOne":"' + kwargs['cystOne'] + '",'
        else:
            xxbc += '"cystOne":"null",'
        if 'cystTwo' in kwargs:
            xxbc += '"cystTwo":"' + kwargs['cystTwo'] + '",'
        else:
            xxbc += '"cystTwo":"null",'
        if 'cystThr' in kwargs:
            xxbc += '"cystThr":"' + kwargs['cystThr'] + '",'
        else:
            xxbc += '"cystThr":"null",'
        if 'cystDescribe' in kwargs:
            xxbc += '"cystDescribe":"' + kwargs['cystDescribe'] + '",'
        else:
            xxbc += '"cystDescribe":"null",'
        if 'testisLeftOne' in kwargs:
            xxbc += '"testisLeftOne":"' + kwargs['testisLeftOne'] + '",'
        else:
            xxbc += '"testisLeftOne":"null",'
        if 'testisLeftTwo' in kwargs:
            xxbc += '"testisLeftTwo":"' + kwargs['testisLeftTwo'] + '",'
        else:
            xxbc += '"testisLeftTwo":"null",'
        if 'testisLeftThr' in kwargs:
            xxbc += '"testisLeftThr":"' + kwargs['testisLeftThr'] + '",'
        else:
            xxbc += '"testisLeftThr":"null",'
        if 'testisLeftLon' in kwargs:
            xxbc += '"testisLeftLon":"' + kwargs['testisLeftLon'] + '",'
        else:
            xxbc += '"testisLeftLon":"null",'
        if 'testisRightOne' in kwargs:
            xxbc += '"testisRightOne":"' + kwargs['testisRightOne'] + '",'
        else:
            xxbc += '"testisRightOne":"null",'
        if 'testisRightTwo' in kwargs:
            xxbc += '"testisRightTwo":"' + kwargs['testisRightTwo'] + '",'
        else:
            xxbc += '"testisRightTwo":"null",'
        if 'testisRightThr' in kwargs:
            xxbc += '"testisRightThr":"' + kwargs['testisRightThr'] + '",'
        else:
            xxbc += '"testisRightThr":"null",'
        if 'MRI' in kwargs:
            xxbc += '"MRI":"' + kwargs['MRI'] + '",'
        else:
            xxbc += '"MRI":"null",'
        if 'ThyroidLBGradation' in kwargs:
            xxbc += '"ThyroidLBGradation":"' + kwargs['ThyroidLBGradation'] + '",'
        else:
            xxbc += '"ThyroidLBGradation":"null",'
        if 'ThyroidLBSize' in kwargs:
            xxbc += '"ThyroidLBSize":"' + kwargs['ThyroidLBSize'] + '",'
        else:
            xxbc += '"ThyroidLBSize":"null",'
        if 'ThyroidLBLesions' in kwargs:
            xxbc += '"ThyroidLBLesions":"' + kwargs['ThyroidLBLesions'] + '",'
        else:
            xxbc += '"ThyroidLBLesions":"null",'
        if 'ThyroidLBOther' in kwargs:
            xxbc += '"ThyroidLBOther":"' + kwargs['ThyroidLBOther'] + '",'
        else:
            xxbc += '"ThyroidLBOther":"null",'
        if 'ThyroidLB' in kwargs:
            xxbc += '"ThyroidLB":"' + kwargs['ThyroidLB'] + '",'
        else:
            xxbc += '"ThyroidLB":"null",'
        if 'ThyroidRB' in kwargs:
            xxbc += '"ThyroidRB":"' + kwargs['ThyroidRB'] + '",'
        else:
            xxbc += '"ThyroidRB":"null",'
        if 'ThyroidRBGradation' in kwargs:
            xxbc += '"ThyroidRBGradation":"' + kwargs['ThyroidRBGradation'] + '",'
        else:
            xxbc += '"ThyroidRBGradation":"null",'
        if 'ThyroidRBSize' in kwargs:
            xxbc += '"ThyroidRBSize":"' + kwargs['ThyroidRBSize'] + '",'
        else:
            xxbc += '"ThyroidRBSize":"null",'
        if 'ThyroidRBLesions' in kwargs:
            xxbc += '"ThyroidRBLesions":"' + kwargs['ThyroidRBLesions'] + '",'
        else:
            xxbc += '"ThyroidRBLesions":"null",'
        if 'ThyroidRBOther' in kwargs:
            xxbc += '"ThyroidRBOther":"' + kwargs['ThyroidRBOther'] + '",'
        else:
            xxbc += '"ThyroidRBOther":"null",'
        if 'mriDescribe' in kwargs:
            xxbc += '"mriDescribe":"' + kwargs['mriDescribe'] + '",'
        else:
            xxbc += '"mriDescribe":"null",'
        if 'testisRightLon' in kwargs:
            xxbc += '"testisRightLon":"' + kwargs['testisRightLon'] + '"}'
        else:
            xxbc += '"testisRightLon":"null"}'
        short.gon_B_ult = xxbc
        #  诊疗方案
        if 'diaPlan' in kwargs:
            zlfa = '{"diaPlan":"' + kwargs['diaPlan'] + '",'
        else:
            zlfa = '{"diaPlan":"null",'
        if 'rhGH' in kwargs:
            zlfa += '"rhGH":"' + kwargs['rhGH'] + '",'
        else:
            zlfa += '"rhGH":"null",'
        if 'GnRHa' in kwargs:
            zlfa += '"GnRHa":"' + kwargs['GnRHa'] + '",'
        else:
            zlfa += '"GnRHa":"null",'
        if 'GnRHadose' in kwargs:
            zlfa += '"GnRHadose":"' + kwargs['GnRHadose'] + '",'
        else:
            zlfa += '"GnRHadose":"null",'
        if 'genData' in kwargs:
            zlfa += '"genData":"' + json.dumps(kwargs['genData']) + '",'
        else:
            zlfa += '"genData":"null",'
        if 'rhGHdoseKG' in kwargs:
            zlfa += '"rhGHdoseKG":"' +  kwargs['rhGHdoseKG'] + '",'
        else:
            zlfa += '"rhGHdoseKG":"null",'
        if 'PEGrhGHdose' in kwargs:
            zlfa += '"PEGrhGHdose":"' +  kwargs['PEGrhGHdose'] + '",'
        else:
            zlfa += '"PEGrhGHdose":"null",'
        if 'rhCustomizationDiaPlan' in kwargs:
            zlfa += '"rhCustomizationDiaPlan":"' +  kwargs['rhCustomizationDiaPlan'] + '",'
        else:
            zlfa += '"rhCustomizationDiaPlan":"null",'
        if 'rhCustomizationPrompt' in kwargs:
            zlfa += '"rhCustomizationPrompt":"' +  kwargs['rhCustomizationPrompt'] + '",'
        else:
            zlfa += '"rhCustomizationPrompt":"null",'
        if 'PEGrhCustomizationPrompt' in kwargs:
            zlfa += '"PEGrhCustomizationPrompt":"' +  kwargs['PEGrhCustomizationPrompt'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPrompt":"null",'
        if 'rhCustomizationPromptKG' in kwargs:
            zlfa += '"rhCustomizationPromptKG":"' +  kwargs['rhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"rhCustomizationPromptKG":"null",'
        if 'PEGrhCustomizationPromptKG' in kwargs:
            zlfa += '"PEGrhCustomizationPromptKG":"' +  kwargs['PEGrhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPromptKG":"null",'
        if 'PEGrhGHdoseKG' in kwargs:
            zlfa += '"PEGrhGHdoseKG":"' +  kwargs['PEGrhGHdoseKG'] + '",'
        else:
            zlfa += '"PEGrhGHdoseKG":"null",'
        if 'planData' in kwargs:
            zlfa += '"planData":"' +  json.dumps(kwargs['planData']) + '",'
        else:
            zlfa += '"planData":"null",'
        if 'otherMedicine' in kwargs:
            zlfa += '"otherMedicine":"' +  json.dumps(kwargs['otherMedicine']) + '",'
        else:
            zlfa += '"otherMedicine":"null",'
        if 'rhGHdose' in kwargs:
            zlfa += '"rhGHdose":"' + kwargs['rhGHdose'] + '"}'
        else:
            zlfa += '"rhGHdose":"null"}'
        short.dia_trea_plan = zlfa
        #  生物样本库
        if 'bioBank' in kwargs:
            swkyb = '{"bioBank":"' + kwargs['bioBank'] + '",'
        else:
            swkyb = '{"bioBank":"null",'
        if 'sampleId' in kwargs:
            swkyb += '"sampleId":"' + kwargs['sampleId'] + '",'
        else:
            swkyb += '"sampleId":"null",'
        if 'sampleClass' in kwargs:
            swkyb += '"sampleClass":"' + str(kwargs['sampleClass']) + '"}'
        else:
            swkyb += '"sampleClass":"null"}'
        short.bio_sam_bank = swkyb
        #  父亲生物样本库
        if 'bioBankFa' in kwargs:
            swkyb = '{"bioBankFa":"' + kwargs['bioBankFa'] + '"'
        else:
            swkyb = '{"bioBankFa":"null"'
        if 'sampleIdFa' in kwargs:
            swkyb += ',"sampleIdFa":"' + kwargs['sampleIdFa'] + '"'
        else:
            swkyb += ',"sampleIdFa":"null"'
        if 'sampleClassFa' in kwargs:
            swkyb += ',"sampleClassFa":"' + str(kwargs['sampleClassFa']) + '"}'
        else:
            swkyb += ',"sampleClassFa":"null"}'
        short.f_bio_sam_bank = swkyb
        #  母亲生物样本库
        if 'bioBankMo' in kwargs:
            swkyb = '{"bioBankMo":"' + kwargs['bioBankMo'] + '"'
        else:
            swkyb = '{"bioBankMo":"null"'
        if 'sampleIdMo' in kwargs:
            swkyb += ',"sampleIdMo":"' + kwargs['sampleIdMo'] + '"'
        else:
            swkyb += ',"sampleIdMo":"null"'
        if 'sampleClassMo' in kwargs:
            swkyb += ',"sampleClassMo":"' + str(kwargs['sampleClassMo']) + '"}'
        else:
            swkyb += ',"sampleClassMo":"null"}'
        short.m_bio_sam_bank = swkyb
        #  主要诊断
        mada = ""
        if 'mainDia' in kwargs:
            mada += '{"mainDia":"' + str(kwargs['mainDia']) + '"'
        else:
            mada += '{"mainDia":"null"'
        if 'mainDiaIllustrate' in kwargs:
            mada += ',"mainDiaIllustrate":"' + str(kwargs['mainDiaIllustrate']) + '"'
        else:
            mada += ',"mainDiaIllustrate":"null"'
        if 'DiaIllustrate' in kwargs:
            mada += ',"DiaIllustrate":"' + str(kwargs['DiaIllustrate']) + '"'
        else:
            mada += ',"DiaIllustrate":"null"'
        if 'peripheralityOther' in kwargs:
            mada += ',"peripheralityOther":"' + str(kwargs['peripheralityOther']) + '"'
        else:
            mada += ',"peripheralityOther":"null"'
        if 'partialityOther' in kwargs:
            mada += ',"partialityOther":"' + str(kwargs['partialityOther']) + '"}'
        else:
            mada += ',"partialityOther":"null"}'
        short.main_dia = mada


            # if kwargs['mainDiaIllustrate'] != "":
            #     # 输入的列表
            #     arrMainDia = [kwargs['mainDia'][0], kwargs['mainDia'][1], kwargs['mainDiaIllustrate']]
            #     # 使用字典推导式和 enumerate 函数生成字典
            #     objMainDia = {f"mainDia{i+1}": value for i, value in enumerate(arrMainDia)}
            #     # 将字典转换为 JSON 字符串
            #     short.main_dia = json.dumps(objMainDia, ensure_ascii=False)
            # elif kwargs['DiaIllustrate'] != "":
            #      arrMainDia = [kwargs['mainDia'][0], kwargs['DiaIllustrate']]
            #      objMainDia = {f"mainDia{i+1}": value for i, value in enumerate(arrMainDia)}
            #      short.main_dia = json.dumps(objMainDia, ensure_ascii=False)
            # else:
            #     arrMainDia = kwargs['mainDia']
            #     objMainDia = {f"mainDia{i+1}": value for i, value in enumerate(arrMainDia)}
            #     short.main_dia = json.dumps(objMainDia, ensure_ascii=False)
        #  次要诊断
        if 'secDia' in kwargs:
            short.sec_dia = kwargs['secDia']
        #  随访
        if 'followUp' in kwargs:
            short.follow_up = kwargs['followUp']

        # 染色体核型
        if 'speKar' in kwargs:
            short.spe_kar = kwargs['speKar']
        # SRY基因
        if 'SRY' in kwargs:
            short.SRY = kwargs['SRY']
        #  变异类型
        if 'mutKind' in kwargs:
            short.mut_kind = kwargs['mutKind']
        # 致病基因名称
        if 'genData' in kwargs:
            short.gen_mut_name = json.dumps(kwargs['genData'])
        # 变异来源
        if 'sourMut' in kwargs:
            short.sour_mut = kwargs['sourMut']
        # 核酸变异
        if 'baseMut' in kwargs:
            short.base_mut = kwargs['baseMut']
        # 氨基酸变异
        if 'amiAciMut' in kwargs:
            short.ami_aci_mut = kwargs['amiAciMut']
        #  ACTH刺激实验
        if 'acth8am' in kwargs:
            acthjf = '{"acth8am":"' + kwargs['acth8am'] + '"'
        else:
            acthjf = '{"acth8am":"null"'
        if 'acthData' in kwargs:
            acthjf += ',"acthData":"' + kwargs['acthData'] + '"'
        else:
            acthjf += ',"acthData":"null"'
        if 'acthTime' in kwargs:
            acthjf += ',"acthTime":"' + kwargs['acthTime'] + '"'
        else:
            acthjf += ',"acthTime":"null"'
        if 'hydroxy17a' in kwargs:
            acthjf += ',"hydroxy17a":"' + kwargs['hydroxy17a'] + '"'
        else:
            acthjf += ',"hydroxy17a":"null"'
        if 'hydroxy17aData' in kwargs:
            acthjf += ',"hydroxy17aData":"' + kwargs['hydroxy17aData'] + '"'
        else:
            acthjf += ',"hydroxy17aData":"null"'
        if 'hydroxy17aTime' in kwargs:
            acthjf += ',"hydroxy17aTime":"' + kwargs['hydroxy17aTime'] + '"'
        else:
            acthjf += ',"hydroxy17aTime":"null"'
        if 'DHEAs' in kwargs:
            acthjf += ',"DHEAs":"' + kwargs['DHEAs'] + '"'
        else:
            acthjf += ',"DHEAs":"null"'
        if 'DHEAsData' in kwargs:
            acthjf += ',"DHEAsData":"' + kwargs['DHEAsData'] + '"'
        else:
            acthjf += ',"DHEAsData":"null"'
        if 'DHEAsTime' in kwargs:
            acthjf += ',"DHEAsTime":"' + str(kwargs['DHEAsTime']) + '"}'
        else:
            acthjf += ',"DHEAsTime":"null"}'
        short.acth_jf = acthjf
        # 其他图片名称
        if 'otherImageNames' in kwargs:
            short.other_ima_name = kwargs['otherImageNames']
        short.c_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result = short
    except:
        result = False
    return result

# 修改或添加家族性矮小
def modifyorAddjzxShort(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            short = models.JzxShort.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            short = models.JzxShort()
        # 病人编号
        if 'userNum' in kwargs:
            short.user_num = kwargs['userNum']
        short.patient_id = casePk
        #  家族史
        if 'famHis' in kwargs:
            short.fam_his = json.dumps(kwargs['famHis'])
        #  运动发育落后
        if 'motDevBack' in kwargs:
            ydfy = '{"motDevBack":"' + kwargs['motDevBack'] + '",'
        else:
            ydfy = '{"motDevBack":"null",'
        if 'sport' in kwargs:
            ydfy += '"sport":"' + kwargs['sport'] + '"}'
        else:
            ydfy += '"sport":"null"}'
        short.mot_dev_back = ydfy
        #  语言发育落后
        if 'lanDevBack' in kwargs:
            yafy = '{"lanDevBack":"' + kwargs['lanDevBack'] + '",'
        else:
            yafy = '{"lanDevBack":"null",'
        if 'language' in kwargs:
            yafy += '"language":"' + kwargs['language'] + '"}'
        else:
            yafy += '"language":"null"}'
        short.lan_dev_back = yafy
        #  智力发育落后
        if 'intDevBack' in kwargs:
            zlfy = '{"intDevBack":"' + kwargs['intDevBack'] + '",'
        else:
            zlfy = '{"intDevBack":"null",'
        if 'intelligence' in kwargs:
            zlfy += '"intelligence":"' + kwargs['intelligence'] + '"}'
        else:
            zlfy += '"intelligence":"null"}'
        short.int_dev_back = zlfy
        #  听力异常
        if 'abnHear' in kwargs:
            tlyc = '{"abnHear":"' + kwargs['abnHear'] + '",'
        else:
            tlyc = '{"abnHear":"null",'
        if 'hear' in kwargs:
            tlyc += '"hear":"' + kwargs['hear'] + '"}'
        else:
            tlyc += '"hear":"null"}'
        short.abn_hear = tlyc
        #  反复感染史
        if 'recInfHis' in kwargs:
            ffgr = '{"recInfHis":"' + kwargs['recInfHis'] + '",'
        else:
            ffgr = '{"recInfHis":"null",'
        if 'infection' in kwargs:
            ffgr += '"infection":"' + kwargs['infection'] + '"}'
        else:
            ffgr += '"infection":"null"}'
        short.rec_inf_his = ffgr
        #  抽搐史
        if 'conHis' in kwargs:
            short.con_his = kwargs['conHis']
        #  其他
        if 'pastOther' in kwargs:
            short.past_other = kwargs['pastOther']
        #  病史
        if 'firVisTime' in kwargs:
            bs = '{"firVisTime":"' + kwargs['firVisTime'] + '",'
        else:
            bs = '{"firVisTime":"null",'
        if 'morbidAge' in kwargs:
            bs += '"morbidAge":"'+kwargs['morbidAge']+'",'
        else:
            bs += '"morbidAge":"null",'
        if 'chiefCom' in kwargs:
            bs += '"chiefCom":"'+kwargs['chiefCom']+'",'
        else:
            bs += '"chiefCom":"null",'
        if 'growRate' in kwargs:
            bs += '"growRate":"'+kwargs['growRate']+'",'
        else:
            bs += '"growRate":"null",'
        if 'rate' in kwargs:
            bs += '"rate":"'+kwargs['rate']+'",'
        else:
            bs += '"rate":"null",'
        if 'menarchyTime' in kwargs:
            bs += '"menarchyTime":"'+kwargs['menarchyTime']+'",'
        else:
            bs += '"menarchyTime":"null",'
        if 'menarchy' in kwargs:
            bs += '"menarchy":"'+kwargs['menarchy']+'"}'
        else:
            bs += '"menarchy":"null"}'
        short.med_his = bs
        #  体格检查
        if 'height' in kwargs:
            tgjc = '{"height":"' + kwargs['height'] + '",'
        else:
            tgjc = '{"height":"null",'
        if 'weight' in kwargs:
            tgjc += '"weight":"' + kwargs['weight'] + '",'
        else:
            tgjc += '"weight":"null",'
        if 'Bmi' in kwargs:
            tgjc += '"Bmi":"' + kwargs['Bmi'] + '",'
        else:
            tgjc += '"Bmi":"null",'
        if 'breastDev' in kwargs:
            tgjc += '"breastDev":"' + kwargs['breastDev'] + '",'
        else:
            tgjc += '"breastDev":"null",'
        if 'breastDevRight' in kwargs:
            tgjc += '"breastDevRight":"' + kwargs['breastDevRight'] + '",'
        else:
            tgjc += '"breastDevRight":"null",'
        if 'exGenitalia' in kwargs:
            tgjc += '"exGenitalia":"' + kwargs['exGenitalia'] + '",'
        else:
            tgjc += '"exGenitalia":"null",'
        if 'pubicHair' in kwargs:
            tgjc += '"pubicHair":"' + kwargs['pubicHair'] + '",'
        else:
            tgjc += '"pubicHair":"null",'
        if 'armLength' in kwargs:
            tgjc += '"armLength":"' + kwargs['armLength'] + '",'
        else:
            tgjc += '"armLength":"null",'
        if 'lowerMeasure' in kwargs:
            tgjc += '"lowerMeasure":"' + kwargs['lowerMeasure'] + '",'
        else:
            tgjc += '"lowerMeasure":"null",'
        if 'specialFace' in kwargs:
            tgjc += '"specialFace":"' + kwargs['specialFace'] + '",'
        else:
            tgjc += '"specialFace":"null",'
        if 'specialFaceDesc' in kwargs:
            tgjc += '"specialFaceDesc":"' + kwargs['specialFaceDesc'] + '",'
        else:
            tgjc += '"specialFaceDesc":"null",'
        if 'scoliosis' in kwargs:
            tgjc += '"scoliosis":"' + kwargs['scoliosis'] + '",'
        else:
            tgjc += '"scoliosis":"null",'
        if 'scoliosisDegree' in kwargs:
            tgjc += '"scoliosisDegree":"' + kwargs['scoliosisDegree'] + '",'
        else:
            tgjc += '"scoliosisDegree":"null",'
        if 'rashDescribe' in kwargs:
            tgjc += '"rashDescribe":"' + kwargs['rashDescribe'] + '",'
        else:
            tgjc += '"rashDescribe":"null",'
        if 'rash' in kwargs:
            tgjc += '"rash":"' + kwargs['rash'] + '"}'
        else:
            tgjc += '"rash":"null"}'
        short.phy_exa = tgjc
        #  实验室检查
        if 'LH' in kwargs:
            sysbg = '{"LH":"' + kwargs['LH'] + '",'
        else:
            sysbg = '{"LH":"null",'
        if 'FSH' in kwargs:
            sysbg += '"FSH":"' + kwargs['FSH'] + '",'
        else:
            sysbg += '"FSH":"null",'
        if 'E2' in kwargs:
            sysbg += '"E2":"' + kwargs['E2'] + '",'
        else:
            sysbg += '"E2":"null",'
        if 'T' in kwargs:
            sysbg += '"T":"' + kwargs['T'] + '",'
        else:
            sysbg += '"T":"null",'
        if 'PRL' in kwargs:
            sysbg += '"PRL":"' + kwargs['PRL'] + '",'
        else:
            sysbg += '"PRL":"null",'
        if 'IGF' in kwargs:
            sysbg += '"IGF":"' + kwargs['IGF'] + '",'
        else:
            sysbg += '"IGF":"null",'
        if 'IGFBP3' in kwargs:
            sysbg += '"IGFBP3":"' + kwargs['IGFBP3'] + '",'
        else:
            sysbg += '"IGFBP3":"null",'
        if 'thyroid' in kwargs:
            sysbg += '"thyroid":"' + kwargs['thyroid'] + '",'
        else:
            sysbg += '"thyroid":"null",'
        if 'thyroidDescribe' in kwargs:
            sysbg += '"thyroidDescribe":"' + kwargs['thyroidDescribe'] + '",'
        else:
            sysbg += '"thyroidDescribe":"null",'
        if 'ACTH' in kwargs:
            sysbg += '"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            sysbg += '"ACTH":"null",'
        if 'cortisol' in kwargs:
            sysbg += '"cortisol":"' + kwargs['cortisol'] + '",'
        else:
            sysbg += '"cortisol":"null",'
        if 'DHEAS' in kwargs:
            sysbg += '"DHEAS":"' + kwargs['DHEAS'] + '",'
        else:
            sysbg += '"DHEAS":"null",'
        if 'OHP' in kwargs:
            sysbg += '"OHP":"' + kwargs['OHP'] + '",'
        else:
            sysbg += '"OHP":"null",'
        if 'blood' in kwargs:
            sysbg += '"blood":"' + kwargs['blood'] + '",'
        else:
            sysbg += '"blood":"null",'
        if 'bloodDescribe' in kwargs:
            sysbg += '"bloodDescribe":"' + kwargs['bloodDescribe'] + '",'
        else:
            sysbg += '"bloodDescribe":"null",'
        if 'urinalysis' in kwargs:
            sysbg += '"urinalysis":"' + kwargs['urinalysis'] + '",'
        else:
            sysbg += '"urinalysis":"null",'
        if 'urinalysisDescribe' in kwargs:
            sysbg += '"urinalysisDescribe":"' + kwargs['urinalysisDescribe'] + '",'
        else:
            sysbg += '"urinalysisDescribe":"null",'
        if 'LAKLGE' in kwargs:
            sysbg += '"LAKLGE":"' + kwargs['LAKLGE'] + '",'
        else:
            sysbg += '"LAKLGE":"null",'
        if 'laklgeDescribe' in kwargs:
            sysbg += '"laklgeDescribe":"' + kwargs['laklgeDescribe'] + '",'
        else:
            sysbg += '"laklgeDescribe":"null",'
        if 'HBs' in kwargs:
            sysbg += '"HBs":"' + kwargs['HBs'] + '",'
        else:
            sysbg += '"HBs":"null",'
        if 'HBsDescribe' in kwargs:
            sysbg += '"HBsDescribe":"' + kwargs['HBsDescribe'] + '",'
        else:
            sysbg += '"HBsDescribe":"null",'
        if 'glyHemA' in kwargs:
            sysbg += '"glyHemA":"' + kwargs['glyHemA'] + '",'
        else:
            sysbg += '"glyHemA":"null",'
        if 'glyHemATime' in kwargs:
            sysbg += '"glyHemATime":"' + kwargs['glyHemATime'] + '",'
        else:
            sysbg += '"glyHemATime":"null",'
        if 'fasBloodGlu' in kwargs:
            sysbg += '"fasBloodGlu":"' + kwargs['fasBloodGlu'] + '",'
        else:
            sysbg += '"fasBloodGlu":"null",'
        if 'fasBloodGluTime' in kwargs:
            sysbg += '"fasBloodGluTime":"' + kwargs['fasBloodGluTime'] + '",'
        else:
            sysbg += '"fasBloodGluTime":"null",'
        if 'fasInsulin' in kwargs:
            sysbg += '"fasInsulin":"' + kwargs['fasInsulin'] + '",'
        else:
            sysbg += '"fasInsulin":"null",'
        if 'fasInsulinTime' in kwargs:
            sysbg += '"fasInsulinTime":"' + kwargs['fasInsulinTime'] + '",'
        else:
            sysbg += '"fasInsulinTime":"null",'
        if 'glyHem' in kwargs:
            sysbg += '"glyHem":"' + kwargs['glyHem'] + '",'
        else:
            sysbg += '"glyHem":"null",'
        if 'glyHemTime' in kwargs:
            sysbg += '"glyHemTime":"' + kwargs['glyHemTime'] + '",'
        else:
            sysbg += '"glyHemTime":"null",'
        if 'LHFSHTime' in kwargs:
            sysbg += '"LHFSHTime":"' + kwargs['LHFSHTime'] + '",'
        else:
            sysbg += '"LHFSHTime":"null",'
        if 'E2Time' in kwargs:
            sysbg += '"E2Time":"' + kwargs['E2Time'] + '",'
        else:
            sysbg += '"E2Time":"null",'
        if 'TTime' in kwargs:
            sysbg += '"TTime":"' + kwargs['TTime'] + '",'
        else:
            sysbg += '"TTime":"null",'
        if 'PRLTime' in kwargs:
            sysbg += '"PRLTime":"' + kwargs['PRLTime'] + '",'
        else:
            sysbg += '"PRLTime":"null",'
        if 'IGFBPTime' in kwargs:
            sysbg += '"IGFBPTime":"' + kwargs['IGFBPTime'] + '",'
        else:
            sysbg += '"IGFBPTime":"null",'
        if 'thyroidTime' in kwargs:
            sysbg += '"thyroidTime":"' + kwargs['thyroidTime'] + '",'
        else:
            sysbg += '"thyroidTime":"null",'
        if 'ACTHTime' in kwargs:
            sysbg += '"ACTHTime":"' + kwargs['ACTHTime'] + '",'
        else:
            sysbg += '"ACTHTime":"null",'
        if 'cortisolTime' in kwargs:
            sysbg += '"cortisolTime":"' + kwargs['cortisolTime'] + '",'
        else:
            sysbg += '"cortisolTime":"null",'
        if 'DHEATime' in kwargs:
            sysbg += '"DHEATime":"' + kwargs['DHEATime'] + '",'
        else:
            sysbg += '"DHEATime":"null",'
        if 'OHPTime' in kwargs:
            sysbg += '"OHPTime":"' + kwargs['OHPTime'] + '",'
        else:
            sysbg += '"OHPTime":"null",'
        if 'bloodTime' in kwargs:
            sysbg += '"bloodTime":"' + kwargs['bloodTime'] + '",'
        else:
            sysbg += '"bloodTime":"null",'
        if 'urinalysisTime' in kwargs:
            sysbg += '"urinalysisTime":"' + kwargs['urinalysisTime'] + '",'
        else:
            sysbg += '"urinalysisTime":"null",'
        if 'LAKLGETime' in kwargs:
            sysbg += '"LAKLGETime":"' + kwargs['LAKLGETime'] + '",'
        else:
            sysbg += '"LAKLGETime":"null",'
        if 'HBsTime' in kwargs:
            sysbg += '"HBsTime":"' + kwargs['HBsTime'] + '",'
        else:
            sysbg += '"HBsTime":"null",'
        if 'ghTime' in kwargs:
            sysbg += '"ghTime":"' + kwargs['ghTime'] + '",'
        else:
            sysbg += '"ghTime":"null",'
        if 'gh' in kwargs:
            sysbg += '"gh":"' + kwargs['gh'] + '"}'
        else:
            sysbg += '"gh":"null"}'
        short.lab_exa = sysbg
        #  心电图
        if 'electdiogram' in kwargs:
            short.electr = kwargs['electdiogram']
        #  性腺B超
        if 'uterusOne' in kwargs:
            xxbc = '{"uterusOne":"' + kwargs['uterusOne'] + '",'
        else:
            xxbc = '{"uterusOne":"null",'
        if 'uterusTwo' in kwargs:
            xxbc += '"uterusTwo":"' + kwargs['uterusTwo'] + '",'
        else:
            xxbc += '"uterusTwo":"null",'
        if 'uterusThr' in kwargs:
            xxbc += '"uterusThr":"' + kwargs['uterusThr'] + '",'
        else:
            xxbc += '"uterusThr":"null",'
        if 'cervixLong' in kwargs:
            xxbc += '"cervixLong":"' + kwargs['cervixLong'] + '",'
        else:
            xxbc += '"cervixLong":"null",'
        if 'intima' in kwargs:
            xxbc += '"intima":"' + kwargs['intima'] + '",'
        else:
            xxbc += '"intima":"null",'
        if 'ovaLeftOne' in kwargs:
            xxbc += '"ovaLeftOne":"' + kwargs['ovaLeftOne'] + '",'
        else:
            xxbc += '"ovaLeftOne":"null",'
        if 'ovaLeftTwo' in kwargs:
            xxbc += '"ovaLeftTwo":"' + kwargs['ovaLeftTwo'] + '",'
        else:
            xxbc += '"ovaLeftTwo":"null",'
        if 'ovaLeftThr' in kwargs:
            xxbc += '"ovaLeftThr":"' + kwargs['ovaLeftThr'] + '",'
        else:
            xxbc += '"ovaLeftThr":"null",'
        if 'ovaRightOne' in kwargs:
            xxbc += '"ovaRightOne":"' + kwargs['ovaRightOne'] + '",'
        else:
            xxbc += '"ovaRightOne":"null",'
        if 'ovaRightTwo' in kwargs:
            xxbc += '"ovaRightTwo":"' + kwargs['ovaRightTwo'] + '",'
        else:
            xxbc += '"ovaRightTwo":"null",'
        if 'ovaRightThr' in kwargs:
            xxbc += '"ovaRightThr":"' + kwargs['ovaRightThr'] + '",'
        else:
            xxbc += '"ovaRightThr":"null",'
        if 'follDiameter' in kwargs:
            xxbc += '"follDiameter":"' + kwargs['follDiameter'] + '",'
        else:
            xxbc += '"follDiameter":"null",'
        if 'isCyst' in kwargs:
            xxbc += '"isCyst":"' + kwargs['isCyst'] + '",'
        else:
            xxbc += '"isCyst":"null",'
        if 'cyst' in kwargs:
            xxbc += '"cyst":"' + kwargs['cyst'] + '",'
        else:
            xxbc += '"cyst":"null",'
        if 'cystOne' in kwargs:
            xxbc += '"cystOne":"' + kwargs['cystOne'] + '",'
        else:
            xxbc += '"cystOne":"null",'
        if 'cystTwo' in kwargs:
            xxbc += '"cystTwo":"' + kwargs['cystTwo'] + '",'
        else:
            xxbc += '"cystTwo":"null",'
        if 'cystThr' in kwargs:
            xxbc += '"cystThr":"' + kwargs['cystThr'] + '",'
        else:
            xxbc += '"cystThr":"null",'
        if 'cystDescribe' in kwargs:
            xxbc += '"cystDescribe":"' + kwargs['cystDescribe'] + '",'
        else:
            xxbc += '"cystDescribe":"null",'
        if 'testisLeftOne' in kwargs:
            xxbc += '"testisLeftOne":"' + kwargs['testisLeftOne'] + '",'
        else:
            xxbc += '"testisLeftOne":"null",'
        if 'testisLeftTwo' in kwargs:
            xxbc += '"testisLeftTwo":"' + kwargs['testisLeftTwo'] + '",'
        else:
            xxbc += '"testisLeftTwo":"null",'
        if 'testisLeftThr' in kwargs:
            xxbc += '"testisLeftThr":"' + kwargs['testisLeftThr'] + '",'
        else:
            xxbc += '"testisLeftThr":"null",'
        if 'testisLeftLon' in kwargs:
            xxbc += '"testisLeftLon":"' + kwargs['testisLeftLon'] + '",'
        else:
            xxbc += '"testisLeftLon":"null",'
        if 'testisRightOne' in kwargs:
            xxbc += '"testisRightOne":"' + kwargs['testisRightOne'] + '",'
        else:
            xxbc += '"testisRightOne":"null",'
        if 'testisRightTwo' in kwargs:
            xxbc += '"testisRightTwo":"' + kwargs['testisRightTwo'] + '",'
        else:
            xxbc += '"testisRightTwo":"null",'
        if 'testisRightThr' in kwargs:
            xxbc += '"testisRightThr":"' + kwargs['testisRightThr'] + '",'
        else:
            xxbc += '"testisRightThr":"null",'
        if 'MRI' in kwargs:
            xxbc += '"MRI":"' + kwargs['MRI'] + '",'
        else:
            xxbc += '"MRI":"null",'
        if 'ThyroidLBGradation' in kwargs:
            xxbc += '"ThyroidLBGradation":"' + kwargs['ThyroidLBGradation'] + '",'
        else:
            xxbc += '"ThyroidLBGradation":"null",'
        if 'ThyroidLBSize' in kwargs:
            xxbc += '"ThyroidLBSize":"' + kwargs['ThyroidLBSize'] + '",'
        else:
            xxbc += '"ThyroidLBSize":"null",'
        if 'ThyroidLBLesions' in kwargs:
            xxbc += '"ThyroidLBLesions":"' + kwargs['ThyroidLBLesions'] + '",'
        else:
            xxbc += '"ThyroidLBLesions":"null",'
        if 'ThyroidLBOther' in kwargs:
            xxbc += '"ThyroidLBOther":"' + kwargs['ThyroidLBOther'] + '",'
        else:
            xxbc += '"ThyroidLBOther":"null",'
        if 'ThyroidLB' in kwargs:
            xxbc += '"ThyroidLB":"' + kwargs['ThyroidLB'] + '",'
        else:
            xxbc += '"ThyroidLB":"null",'
        if 'ThyroidRB' in kwargs:
            xxbc += '"ThyroidRB":"' + kwargs['ThyroidRB'] + '",'
        else:
            xxbc += '"ThyroidRB":"null",'
        if 'ThyroidRBGradation' in kwargs:
            xxbc += '"ThyroidRBGradation":"' + kwargs['ThyroidRBGradation'] + '",'
        else:
            xxbc += '"ThyroidRBGradation":"null",'
        if 'ThyroidRBSize' in kwargs:
            xxbc += '"ThyroidRBSize":"' + kwargs['ThyroidRBSize'] + '",'
        else:
            xxbc += '"ThyroidRBSize":"null",'
        if 'ThyroidRBLesions' in kwargs:
            xxbc += '"ThyroidRBLesions":"' + kwargs['ThyroidRBLesions'] + '",'
        else:
            xxbc += '"ThyroidRBLesions":"null",'
        if 'ThyroidRBOther' in kwargs:
            xxbc += '"ThyroidRBOther":"' + kwargs['ThyroidRBOther'] + '",'
        else:
            xxbc += '"ThyroidRBOther":"null",'
        if 'mriDescribe' in kwargs:
            xxbc += '"mriDescribe":"' + kwargs['mriDescribe'] + '",'
        else:
            xxbc += '"mriDescribe":"null",'
        if 'testisRightLon' in kwargs:
            xxbc += '"testisRightLon":"' + kwargs['testisRightLon'] + '"}'
        else:
            xxbc += '"testisRightLon":"null"}'
        short.gon_B_ult = xxbc
        #  诊疗方案
        if 'diaPlan' in kwargs:
            zlfa = '{"diaPlan":"' + kwargs['diaPlan'] + '",'
        else:
            zlfa = '{"diaPlan":"null",'
        if 'rhGH' in kwargs:
            zlfa += '"rhGH":"' + kwargs['rhGH'] + '",'
        else:
            zlfa += '"rhGH":"null",'
        if 'GnRHa' in kwargs:
            zlfa += '"GnRHa":"' + kwargs['GnRHa'] + '",'
        else:
            zlfa += '"GnRHa":"null",'
        if 'GnRHadose' in kwargs:
            zlfa += '"GnRHadose":"' + kwargs['GnRHadose'] + '",'
        else:
            zlfa += '"GnRHadose":"null",'
        if 'genData' in kwargs:
            zlfa += '"genData":"' + json.dumps(kwargs['genData']) + '",'
        else:
            zlfa += '"genData":"null",'
        if 'rhGHdoseKG' in kwargs:
            zlfa += '"rhGHdoseKG":"' +  kwargs['rhGHdoseKG'] + '",'
        else:
            zlfa += '"rhGHdoseKG":"null",'
        if 'PEGrhGHdose' in kwargs:
            zlfa += '"PEGrhGHdose":"' +  kwargs['PEGrhGHdose'] + '",'
        else:
            zlfa += '"PEGrhGHdose":"null",'
        if 'rhCustomizationDiaPlan' in kwargs:
            zlfa += '"rhCustomizationDiaPlan":"' +  kwargs['rhCustomizationDiaPlan'] + '",'
        else:
            zlfa += '"rhCustomizationDiaPlan":"null",'
        if 'rhCustomizationPrompt' in kwargs:
            zlfa += '"rhCustomizationPrompt":"' +  kwargs['rhCustomizationPrompt'] + '",'
        else:
            zlfa += '"rhCustomizationPrompt":"null",'
        if 'PEGrhCustomizationPrompt' in kwargs:
            zlfa += '"PEGrhCustomizationPrompt":"' +  kwargs['PEGrhCustomizationPrompt'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPrompt":"null",'
        if 'rhCustomizationPromptKG' in kwargs:
            zlfa += '"rhCustomizationPromptKG":"' +  kwargs['rhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"rhCustomizationPromptKG":"null",'
        if 'PEGrhCustomizationPromptKG' in kwargs:
            zlfa += '"PEGrhCustomizationPromptKG":"' +  kwargs['PEGrhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPromptKG":"null",'
        if 'PEGrhGHdoseKG' in kwargs:
            zlfa += '"PEGrhGHdoseKG":"' +  kwargs['PEGrhGHdoseKG'] + '",'
        else:
            zlfa += '"PEGrhGHdoseKG":"null",'
        if 'planData' in kwargs:
            zlfa += '"planData":"' +  json.dumps(kwargs['planData']) + '",'
        else:
            zlfa += '"planData":"null",'
        if 'otherMedicine' in kwargs:
            zlfa += '"otherMedicine":"' +  json.dumps(kwargs['otherMedicine']) + '",'
        else:
            zlfa += '"otherMedicine":"null",'
        if 'rhGHdose' in kwargs:
            zlfa += '"rhGHdose":"' + kwargs['rhGHdose'] + '"}'
        else:
            zlfa += '"rhGHdose":"null"}'
        short.dia_trea_plan = zlfa
        #  生物样本库
        if 'bioBank' in kwargs:
            swkyb = '{"bioBank":"' + kwargs['bioBank'] + '",'
        else:
            swkyb = '{"bioBank":"null",'
        if 'sampleId' in kwargs:
            swkyb += '"sampleId":"' + kwargs['sampleId'] + '",'
        else:
            swkyb += '"sampleId":"null",'
        if 'sampleClass' in kwargs:
            swkyb += '"sampleClass":"' + str(kwargs['sampleClass']) + '"}'
        else:
            swkyb += '"sampleClass":"null"}'
        short.bio_sam_bank = swkyb
        #  父亲生物样本库
        if 'bioBankFa' in kwargs:
            swkyb = '{"bioBankFa":"' + kwargs['bioBankFa'] + '"'
        else:
            swkyb = '{"bioBankFa":"null"'
        if 'sampleIdFa' in kwargs:
            swkyb += ',"sampleIdFa":"' + kwargs['sampleIdFa'] + '"'
        else:
            swkyb += ',"sampleIdFa":"null"'
        if 'sampleClassFa' in kwargs:
            swkyb += ',"sampleClassFa":"' + str(kwargs['sampleClassFa']) + '"}'
        else:
            swkyb += ',"sampleClassFa":"null"}'
        short.f_bio_sam_bank = swkyb
        #  母亲生物样本库
        if 'bioBankMo' in kwargs:
            swkyb = '{"bioBankMo":"' + kwargs['bioBankMo'] + '"'
        else:
            swkyb = '{"bioBankMo":"null"'
        if 'sampleIdMo' in kwargs:
            swkyb += ',"sampleIdMo":"' + kwargs['sampleIdMo'] + '"'
        else:
            swkyb += ',"sampleIdMo":"null"'
        if 'sampleClassMo' in kwargs:
            swkyb += ',"sampleClassMo":"' + str(kwargs['sampleClassMo']) + '"}'
        else:
            swkyb += ',"sampleClassMo":"null"}'
        short.m_bio_sam_bank = swkyb
        #  主要诊断
        mada = ""
        if 'mainDia' in kwargs:
            mada += '{"mainDia":"' + str(kwargs['mainDia']) + '"'
        else:
            mada += '{"mainDia":"null"'
        if 'mainDiaIllustrate' in kwargs:
            mada += ',"mainDiaIllustrate":"' + str(kwargs['mainDiaIllustrate']) + '"'
        else:
            mada += ',"mainDiaIllustrate":"null"'
        if 'DiaIllustrate' in kwargs:
            mada += ',"DiaIllustrate":"' + str(kwargs['DiaIllustrate']) + '"'
        else:
            mada += ',"DiaIllustrate":"null"'
        if 'peripheralityOther' in kwargs:
            mada += ',"peripheralityOther":"' + str(kwargs['peripheralityOther']) + '"'
        else:
            mada += ',"peripheralityOther":"null"'
        if 'partialityOther' in kwargs:
            mada += ',"partialityOther":"' + str(kwargs['partialityOther']) + '"}'
        else:
            mada += ',"partialityOther":"null"}'
        short.main_dia = mada
        #  次要诊断
        if 'secDia' in kwargs:
            short.sec_dia = kwargs['secDia']
        #  随访
        if 'followUp' in kwargs:
            short.follow_up = kwargs['followUp']

        # 染色体核型
        if 'speKar' in kwargs:
            short.spe_kar = kwargs['speKar']
        # SRY基因
        if 'SRY' in kwargs:
            short.SRY = kwargs['SRY']
        #  变异类型
        if 'mutKind' in kwargs:
            short.mut_kind = kwargs['mutKind']
        # 致病基因名称
        if 'genData' in kwargs:
            short.gen_mut_name = json.dumps(kwargs['genData'])
        # 变异来源
        if 'sourMut' in kwargs:
            short.sour_mut = kwargs['sourMut']
        # 核酸变异
        if 'baseMut' in kwargs:
            short.base_mut = kwargs['baseMut']
        # 氨基酸变异
        if 'amiAciMut' in kwargs:
            short.ami_aci_mut = kwargs['amiAciMut']
        #  ACTH刺激实验
        if 'acth8am' in kwargs:
            acthjf = '{"acth8am":"' + kwargs['acth8am'] + '"'
        else:
            acthjf = '{"acth8am":"null"'
        if 'acthData' in kwargs:
            acthjf += ',"acthData":"' + kwargs['acthData'] + '"'
        else:
            acthjf += ',"acthData":"null"'
        if 'acthTime' in kwargs:
            acthjf += ',"acthTime":"' + kwargs['acthTime'] + '"'
        else:
            acthjf += ',"acthTime":"null"'
        if 'hydroxy17a' in kwargs:
            acthjf += ',"hydroxy17a":"' + kwargs['hydroxy17a'] + '"'
        else:
            acthjf += ',"hydroxy17a":"null"'
        if 'hydroxy17aData' in kwargs:
            acthjf += ',"hydroxy17aData":"' + kwargs['hydroxy17aData'] + '"'
        else:
            acthjf += ',"hydroxy17aData":"null"'
        if 'hydroxy17aTime' in kwargs:
            acthjf += ',"hydroxy17aTime":"' + kwargs['hydroxy17aTime'] + '"'
        else:
            acthjf += ',"hydroxy17aTime":"null"'
        if 'DHEAs' in kwargs:
            acthjf += ',"DHEAs":"' + kwargs['DHEAs'] + '"'
        else:
            acthjf += ',"DHEAs":"null"'
        if 'DHEAsData' in kwargs:
            acthjf += ',"DHEAsData":"' + kwargs['DHEAsData'] + '"'
        else:
            acthjf += ',"DHEAsData":"null"'
        if 'DHEAsTime' in kwargs:
            acthjf += ',"DHEAsTime":"' + str(kwargs['DHEAsTime']) + '"}'
        else:
            acthjf += ',"DHEAsTime":"null"}'
        short.acth_jf = acthjf
        short.c_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 其他图片名称
        if 'otherImageNames' in kwargs:
            short.other_ima_name = kwargs['otherImageNames']
        result = short
    except:
        result = False
    return result

# 修改SGA
def modifyorAddSGA(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            short = models.SGA.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            short = models.SGA()
        # 病人编号
        if 'userNum' in kwargs:
            short.user_num = kwargs['userNum']
        short.patient_id = casePk
        # 母亲孕期疾病
        if 'isPregnancySickness' in kwargs:
            short.mot_pre_dis = kwargs['isPregnancySickness']
        # 母亲孕期疾病其他描述
        if 'PregnancySickness' in kwargs:
            short.mot_pre_dis_ms = kwargs['PregnancySickness']
        # 是否多胎
        if 'isMultiplePregnancies' in kwargs:
            short.is_mul_bir = kwargs['isMultiplePregnancies']
        # 多胎描述
        if 'MultiplePregnancies' in kwargs:
            short.mul_bir_ms = kwargs['MultiplePregnancies']
        # 胎次
        if 'isParityG' in kwargs and kwargs['isParityG'] != "":
            short.parity = kwargs['isParityG']
        # 产次
        if 'isParityP' in kwargs and kwargs['isParityP'] != "":
            short.pronum = kwargs['isParityP']
        #  家族史
        if 'famHis' in kwargs:
            short.fam_his = json.dumps(kwargs['famHis'])
        #  运动发育落后
        if 'motDevBack' in kwargs:
            ydfy = '{"motDevBack":"' + kwargs['motDevBack'] + '",'
        else:
            ydfy = '{"motDevBack":"null",'
        if 'sport' in kwargs:
            ydfy += '"sport":"' + kwargs['sport'] + '"}'
        else:
            ydfy += '"sport":"null"}'
        short.mot_dev_back = ydfy
        #  语言发育落后
        if 'lanDevBack' in kwargs:
            yafy = '{"lanDevBack":"' + kwargs['lanDevBack'] + '",'
        else:
            yafy = '{"lanDevBack":"null",'
        if 'language' in kwargs:
            yafy += '"language":"' + kwargs['language'] + '"}'
        else:
            yafy += '"language":"null"}'
        short.lan_dev_back = yafy
        #  智力发育落后
        if 'intDevBack' in kwargs:
            zlfy = '{"intDevBack":"' + kwargs['intDevBack'] + '",'
        else:
            zlfy = '{"intDevBack":"null",'
        if 'intelligence' in kwargs:
            zlfy += '"intelligence":"' + kwargs['intelligence'] + '"}'
        else:
            zlfy += '"intelligence":"null"}'
        short.int_dev_back = zlfy
        #  听力异常
        if 'abnHear' in kwargs:
            tlyc = '{"abnHear":"' + kwargs['abnHear'] + '",'
        else:
            tlyc = '{"abnHear":"null",'
        if 'hear' in kwargs:
            tlyc += '"hear":"' + kwargs['hear'] + '"}'
        else:
            tlyc += '"hear":"null"}'
        short.abn_hear = tlyc
        #  反复感染史
        if 'recInfHis' in kwargs:
            ffgr = '{"recInfHis":"' + kwargs['recInfHis'] + '",'
        else:
            ffgr = '{"recInfHis":"null",'
        if 'infection' in kwargs:
            ffgr += '"infection":"' + kwargs['infection'] + '"}'
        else:
            ffgr += '"infection":"null"}'
        short.rec_inf_his = ffgr
        #  抽搐史
        if 'conHis' in kwargs:
            short.con_his = kwargs['conHis']
        #  其他
        if 'pastOther' in kwargs:
            short.past_other = kwargs['pastOther']
        #  病史
        if 'firVisTime' in kwargs:
            bs = '{"firVisTime":"' + kwargs['firVisTime'] + '",'
        else:
            bs = '{"firVisTime":"null",'
        if 'morbidAge' in kwargs:
            bs += '"morbidAge":"'+kwargs['morbidAge']+'",'
        else:
            bs += '"morbidAge":"null",'
        if 'chiefCom' in kwargs:
            bs += '"chiefCom":"'+kwargs['chiefCom']+'",'
        else:
            bs += '"chiefCom":"null",'
        if 'growRate' in kwargs:
            bs += '"growRate":"'+kwargs['growRate']+'",'
        else:
            bs += '"growRate":"null",'
        if 'rate' in kwargs:
            bs += '"rate":"'+kwargs['rate']+'",'
        else:
            bs += '"rate":"null",'
        if 'menarchyTime' in kwargs:
            bs += '"menarchyTime":"'+kwargs['menarchyTime']+'",'
        else:
            bs += '"menarchyTime":"null",'
        if 'menarchy' in kwargs:
            bs += '"menarchy":"'+kwargs['menarchy']+'"}'
        else:
            bs += '"menarchy":"null"}'
        short.med_his = bs
        #  体格检查
        if 'height' in kwargs:
            tgjc = '{"height":"' + kwargs['height'] + '",'
        else:
            tgjc = '{"height":"null",'
        if 'weight' in kwargs:
            tgjc += '"weight":"' + kwargs['weight'] + '",'
        else:
            tgjc += '"weight":"null",'
        if 'Bmi' in kwargs:
            tgjc += '"Bmi":"' + kwargs['Bmi'] + '",'
        else:
            tgjc += '"Bmi":"null",'
        if 'breastDev' in kwargs:
            tgjc += '"breastDev":"' + kwargs['breastDev'] + '",'
        else:
            tgjc += '"breastDev":"null",'
        if 'breastDevRight' in kwargs:
            tgjc += '"breastDevRight":"' + kwargs['breastDevRight'] + '",'
        else:
            tgjc += '"breastDevRight":"null",'
        if 'exGenitalia' in kwargs:
            tgjc += '"exGenitalia":"' + kwargs['exGenitalia'] + '",'
        else:
            tgjc += '"exGenitalia":"null",'
        if 'pubicHair' in kwargs:
            tgjc += '"pubicHair":"' + kwargs['pubicHair'] + '",'
        else:
            tgjc += '"pubicHair":"null",'
        if 'armLength' in kwargs:
            tgjc += '"armLength":"' + kwargs['armLength'] + '",'
        else:
            tgjc += '"armLength":"null",'
        if 'specialFace' in kwargs:
            tgjc += '"specialFace":"' + kwargs['specialFace'] + '",'
        else:
            tgjc += '"specialFace":"null",'
        if 'specialFaceDesc' in kwargs:
            tgjc += '"specialFaceDesc":"' + kwargs['specialFaceDesc'] + '",'
        else:
            tgjc += '"specialFaceDesc":"null",'
        if 'scoliosis' in kwargs:
            tgjc += '"scoliosis":"' + kwargs['scoliosis'] + '",'
        else:
            tgjc += '"scoliosis":"null",'
        if 'scoliosisDegree' in kwargs:
            tgjc += '"scoliosisDegree":"' + kwargs['scoliosisDegree'] + '",'
        else:
            tgjc += '"scoliosisDegree":"null",'
        if 'rashDescribe' in kwargs:
            tgjc += '"rashDescribe":"' + kwargs['rashDescribe'] + '",'
        else:
            tgjc += '"rashDescribe":"null",'
        if 'rash' in kwargs:
            tgjc += '"rash":"' + kwargs['rash'] + '"}'
        else:
            tgjc += '"rash":"null"}'
        short.phy_exa = tgjc
        #  实验室检查
        if 'LH' in kwargs:
            sysbg = '{"LH":"' + kwargs['LH'] + '",'
        else:
            sysbg = '{"LH":"null",'
        if 'FSH' in kwargs:
            sysbg += '"FSH":"' + kwargs['FSH'] + '",'
        else:
            sysbg += '"FSH":"null",'
        if 'E2' in kwargs:
            sysbg += '"E2":"' + kwargs['E2'] + '",'
        else:
            sysbg += '"E2":"null",'
        if 'T' in kwargs:
            sysbg += '"T":"' + kwargs['T'] + '",'
        else:
            sysbg += '"T":"null",'
        if 'PRL' in kwargs:
            sysbg += '"PRL":"' + kwargs['PRL'] + '",'
        else:
            sysbg += '"PRL":"null",'
        if 'IGF' in kwargs:
            sysbg += '"IGF":"' + kwargs['IGF'] + '",'
        else:
            sysbg += '"IGF":"null",'
        if 'IGFBP3' in kwargs:
            sysbg += '"IGFBP3":"' + kwargs['IGFBP3'] + '",'
        else:
            sysbg += '"IGFBP3":"null",'
        if 'thyroid' in kwargs:
            sysbg += '"thyroid":"' + kwargs['thyroid'] + '",'
        else:
            sysbg += '"thyroid":"null",'
        if 'thyroidDescribe' in kwargs:
            sysbg += '"thyroidDescribe":"' + kwargs['thyroidDescribe'] + '",'
        else:
            sysbg += '"thyroidDescribe":"null",'
        if 'ACTH' in kwargs:
            sysbg += '"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            sysbg += '"ACTH":"null",'
        if 'cortisol' in kwargs:
            sysbg += '"cortisol":"' + kwargs['cortisol'] + '",'
        else:
            sysbg += '"cortisol":"null",'
        if 'DHEAS' in kwargs:
            sysbg += '"DHEAS":"' + kwargs['DHEAS'] + '",'
        else:
            sysbg += '"DHEAS":"null",'
        if 'OHP' in kwargs:
            sysbg += '"OHP":"' + kwargs['OHP'] + '",'
        else:
            sysbg += '"OHP":"null",'
        if 'blood' in kwargs:
            sysbg += '"blood":"' + kwargs['blood'] + '",'
        else:
            sysbg += '"blood":"null",'
        if 'bloodDescribe' in kwargs:
            sysbg += '"bloodDescribe":"' + kwargs['bloodDescribe'] + '",'
        else:
            sysbg += '"bloodDescribe":"null",'
        if 'urinalysis' in kwargs:
            sysbg += '"urinalysis":"' + kwargs['urinalysis'] + '",'
        else:
            sysbg += '"urinalysis":"null",'
        if 'urinalysisDescribe' in kwargs:
            sysbg += '"urinalysisDescribe":"' + kwargs['urinalysisDescribe'] + '",'
        else:
            sysbg += '"urinalysisDescribe":"null",'
        if 'LAKLGE' in kwargs:
            sysbg += '"LAKLGE":"' + kwargs['LAKLGE'] + '",'
        else:
            sysbg += '"LAKLGE":"null",'
        if 'laklgeDescribe' in kwargs:
            sysbg += '"laklgeDescribe":"' + kwargs['laklgeDescribe'] + '",'
        else:
            sysbg += '"laklgeDescribe":"null",'
        if 'HBs' in kwargs:
            sysbg += '"HBs":"' + kwargs['HBs'] + '",'
        else:
            sysbg += '"HBs":"null",'
        if 'HBsDescribe' in kwargs:
            sysbg += '"HBsDescribe":"' + kwargs['HBsDescribe'] + '",'
        else:
            sysbg += '"HBsDescribe":"null",'
        if 'fasBloodGlu' in kwargs:
            sysbg += '"fasBloodGlu":"' + kwargs['fasBloodGlu'] + '",'
        else:
            sysbg += '"fasBloodGlu":"null",'
        if 'fasBloodGluTime' in kwargs:
            sysbg += '"fasBloodGluTime":"' + kwargs['fasBloodGluTime'] + '",'
        else:
            sysbg += '"fasBloodGluTime":"null",'
        if 'fasInsulin' in kwargs:
            sysbg += '"fasInsulin":"' + kwargs['fasInsulin'] + '",'
        else:
            sysbg += '"fasInsulin":"null",'
        if 'fasInsulinTime' in kwargs:
            sysbg += '"fasInsulinTime":"' + kwargs['fasInsulinTime'] + '",'
        else:
            sysbg += '"fasInsulinTime":"null",'
        if 'glyHem' in kwargs:
            sysbg += '"glyHem":"' + kwargs['glyHem'] + '",'
        else:
            sysbg += '"glyHem":"null",'
        if 'glyHemTime' in kwargs:
            sysbg += '"glyHemTime":"' + kwargs['glyHemTime'] + '",'
        else:
            sysbg += '"glyHemTime":"null",'
        if 'LHFSHTime' in kwargs:
            sysbg += '"LHFSHTime":"' + kwargs['LHFSHTime'] + '",'
        else:
            sysbg += '"LHFSHTime":"null",'
        if 'E2Time' in kwargs:
            sysbg += '"E2Time":"' + kwargs['E2Time'] + '",'
        else:
            sysbg += '"E2Time":"null",'
        if 'TTime' in kwargs:
            sysbg += '"TTime":"' + kwargs['TTime'] + '",'
        else:
            sysbg += '"TTime":"null",'
        if 'PRLTime' in kwargs:
            sysbg += '"PRLTime":"' + kwargs['PRLTime'] + '",'
        else:
            sysbg += '"PRLTime":"null",'
        if 'IGFBPTime' in kwargs:
            sysbg += '"IGFBPTime":"' + kwargs['IGFBPTime'] + '",'
        else:
            sysbg += '"IGFBPTime":"null",'
        if 'thyroidTime' in kwargs:
            sysbg += '"thyroidTime":"' + kwargs['thyroidTime'] + '",'
        else:
            sysbg += '"thyroidTime":"null",'
        if 'ACTHTime' in kwargs:
            sysbg += '"ACTHTime":"' + kwargs['ACTHTime'] + '",'
        else:
            sysbg += '"ACTHTime":"null",'
        if 'cortisolTime' in kwargs:
            sysbg += '"cortisolTime":"' + kwargs['cortisolTime'] + '",'
        else:
            sysbg += '"cortisolTime":"null",'
        if 'DHEATime' in kwargs:
            sysbg += '"DHEATime":"' + kwargs['DHEATime'] + '",'
        else:
            sysbg += '"DHEATime":"null",'
        if 'OHPTime' in kwargs:
            sysbg += '"OHPTime":"' + kwargs['OHPTime'] + '",'
        else:
            sysbg += '"OHPTime":"null",'
        if 'bloodTime' in kwargs:
            sysbg += '"bloodTime":"' + kwargs['bloodTime'] + '",'
        else:
            sysbg += '"bloodTime":"null",'
        if 'urinalysisTime' in kwargs:
            sysbg += '"urinalysisTime":"' + kwargs['urinalysisTime'] + '",'
        else:
            sysbg += '"urinalysisTime":"null",'
        if 'LAKLGETime' in kwargs:
            sysbg += '"LAKLGETime":"' + kwargs['LAKLGETime'] + '",'
        else:
            sysbg += '"LAKLGETime":"null",'
        if 'HBsTime' in kwargs:
            sysbg += '"HBsTime":"' + kwargs['HBsTime'] + '",'
        else:
            sysbg += '"HBsTime":"null",'
        if 'ghTime' in kwargs:
            sysbg += '"ghTime":"' + kwargs['ghTime'] + '",'
        else:
            sysbg += '"ghTime":"null",'
        if 'gh' in kwargs:
            sysbg += '"gh":"' + kwargs['gh'] + '"}'
        else:
            sysbg += '"gh":"null"}'
        short.lab_exa = sysbg
        #  心电图
        if 'electdiogram' in kwargs:
            short.electr = kwargs['electdiogram']
        #  性腺B超
        if 'uterusOne' in kwargs:
            xxbc = '{"uterusOne":"' + kwargs['uterusOne'] + '",'
        else:
            xxbc = '{"uterusOne":"null",'
        if 'uterusTwo' in kwargs:
            xxbc += '"uterusTwo":"' + kwargs['uterusTwo'] + '",'
        else:
            xxbc += '"uterusTwo":"null",'
        if 'uterusThr' in kwargs:
            xxbc += '"uterusThr":"' + kwargs['uterusThr'] + '",'
        else:
            xxbc += '"uterusThr":"null",'
        if 'cervixLong' in kwargs:
            xxbc += '"cervixLong":"' + kwargs['cervixLong'] + '",'
        else:
            xxbc += '"cervixLong":"null",'
        if 'intima' in kwargs:
            xxbc += '"intima":"' + kwargs['intima'] + '",'
        else:
            xxbc += '"intima":"null",'
        if 'ovaLeftOne' in kwargs:
            xxbc += '"ovaLeftOne":"' + kwargs['ovaLeftOne'] + '",'
        else:
            xxbc += '"ovaLeftOne":"null",'
        if 'ovaLeftTwo' in kwargs:
            xxbc += '"ovaLeftTwo":"' + kwargs['ovaLeftTwo'] + '",'
        else:
            xxbc += '"ovaLeftTwo":"null",'
        if 'ovaLeftThr' in kwargs:
            xxbc += '"ovaLeftThr":"' + kwargs['ovaLeftThr'] + '",'
        else:
            xxbc += '"ovaLeftThr":"null",'
        if 'ovaRightOne' in kwargs:
            xxbc += '"ovaRightOne":"' + kwargs['ovaRightOne'] + '",'
        else:
            xxbc += '"ovaRightOne":"null",'
        if 'ovaRightTwo' in kwargs:
            xxbc += '"ovaRightTwo":"' + kwargs['ovaRightTwo'] + '",'
        else:
            xxbc += '"ovaRightTwo":"null",'
        if 'ovaRightThr' in kwargs:
            xxbc += '"ovaRightThr":"' + kwargs['ovaRightThr'] + '",'
        else:
            xxbc += '"ovaRightThr":"null",'
        if 'follDiameter' in kwargs:
            xxbc += '"follDiameter":"' + kwargs['follDiameter'] + '",'
        else:
            xxbc += '"follDiameter":"null",'
        if 'isCyst' in kwargs:
            xxbc += '"isCyst":"' + kwargs['isCyst'] + '",'
        else:
            xxbc += '"isCyst":"null",'
        if 'cyst' in kwargs:
            xxbc += '"cyst":"' + kwargs['cyst'] + '",'
        else:
            xxbc += '"cyst":"null",'
        if 'cystOne' in kwargs:
            xxbc += '"cystOne":"' + kwargs['cystOne'] + '",'
        else:
            xxbc += '"cystOne":"null",'
        if 'cystTwo' in kwargs:
            xxbc += '"cystTwo":"' + kwargs['cystTwo'] + '",'
        else:
            xxbc += '"cystTwo":"null",'
        if 'cystThr' in kwargs:
            xxbc += '"cystThr":"' + kwargs['cystThr'] + '",'
        else:
            xxbc += '"cystThr":"null",'
        if 'cystDescribe' in kwargs:
            xxbc += '"cystDescribe":"' + kwargs['cystDescribe'] + '",'
        else:
            xxbc += '"cystDescribe":"null",'
        if 'testisLeftOne' in kwargs:
            xxbc += '"testisLeftOne":"' + kwargs['testisLeftOne'] + '",'
        else:
            xxbc += '"testisLeftOne":"null",'
        if 'testisLeftTwo' in kwargs:
            xxbc += '"testisLeftTwo":"' + kwargs['testisLeftTwo'] + '",'
        else:
            xxbc += '"testisLeftTwo":"null",'
        if 'testisLeftThr' in kwargs:
            xxbc += '"testisLeftThr":"' + kwargs['testisLeftThr'] + '",'
        else:
            xxbc += '"testisLeftThr":"null",'
        if 'testisLeftLon' in kwargs:
            xxbc += '"testisLeftLon":"' + kwargs['testisLeftLon'] + '",'
        else:
            xxbc += '"testisLeftLon":"null",'
        if 'testisRightOne' in kwargs:
            xxbc += '"testisRightOne":"' + kwargs['testisRightOne'] + '",'
        else:
            xxbc += '"testisRightOne":"null",'
        if 'testisRightTwo' in kwargs:
            xxbc += '"testisRightTwo":"' + kwargs['testisRightTwo'] + '",'
        else:
            xxbc += '"testisRightTwo":"null",'
        if 'testisRightThr' in kwargs:
            xxbc += '"testisRightThr":"' + kwargs['testisRightThr'] + '",'
        else:
            xxbc += '"testisRightThr":"null",'
        if 'MRI' in kwargs:
            xxbc += '"MRI":"' + kwargs['MRI'] + '",'
        else:
            xxbc += '"MRI":"null",'
        if 'ThyroidLBGradation' in kwargs:
            xxbc += '"ThyroidLBGradation":"' + kwargs['ThyroidLBGradation'] + '",'
        else:
            xxbc += '"ThyroidLBGradation":"null",'
        if 'ThyroidLBSize' in kwargs:
            xxbc += '"ThyroidLBSize":"' + kwargs['ThyroidLBSize'] + '",'
        else:
            xxbc += '"ThyroidLBSize":"null",'
        if 'ThyroidLBLesions' in kwargs:
            xxbc += '"ThyroidLBLesions":"' + kwargs['ThyroidLBLesions'] + '",'
        else:
            xxbc += '"ThyroidLBLesions":"null",'
        if 'ThyroidLBOther' in kwargs:
            xxbc += '"ThyroidLBOther":"' + kwargs['ThyroidLBOther'] + '",'
        else:
            xxbc += '"ThyroidLBOther":"null",'
        if 'ThyroidLB' in kwargs:
            xxbc += '"ThyroidLB":"' + kwargs['ThyroidLB'] + '",'
        else:
            xxbc += '"ThyroidLB":"null",'
        if 'ThyroidRB' in kwargs:
            xxbc += '"ThyroidRB":"' + kwargs['ThyroidRB'] + '",'
        else:
            xxbc += '"ThyroidRB":"null",'
        if 'ThyroidRBGradation' in kwargs:
            xxbc += '"ThyroidRBGradation":"' + kwargs['ThyroidRBGradation'] + '",'
        else:
            xxbc += '"ThyroidRBGradation":"null",'
        if 'ThyroidRBSize' in kwargs:
            xxbc += '"ThyroidRBSize":"' + kwargs['ThyroidRBSize'] + '",'
        else:
            xxbc += '"ThyroidRBSize":"null",'
        if 'ThyroidRBLesions' in kwargs:
            xxbc += '"ThyroidRBLesions":"' + kwargs['ThyroidRBLesions'] + '",'
        else:
            xxbc += '"ThyroidRBLesions":"null",'
        if 'ThyroidRBOther' in kwargs:
            xxbc += '"ThyroidRBOther":"' + kwargs['ThyroidRBOther'] + '",'
        else:
            xxbc += '"ThyroidRBOther":"null",'
        if 'mriDescribe' in kwargs:
            xxbc += '"mriDescribe":"' + kwargs['mriDescribe'] + '",'
        else:
            xxbc += '"mriDescribe":"null",'
        if 'testisRightLon' in kwargs:
            xxbc += '"testisRightLon":"' + kwargs['testisRightLon'] + '"}'
        else:
            xxbc += '"testisRightLon":"null"}'
        short.gon_B_ult = xxbc
        #  诊疗方案
        if 'diaPlan' in kwargs:
            zlfa = '{"diaPlan":"' + kwargs['diaPlan'] + '",'
        else:
            zlfa = '{"diaPlan":"null",'
        if 'rhGH' in kwargs:
            zlfa += '"rhGH":"' + kwargs['rhGH'] + '",'
        else:
            zlfa += '"rhGH":"null",'
        if 'GnRHa' in kwargs:
            zlfa += '"GnRHa":"' + kwargs['GnRHa'] + '",'
        else:
            zlfa += '"GnRHa":"null",'
        if 'GnRHadose' in kwargs:
            zlfa += '"GnRHadose":"' + kwargs['GnRHadose'] + '",'
        else:
            zlfa += '"GnRHadose":"null",'
        if 'rhGHdoseKG' in kwargs:
            zlfa += '"rhGHdoseKG":"' +  kwargs['rhGHdoseKG'] + '",'
        else:
            zlfa += '"rhGHdoseKG":"null",'
        if 'PEGrhGHdose' in kwargs:
            zlfa += '"PEGrhGHdose":"' +  kwargs['PEGrhGHdose'] + '",'
        else:
            zlfa += '"PEGrhGHdose":"null",'
        if 'rhCustomizationDiaPlan' in kwargs:
            zlfa += '"rhCustomizationDiaPlan":"' +  kwargs['rhCustomizationDiaPlan'] + '",'
        else:
            zlfa += '"rhCustomizationDiaPlan":"null",'
        if 'rhCustomizationPrompt' in kwargs:
            zlfa += '"rhCustomizationPrompt":"' +  kwargs['rhCustomizationPrompt'] + '",'
        else:
            zlfa += '"rhCustomizationPrompt":"null",'
        if 'PEGrhCustomizationPrompt' in kwargs:
            zlfa += '"PEGrhCustomizationPrompt":"' +  kwargs['PEGrhCustomizationPrompt'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPrompt":"null",'
        if 'rhCustomizationPromptKG' in kwargs:
            zlfa += '"rhCustomizationPromptKG":"' +  kwargs['rhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"rhCustomizationPromptKG":"null",'
        if 'PEGrhCustomizationPromptKG' in kwargs:
            zlfa += '"PEGrhCustomizationPromptKG":"' +  kwargs['PEGrhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPromptKG":"null",'
        if 'PEGrhGHdoseKG' in kwargs:
            zlfa += '"PEGrhGHdoseKG":"' +  kwargs['PEGrhGHdoseKG'] + '",'
        else:
            zlfa += '"PEGrhGHdoseKG":"null",'
        if 'planData' in kwargs:
            zlfa += '"planData":"' +  json.dumps(kwargs['planData']) + '",'
        else:
            zlfa += '"planData":"null",'
        if 'otherMedicine' in kwargs:
            zlfa += '"otherMedicine":"' +  json.dumps(kwargs['otherMedicine']) + '",'
        else:
            zlfa += '"otherMedicine":"null",'
        if 'rhGHdose' in kwargs:
            zlfa += '"rhGHdose":"' + kwargs['rhGHdose'] + '"}'
        else:
            zlfa += '"rhGHdose":"null"}'
        short.dia_trea_plan = zlfa
        #  生物样本库
        if 'bioBank' in kwargs:
            swkyb = '{"bioBank":"' + kwargs['bioBank'] + '",'
        else:
            swkyb = '{"bioBank":"null",'
        if 'sampleId' in kwargs:
            swkyb += '"sampleId":"' + kwargs['sampleId'] + '",'
        else:
            swkyb += '"sampleId":"null",'
        if 'sampleClass' in kwargs:
            swkyb += '"sampleClass":"' + str(kwargs['sampleClass']) + '"}'
        else:
            swkyb += '"sampleClass":"null"}'
        short.bio_sam_bank = swkyb
        #  父亲生物样本库
        if 'bioBankFa' in kwargs:
            swkyb = '{"bioBankFa":"' + kwargs['bioBankFa'] + '"'
        else:
            swkyb = '{"bioBankFa":"null"'
        if 'sampleIdFa' in kwargs:
            swkyb += ',"sampleIdFa":"' + kwargs['sampleIdFa'] + '"'
        else:
            swkyb += ',"sampleIdFa":"null"'
        if 'sampleClassFa' in kwargs:
            swkyb += ',"sampleClassFa":"' + str(kwargs['sampleClassFa']) + '"}'
        else:
            swkyb += ',"sampleClassFa":"null"}'
        short.f_bio_sam_bank = swkyb
        #  母亲生物样本库
        if 'bioBankMo' in kwargs:
            swkyb = '{"bioBankMo":"' + kwargs['bioBankMo'] + '"'
        else:
            swkyb = '{"bioBankMo":"null"'
        if 'sampleIdMo' in kwargs:
            swkyb += ',"sampleIdMo":"' + kwargs['sampleIdMo'] + '"'
        else:
            swkyb += ',"sampleIdMo":"null"'
        if 'sampleClassMo' in kwargs:
            swkyb += ',"sampleClassMo":"' + str(kwargs['sampleClassMo']) + '"}'
        else:
            swkyb += ',"sampleClassMo":"null"}'
        short.m_bio_sam_bank = swkyb
        #  主要诊断
        mada = ""
        if 'mainDia' in kwargs:
            mada += '{"mainDia":"' + str(kwargs['mainDia']) + '"'
        else:
            mada += '{"mainDia":"null"'
        if 'mainDiaIllustrate' in kwargs:
            mada += ',"mainDiaIllustrate":"' + str(kwargs['mainDiaIllustrate']) + '"'
        else:
            mada += ',"mainDiaIllustrate":"null"'
        if 'DiaIllustrate' in kwargs:
            mada += ',"DiaIllustrate":"' + str(kwargs['DiaIllustrate']) + '"'
        else:
            mada += ',"DiaIllustrate":"null"'
        if 'peripheralityOther' in kwargs:
            mada += ',"peripheralityOther":"' + str(kwargs['peripheralityOther']) + '"'
        else:
            mada += ',"peripheralityOther":"null"'
        if 'partialityOther' in kwargs:
            mada += ',"partialityOther":"' + str(kwargs['partialityOther']) + '"}'
        else:
            mada += ',"partialityOther":"null"}'
        short.main_dia = mada
        #  次要诊断
        if 'secDia' in kwargs:
            short.sec_dia = kwargs['secDia']
        #  随访
        if 'followUp' in kwargs:
            short.follow_up = kwargs['followUp']

        # 染色体核型
        if 'speKar' in kwargs:
            short.spe_kar = kwargs['speKar']
        # SRY基因
        if 'SRY' in kwargs:
            short.SRY = kwargs['SRY']
        #  变异类型
        if 'mutKind' in kwargs:
            short.mut_kind = kwargs['mutKind']
        # 致病基因名称
        if 'genData' in kwargs:
            short.gen_mut_name = json.dumps(kwargs['genData'])
        # 变异来源
        if 'sourMut' in kwargs:
            short.sour_mut = kwargs['sourMut']
        # 核酸变异
        if 'baseMut' in kwargs:
            short.base_mut = kwargs['baseMut']
        # 氨基酸变异
        if 'amiAciMut' in kwargs:
            short.ami_aci_mut = kwargs['amiAciMut']
        short.c_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result = short
    except:
        result = False
    return result

# 修改或添加中枢性性早熟
def modifyorAddSexpre(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            sexp = models.Sexprecocity.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            sexp = models.Sexprecocity()
        # 病人编号
        if 'userNum' in kwargs:
            sexp.user_num = kwargs['userNum']
        sexp.patient_id = casePk
        #  家族史
        if 'fHeight' in kwargs:
            tgjc = '{"fHeight":"' + kwargs['fHeight'] + '",'
        else:
            tgjc = '{"fHeight":"null",'
        if 'fWeight' in kwargs:
            tgjc += '"fWeight":"' + kwargs['fWeight'] + '",'
        else:
            tgjc += '"fWeight":"null",'
        if 'mHeight' in kwargs:
            tgjc += '"mHeight":"' + kwargs['mHeight'] + '",'
        else:
            tgjc += '"mHeight":"null",'
        if 'mWeight' in kwargs:
            tgjc += '"mWeight":"' + kwargs['mWeight'] + '",'
        else:
            tgjc += '"mWeight":"null",'
        if 'firstAge' in kwargs:
            tgjc += '"firstAge":"' + kwargs['firstAge'] + '",'
        else:
            tgjc += '"firstAge":"null",'
        if 'bro' in kwargs:
            tgjc += '"bro":"' + kwargs['bro'] + '",'
        else:
            tgjc += '"bro":"null",'
        if 'isHis' in kwargs:
            tgjc += '"isHis":"' + kwargs['isHis'] + '",'
        else:
            tgjc += '"isHis":"null",'
        if 'oldHis' in kwargs:
            tgjc += '"oldHis":"' + kwargs['oldHis'] + '",'
        else:
            tgjc += '"oldHis":"null",'
        if 'familyData' in kwargs:
            tgjc += '"familyData":' + json.dumps(kwargs['familyData']) + '}'
        else:
            tgjc += '"familyData":"null"}'
        sexp.fam_his = tgjc
        #  初次就诊时间
        if 'firVisTime' in kwargs and kwargs['firVisTime'] is not None and len(kwargs['firVisTime']):
            time = kwargs['firVisTime'][0:10]
            sexp.first_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  发病年龄
        if 'morbidAge' in kwargs:
            sexp.age_ons = kwargs['morbidAge']
        #  主诉
        if 'chiefCom' in kwargs:
            sexp.chi_com = kwargs['chiefCom']
        # #  生长加速
        # if 'growRate' in kwargs:
        #     sexp.acc_growth = kwargs['growRate']
        #  生长加速修改为生长速率
        if 'growRate' in kwargs:
            szjs = '{"growRate":"' + kwargs['growRate'] + '",'
        else:
            szjs = '{"growRate":"null",'
        if 'rate' in kwargs:
            szjs += '"rate":"' + kwargs['rate'] + '"}'
        else:
            szjs += '"rate":"null"}'
        sexp.acc_growth = szjs
        #  月经初潮情况
        if 'menarchy' in kwargs:
            yjcc = '{"menarchy":"' + kwargs['menarchy'] + '",'
        else:
            yjcc = '{"menarchy":"null",'
        if 'menarchyTime' in kwargs:
            yjcc += '"menarchyTime":"' + kwargs['menarchyTime'] + '"}'
        else:
            yjcc += '"menarchyTime":"null"}'
        sexp.menarche = yjcc
        #  体格检查
        if 'height' in kwargs:
            tgjc = '{"height":"' + kwargs['height'] + '",'
        else:
            tgjc = '{"height":"null",'
        if 'weight' in kwargs:
            tgjc += '"weight":"' + kwargs['weight'] + '",'
        else:
            tgjc += '"weight":"null",'
        if 'breastDev' in kwargs:
            tgjc += '"breastDev":"'+kwargs['breastDev']+'",'
        else:
            tgjc += '"breastDev":"null",'
        if 'breastDevRight' in kwargs:
            tgjc += '"breastDevRight":"'+kwargs['breastDevRight']+'",'
        else:
            tgjc += '"breastDevRight":"null",'
        if 'exGenitalia' in kwargs:
            tgjc += '"exGenitalia":"'+kwargs['exGenitalia']+'",'
        else:
            tgjc += '"exGenitalia":"null",'
        if 'pubicHair' in kwargs:
            tgjc += '"pubicHair":"' + kwargs['pubicHair'] + '"}'
        else:
            tgjc += '"pubicHair":"null"}'
        sexp.phy_exa = tgjc
        #  实验室检查
        if 'LH' in kwargs:
            sysjc = '{"LH":"' + kwargs['LH'] + '",'
        else:
            sysjc = '{"LH":"null",'
        if 'FSH' in kwargs:
            sysjc += '"FSH":"' + kwargs['FSH'] + '",'
        else:
            sysjc += '"FSH":"null",'
        if 'E2' in kwargs:
            sysjc += '"E2":"' + kwargs['E2'] + '",'
        else:
            sysjc += '"E2":"null",'
        if 'T' in kwargs:
            sysjc += '"T":"' + kwargs['T'] + '",'
        else:
            sysjc += '"T":"null",'
        if 'HSBG' in kwargs:
            sysjc += '"HSBG":"' + kwargs['HSBG'] + '",'
        else:
            sysjc += '"HSBG":"null",'
        if 'PRL' in kwargs:
            sysjc += '"PRL":"' + kwargs['PRL'] + '",'
        else:
            sysjc += '"PRL":"null",'
        if 'IGF' in kwargs:
            sysjc += '"IGF":"' + kwargs['IGF'] + '",'
        else:
            sysjc += '"IGF":"null",'
        if 'IGFBP3' in kwargs:
            sysjc += '"IGFBP3":"' + kwargs['IGFBP3'] + '",'
        else:
            sysjc += '"IGFBP3":"null",'
        if 'thyroid' in kwargs:
            sysjc += '"thyroid":"' + kwargs['thyroid'] + '",'
        else:
            sysjc += '"thyroid":"null",'
        if 'thyroidDescribe' in kwargs:
            sysjc += '"thyroidDescribe":"' + kwargs['thyroidDescribe'] + '",'
        else:
            sysjc += '"thyroidDescribe":"null",'
        if 'ACTH' in kwargs:
            sysjc += '"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            sysjc += '"ACTH":"null",'
        if 'cortisol' in kwargs:
            sysjc += '"cortisol":"' + kwargs['cortisol'] + '",'
        else:
            sysjc += '"cortisol":"null",'
        if 'DHEAS' in kwargs:
            sysjc += '"DHEAS":"' + kwargs['DHEAS'] + '",'
        else:
            sysjc += '"DHEAS":"null",'
        if 'OHP' in kwargs:
            sysjc += '"OHP":"' + kwargs['OHP'] + '",'
        else:
            sysjc += '"OHP":"null",'
        if 'LAKLGE' in kwargs:
            sysjc += '"LAKLGE":"' + kwargs['LAKLGE'] + '",'
        else:
            sysjc += '"LAKLGE":"null",'
        if 'laklgeDescribe' in kwargs:
            sysjc += '"laklgeDescribe":"' + kwargs['laklgeDescribe'] + '"}'
        else:
            sysjc += '"laklgeDescribe":"null"}'
        sexp.lab_exa = sysjc
        #  心电图
        if 'electdiogram' in kwargs:
            sexp.electr = kwargs['electdiogram']
        #  性腺B超
        if 'uterusOne' in kwargs:
            xxbc = '{"uterusOne":"' + kwargs['uterusOne'] + '",'
        else:
            xxbc = '{"uterusOne":"null",'
        if 'uterusTwo' in kwargs:
            xxbc += '"uterusTwo":"' + kwargs['uterusTwo'] + '",'
        else:
            xxbc += '"uterusTwo":"null",'
        if 'uterusThr' in kwargs:
            xxbc += '"uterusThr":"' + kwargs['uterusThr'] + '",'
        else:
            xxbc += '"uterusThr":"null",'
        if 'cervixLong' in kwargs:
            xxbc += '"cervixLong":"' + kwargs['cervixLong'] + '",'
        else:
            xxbc += '"cervixLong":"null",'
        if 'intima' in kwargs:
            xxbc += '"intima":"' + kwargs['intima'] + '",'
        else:
            xxbc += '"intima":"null",'
        if 'ovaLeftOne' in kwargs:
            xxbc += '"ovaLeftOne":"' + kwargs['ovaLeftOne'] + '",'
        else:
            xxbc += '"ovaLeftOne":"null",'
        if 'ovaLeftTwo' in kwargs:
            xxbc += '"ovaLeftTwo":"' + kwargs['ovaLeftTwo'] + '",'
        else:
            xxbc += '"ovaLeftTwo":"null",'
        if 'ovaLeftThr' in kwargs:
            xxbc += '"ovaLeftThr":"' + kwargs['ovaLeftThr'] + '",'
        else:
            xxbc += '"ovaLeftThr":"null",'
        if 'ovaRightOne' in kwargs:
            xxbc += '"ovaRightOne":"' + kwargs['ovaRightOne'] + '",'
        else:
            xxbc += '"ovaRightOne":"null",'
        if 'ovaRightTwo' in kwargs:
            xxbc += '"ovaRightTwo":"' + kwargs['ovaRightTwo'] + '",'
        else:
            xxbc += '"ovaRightTwo":"null",'
        if 'ovaRightThr' in kwargs:
            xxbc += '"ovaRightThr":"' + kwargs['ovaRightThr'] + '",'
        else:
            xxbc += '"ovaRightThr":"null",'
        if 'follDiameter' in kwargs:
            xxbc += '"follDiameter":"' + kwargs['follDiameter'] + '",'
        else:
            xxbc += '"follDiameter":"null",'
        if 'isCyst' in kwargs:
            xxbc += '"isCyst":"' + kwargs['isCyst'] + '",'
        else:
            xxbc += '"isCyst":"null",'
        if 'cyst' in kwargs:
            xxbc += '"cyst":"' + kwargs['cyst'] + '",'
        else:
            xxbc += '"cyst":"null",'
        if 'cystOne' in kwargs:
            xxbc += '"cystOne":"' + kwargs['cystOne'] + '",'
        else:
            xxbc += '"cystOne":"null",'
        if 'cystTwo' in kwargs:
            xxbc += '"cystTwo":"' + kwargs['cystTwo'] + '",'
        else:
            xxbc += '"cystTwo":"null",'
        if 'cystThr' in kwargs:
            xxbc += '"cystThr":"' + kwargs['cystThr'] + '",'
        else:
            xxbc += '"cystThr":"null",'
        if 'cystDescribe' in kwargs:
            xxbc += '"cystDescribe":"' + kwargs['cystDescribe'] + '",'
        else:
            xxbc += '"cystDescribe":"null",'
        if 'testisLeftOne' in kwargs:
            xxbc += '"testisLeftOne":"' + kwargs['testisLeftOne'] + '",'
        else:
            xxbc += '"testisLeftOne":"null",'
        if 'testisLeftTwo' in kwargs:
            xxbc += '"testisLeftTwo":"' + kwargs['testisLeftTwo'] + '",'
        else:
            xxbc += '"testisLeftTwo":"null",'
        if 'testisLeftThr' in kwargs:
            xxbc += '"testisLeftThr":"' + kwargs['testisLeftThr'] + '",'
        else:
            xxbc += '"testisLeftThr":"null",'
        if 'testisLeftLon' in kwargs:
            xxbc += '"testisLeftLon":"' + kwargs['testisLeftLon'] + '",'
        else:
            xxbc += '"testisLeftLon":"null",'
        if 'testisRightOne' in kwargs:
            xxbc += '"testisRightOne":"' + kwargs['testisRightOne'] + '",'
        else:
            xxbc += '"testisRightOne":"null",'
        if 'testisRightTwo' in kwargs:
            xxbc += '"testisRightTwo":"' + kwargs['testisRightTwo'] + '",'
        else:
            xxbc += '"testisRightTwo":"null",'
        if 'testisRightThr' in kwargs:
            xxbc += '"testisRightThr":"' + kwargs['testisRightThr'] + '",'
        else:
            xxbc += '"testisRightThr":"null",'
        if 'MRI' in kwargs:
            xxbc += '"MRI":"' + kwargs['MRI'] + '",'
        else:
            xxbc += '"MRI":"null",'
        if 'CThyroidLB' in kwargs:
            xxbc += '"CThyroidLB":"' + kwargs['CThyroidLB'] + '",'
        else:
            xxbc += '"CThyroidLB":"null",'
        if 'CThyroidLBGradation' in kwargs:
            xxbc += '"CThyroidLBGradation":"' + kwargs['CThyroidLBGradation'] + '",'
        else:
            xxbc += '"CThyroidLBGradation":"null",'
        if 'CThyroidLBSize' in kwargs:
            xxbc += '"CThyroidLBSize":"' + kwargs['CThyroidLBSize'] + '",'
        else:
            xxbc += '"CThyroidLBSize":"null",'
        if 'CThyroidLBLesions' in kwargs:
            xxbc += '"CThyroidLBLesions":"' + kwargs['CThyroidLBLesions'] + '",'
        else:
            xxbc += '"CThyroidLBLesions":"null",'
        if 'CThyroidLBOther' in kwargs:
            xxbc += '"CThyroidLBOther":"' + kwargs['CThyroidLBOther'] + '",'
        else:
            xxbc += '"CThyroidLBOther":"null",'
        if 'CThyroidRB' in kwargs:
            xxbc += '"CThyroidRB":"' + kwargs['CThyroidRB'] + '",'
        else:
            xxbc += '"CThyroidRB":"null",'
        if 'CThyroidRBGradation' in kwargs:
            xxbc += '"CThyroidRBGradation":"' + kwargs['CThyroidRBGradation'] + '",'
        else:
            xxbc += '"CThyroidRBGradation":"null",'
        if 'CThyroidRBSize' in kwargs:
            xxbc += '"CThyroidRBSize":"' + kwargs['CThyroidRBSize'] + '",'
        else:
            xxbc += '"CThyroidRBSize":"null",'
        if 'CThyroidRBLesions' in kwargs:
            xxbc += '"CThyroidRBLesions":"' + kwargs['CThyroidRBLesions'] + '",'
        else:
            xxbc += '"CThyroidRBLesions":"null",'
        if 'CThyroidRBOther' in kwargs:
            xxbc += '"CThyroidRBOther":"' + kwargs['CThyroidRBOther'] + '",'
        else:
            xxbc += '"CThyroidRBOther":"null",'
        if 'mriDescribe' in kwargs:
            xxbc += '"mriDescribe":"' + kwargs['mriDescribe'] + '",'
        else:
            xxbc += '"mriDescribe":"null",'
        if 'testisRightLon' in kwargs:
            xxbc += '"testisRightLon":"' + kwargs['testisRightLon'] + '"}'
        else:
            xxbc += '"testisRightLon":"null"}'
        sexp.gon_B_ult = xxbc
        #  诊疗方案
        if 'diaPlan' in kwargs:
            zlfa = '{"diaPlan":"' + kwargs['diaPlan'] + '",'
        else:
            zlfa = '{"diaPlan":"null",'
        if 'rhGH' in kwargs:
            zlfa += '"rhGH":"' + kwargs['rhGH'] + '",'
        else:
            zlfa += '"rhGH":"null",'
        if 'GnRHa' in kwargs:
            zlfa += '"GnRHa":"' + kwargs['GnRHa'] + '",'
        else:
            zlfa += '"GnRHa":"null",'
        if 'GnRHadose' in kwargs:
            zlfa += '"GnRHadose":"' + kwargs['GnRHadose'] + '",'
        else:
            zlfa += '"GnRHadose":"null",'
        if 'rhGHdoseKG' in kwargs:
            zlfa += '"rhGHdoseKG":"' +  kwargs['rhGHdoseKG'] + '",'
        else:
            zlfa += '"rhGHdoseKG":"null",'
        if 'PEGrhGHdose' in kwargs:
            zlfa += '"PEGrhGHdose":"' +  kwargs['PEGrhGHdose'] + '",'
        else:
            zlfa += '"PEGrhGHdose":"null",'
        if 'rhCustomizationDiaPlan' in kwargs:
            zlfa += '"rhCustomizationDiaPlan":"' +  kwargs['rhCustomizationDiaPlan'] + '",'
        else:
            zlfa += '"rhCustomizationDiaPlan":"null",'
        if 'rhCustomizationPrompt' in kwargs:
            zlfa += '"rhCustomizationPrompt":"' +  kwargs['rhCustomizationPrompt'] + '",'
        else:
            zlfa += '"rhCustomizationPrompt":"null",'
        if 'PEGrhCustomizationPrompt' in kwargs:
            zlfa += '"PEGrhCustomizationPrompt":"' +  kwargs['PEGrhCustomizationPrompt'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPrompt":"null",'
        if 'rhCustomizationPromptKG' in kwargs:
            zlfa += '"rhCustomizationPromptKG":"' +  kwargs['rhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"rhCustomizationPromptKG":"null",'
        if 'PEGrhCustomizationPromptKG' in kwargs:
            zlfa += '"PEGrhCustomizationPromptKG":"' +  kwargs['PEGrhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPromptKG":"null",'
        if 'PEGrhGHdoseKG' in kwargs:
            zlfa += '"PEGrhGHdoseKG":"' +  kwargs['PEGrhGHdoseKG'] + '",'
        else:
            zlfa += '"PEGrhGHdoseKG":"null",'
        if 'rhGHdose' in kwargs:
            zlfa += '"rhGHdose":"' + kwargs['rhGHdose'] + '"}'
        else:
            zlfa += '"rhGHdose":"null"}'
        sexp.dia_trea_plan = zlfa
        #  生物样本库
        if 'bioBank' in kwargs:
            ybk = '{"bioBank":"' + kwargs['bioBank'] + '",'
        else:
            ybk = '{"bioBank":"null",'
        if 'sampleId' in kwargs:
            ybk += '"sampleId":"' + kwargs['sampleId'] + '",'
        else:
            ybk += '"sampleId":"null",'
        if 'sampleClass' in kwargs:
            ybk += '"sampleClass":"' + str(kwargs['sampleClass']) + '"}'
        else:
            ybk += '"sampleClass":"null"}'
        sexp.bio_sam_bank = ybk
        #  主要诊断
        mada = ""
        if 'mainDia' in kwargs:
            mada += '{"mainDia":"' + str(kwargs['mainDia']) + '"'
        else:
            mada += '{"mainDia":"null"'
        if 'mainDiaIllustrate' in kwargs:
            mada += ',"mainDiaIllustrate":"' + str(kwargs['mainDiaIllustrate']) + '"'
        else:
            mada += ',"mainDiaIllustrate":"null"'
        if 'DiaIllustrate' in kwargs:
            mada += ',"DiaIllustrate":"' + str(kwargs['DiaIllustrate']) + '"'
        else:
            mada += ',"DiaIllustrate":"null"'
        if 'peripheralityOther' in kwargs:
            mada += ',"peripheralityOther":"' + str(kwargs['peripheralityOther']) + '"'
        else:
            mada += ',"peripheralityOther":"null"'
        if 'partialityOther' in kwargs:
            mada += ',"partialityOther":"' + str(kwargs['partialityOther']) + '"}'
        else:
            mada += ',"partialityOther":"null"}'
        sexp.main_dia = mada
        #  次要诊断
        if 'secDia' in kwargs:
            sexp.sec_dia = kwargs['secDia']
        #  随访
        if 'followUp' in kwargs:
            sexp.follow_up = kwargs['followUp']
        # 染色体核型
        if 'speKar' in kwargs:
            sexp.spe_kar = kwargs['speKar']
        # SRY基因
        if 'SRY' in kwargs:
            sexp.SRY = kwargs['SRY']
        #  变异类型
        if 'mutKind' in kwargs:
            sexp.mut_kind = kwargs['mutKind']
        # 致病基因名称
        if 'genData' in kwargs:
            sexp.gen_mut_name = json.dumps(kwargs['genData'])
        # 变异来源
        if 'sourMut' in kwargs:
            sexp.sour_mut = kwargs['sourMut']
        # 核酸变异
        if 'baseMut' in kwargs:
            sexp.base_mut = kwargs['baseMut']
        # 氨基酸变异
        if 'amiAciMut' in kwargs:
            sexp.ami_aci_mut = kwargs['amiAciMut']
        # LHmax
        if 'LHmax' in kwargs:
            sexp.LHmax = kwargs['LHmax']
        # FSHmax
        if 'FSHmax' in kwargs:
            sexp.FSHmax = kwargs['FSHmax']
        # FSHmax
        if 'LFmax' in kwargs:
            sexp.LFmax = kwargs['LFmax']
        sexp.c_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result = sexp
    except:
        result = False
    return result

# 修改或添加MAS
def modifyorAddMas(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            mas = models.Mas.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            mas = models.Mas()
        # 病人编号
        if 'userNum' in kwargs:
            mas.user_num = kwargs['userNum']
        mas.patient_id = casePk
        #  首发表现
        if 'isMainBehave' in kwargs:
            mas.ini_per = kwargs['isMainBehave']
        # 首发表现二
        # if 'isMinorBehave' in kwargs:
        #     tgjc = '{"isMinorBehave":"' + kwargs['isMinorBehave'] + '",'
        # else:
        #     tgjc = '{"isMinorBehave":"null",'
        # if 'isMinorPlace' in kwargs:
        #     tgjc += '"isMinorPlace":"' + kwargs['isMinorPlace'] + '",'
        # else:
        #     tgjc += '"isMinorPlace":"null",'
        # if 'isTestisPlace' in kwargs:
        #     tgjc += '"isTestisPlace":"' + kwargs['isTestisPlace'] + '",'
        # else:
        #     tgjc += '"isTestisPlace":"null",'
        # if 'isSkeletalLesions' in kwargs:
        #     tgjc += '"isSkeletalLesions":"' + kwargs['isSkeletalLesions'] + '",'
        # else:
        #     tgjc += '"isSkeletalLesions":"null",'
        # if 'isSkeletalParts' in kwargs:
        #     tgjc += '"isSkeletalParts":"' + kwargs['isSkeletalParts'] + '",'
        # else:
        #     tgjc += '"isSkeletalParts":"null",'
        # if 'isLimbParts' in kwargs:
        #     tgjc += '"isLimbParts":"' + kwargs['isLimbParts'] + '",'
        # else:
        #     tgjc += '"isLimbParts":"null",'
        # if 'isTorsoBones' in kwargs:
        #     tgjc += '"isTorsoBones":"' + kwargs['isTorsoBones'] + '",'
        # else:
        #     tgjc += '"isTorsoBones":"null",'
        # if 'isCraniofacialBones' in kwargs:
        #     tgjc += '"isCraniofacialBones":"' + kwargs['isCraniofacialBones'] + '",'
        # else:
        #     tgjc += '"isCraniofacialBones":"null",'
        # if 'isCoffeeSpots' in kwargs:
        #     tgjc += '"isCoffeeSpots":"' + kwargs['isCoffeeSpots'] + '",'
        # else:
        #     tgjc += '"isCoffeeSpots":"null",'
        # if 'isCoffeeSpotArea' in kwargs:
        #     tgjc += '"isCoffeeSpotArea":"' + kwargs['isCoffeeSpotArea'] + '",'
        # else:
        #     tgjc += '"isCoffeeSpotArea":"null",'
        # if 'isLocationVision' in kwargs:
        #     tgjc += '"isLocationVision":"' + kwargs['isLocationVision'] + '",'
        # else:
        #     tgjc += '"isLocationVision":"null",'
        # if 'isHearingArea' in kwargs:
        #     tgjc += '"isHearingArea":"' + kwargs['isHearingArea'] + '",'
        # else:
        #     tgjc += '"isHearingArea":"null",'
        # if 'isOtherPresentations' in kwargs:
        #     tgjc += '"isOtherPresentations":"' + kwargs['isOtherPresentations'] + '"}'
        # else:
        #     tgjc += '"isOtherPresentations":"null"}'
        if 'MainPerformanceData' in kwargs:
            mas.ini_per_ci = json.dumps(kwargs['MainPerformanceData'])
        else:
            mas.ini_per_ci = None
        #  首发表现日期
        if 'manifestationCheckTime' in kwargs and kwargs['manifestationCheckTime'] is not None and len(kwargs['manifestationCheckTime']):
            time = kwargs['manifestationCheckTime'][0:10]
            mas.ini_per_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  家族史
        if 'faHeight' in kwargs:
            tgjc = '{"faHeight":"' + kwargs['faHeight'] + '",'
        else:
            tgjc = '{"faHeight":"null",'
        if 'moHeight' in kwargs:
            tgjc += '"moHeight":"' + kwargs['moHeight'] + '",'
        else:
            tgjc += '"moHeight":"null",'
        if 'moWeight' in kwargs:
            tgjc += '"moWeight":"' + kwargs['moWeight'] + '",'
        else:
            tgjc += '"moWeight":"null",'
        if 'faWeight' in kwargs:
            tgjc += '"faWeight":"' + kwargs['faWeight'] + '",'
        else:
            tgjc += '"faWeight":"null",'
        if 'isDiabetesFamily' in kwargs:
            tgjc += '"isDiabetesFamily":"' + kwargs['isDiabetesFamily'] + '",'
        else:
            tgjc += '"isDiabetesFamily":"null",'
        if 'DiabetesDescription' in kwargs:
            tgjc += '"DiabetesDescription":"' + kwargs['DiabetesDescription'] + '",'
        else:
            tgjc += '"DiabetesDescription":"null",'
        if 'isThyroidFamily' in kwargs:
            tgjc += '"isThyroidFamily":"' + kwargs['isThyroidFamily'] + '",'
        else:
            tgjc += '"isThyroidFamily":"null",'
        if 'ThyroidDescription' in kwargs:
            tgjc += '"ThyroidDescription":"' + kwargs['ThyroidDescription'] + '",'
        else:
            tgjc += '"ThyroidDescription":"null",'
        if 'isTumorFamily' in kwargs:
            tgjc += '"isTumorFamily":"' + kwargs['isTumorFamily'] + '",'
        else:
            tgjc += '"isTumorFamily":"null",'
        if 'TumorDescription' in kwargs:
            tgjc += '"TumorDescription":"' + kwargs['TumorDescription'] + '",'
        else:
            tgjc += '"TumorDescription":"null",'
        if 'OtherDiseaseDescriptions' in kwargs:
            tgjc += '"OtherDiseaseDescriptions":"' + kwargs['OtherDiseaseDescriptions'] + '",'
        else:
            tgjc += '"OtherDiseaseDescriptions":"null",'
        if 'isOtherFamily' in kwargs:
            tgjc += '"isOtherFamily":"' + kwargs['isOtherFamily'] + '"}'
        else:
            tgjc += '"isOtherFamily":"null"}'
        mas.fam_his = tgjc
        #  检查日期
        if 'checkTime' in kwargs and kwargs['checkTime'] is not None and len(kwargs['checkTime']):
            time = kwargs['checkTime'][0:10]
            mas.check_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  一般情况
        if 'height' in kwargs:
            sjzh = '{"height":"' + kwargs['height'] + '",'
        else:
            sjzh = '{"height":"null",'
        if 'heightRate' in kwargs:
            sjzh += '"heightRate":"' + kwargs['heightRate'] + '",'
        else:
            sjzh += '"heightRate":"null",'
        if 'weight' in kwargs:
            sjzh += '"weight":"' + kwargs['weight'] + '",'
        else:
            sjzh += '"weight":"null",'
        if 'systolic' in kwargs:
            sjzh += '"systolic":"' + kwargs['systolic'] + '",'
        else:
            sjzh += '"systolic":"null",'
        if 'isTumorFamily' in kwargs:
            sjzh += '"isTumorFamily":"' + kwargs['isTumorFamily'] + '",'
        else:
            sjzh += '"isTumorFamily":"null",'
        if 'diastolic' in kwargs:
            sjzh += '"diastolic":"' + kwargs['diastolic'] + '",'
        else:
            sjzh += '"diastolic":"null",'
        if 'heartRate' in kwargs:
            sjzh += '"heartRate":"' + kwargs['heartRate'] + '"}'
        else:
            sjzh += '"heartRate":"null"}'
        mas.gen_sit = sjzh
        #  女孩发育分期
        if 'leftBreastDev' in kwargs:
            sjzh = '{"leftBreastDev":"' + kwargs['leftBreastDev'] + '",'
        else:
            sjzh = '{"leftBreastDev":"null",'
        if 'rightBreastDev' in kwargs:
            sjzh += '"rightBreastDev":"' + kwargs['rightBreastDev'] + '",'
        else:
            sjzh += '"rightBreastDev":"null",'
        if 'pubicHair' in kwargs:
            sjzh += '"pubicHair":"' + kwargs['pubicHair'] + '",'
        else:
            sjzh += '"pubicHair":"null",'
        if 'breastTend' in kwargs:
            sjzh += '"breastTend":"' + kwargs['breastTend'] + '",'
        else:
            sjzh += '"breastTend":"null",'
        if 'clitoralHypertrophy' in kwargs:
            sjzh += '"clitoralHypertrophy":"' + kwargs['clitoralHypertrophy'] + '",'
        else:
            sjzh += '"clitoralHypertrophy":"null",'
        if 'labialColoration' in kwargs:
            sjzh += '"labialColoration":"' + kwargs['labialColoration'] + '"}'
        else:
            sjzh += '"labialColoration":"null"}'
        mas.girl_sta_dev = sjzh
        #  男孩发育分期
        if 'leftTesticleDev' in kwargs:
            sjzh = '{"leftTesticleDev":"' + kwargs['leftTesticleDev'] + '",'
        else:
            sjzh = '{"leftTesticleDev":"null",'
        if 'rightTesticleDev' in kwargs:
            sjzh += '"rightTesticleDev":"' + kwargs['rightTesticleDev'] + '",'
        else:
            sjzh += '"rightTesticleDev":"null",'
        if 'appleProtrusion' in kwargs:
            sjzh += '"appleProtrusion":"' + kwargs['appleProtrusion'] + '",'
        else:
            sjzh += '"appleProtrusion":"null",'
        if 'breastEnlarg' in kwargs:
            sjzh += '"breastEnlarg":"' + kwargs['breastEnlarg'] + '",'
        else:
            sjzh += '"breastEnlarg":"null",'
        if 'penileGrowth' in kwargs:
            sjzh += '"penileGrowth":"' + kwargs['penileGrowth'] + '"}'
        else:
            sjzh += '"penileGrowth":"null"}'
        mas.boy_sta_dev = sjzh
        #  甲状腺肿大
        if 'goiter' in kwargs:
            mas.goiter = kwargs['goiter']
        #  皮肤检查（多选）
        if 'skinExamination' in kwargs:
            sjzh = '{"skinExamination":' + json.dumps(kwargs['skinExamination']) + ','
        else:
            sjzh = '{"skinExamination":"null",'
        if 'cafeMilkPoint' in kwargs:
            sjzh += '"cafeMilkPoint":"' + kwargs['cafeMilkPoint'] + '"}'
        else:
            sjzh += '"cafeMilkPoint":"null"}'
        mas.skin_exam = sjzh
        #  骨骼检查（多选）
        if 'boneExamination' in kwargs:
            sjzh = '{"boneExamination":' + json.dumps(kwargs['boneExamination']) + ','
        else:
            sjzh = '{"boneExamination":"null",'
        if 'boneSwelling' in kwargs:
            sjzh += '"boneSwelling":"' + kwargs['boneSwelling'] + '",'
        else:
            sjzh += '"boneSwelling":"null",'
        if 'jointDeformity' in kwargs:
            sjzh += '"jointDeformity":"' + kwargs['jointDeformity'] + '",'
        else:
            sjzh += '"jointDeformity":"null",'
        if 'jointPain' in kwargs:
            sjzh += '"jointPain":"' + kwargs['jointPain'] + '",'
        else:
            sjzh += '"jointPain":"null",'
        if 'bonePain' in kwargs:
            sjzh += '"bonePain":"' + kwargs['bonePain'] + '"}'
        else:
            sjzh += '"bonePain":"null"}'
        mas.ske_sur = sjzh
        #  子宫卵巢B超检查日期
        if 'bCheckTime' in kwargs and kwargs['bCheckTime'] is not None and len(kwargs['bCheckTime']):
            time = kwargs['bCheckTime'][0:10]
            mas.ult_exam_ova_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  子宫超声情况
        if 'uterusUlt' in kwargs:
            mas.ute_ult_con = kwargs['uterusUlt']
        #  子宫情况具体描述
        if 'uterusLength' in kwargs:
            sjzh = '{"uterusLength":"' + kwargs['uterusLength'] + '",'
        else:
            sjzh = '{"uterusLength":"null",'
        if 'uterusWidth' in kwargs:
            sjzh += '"uterusWidth":"' + kwargs['uterusWidth'] + '",'
        else:
            sjzh += '"uterusWidth":"null",'
        if 'cervixLength' in kwargs:
            sjzh += '"cervixLength":"' + kwargs['cervixLength'] + '",'
        else:
            sjzh += '"cervixLength":"null",'
        if 'doubleIntima' in kwargs:
            sjzh += '"doubleIntima":"' + kwargs['doubleIntima'] + '",'
        else:
            sjzh += '"doubleIntima":"null",'
        if 'uterineThickness' in kwargs:
            sjzh += '"uterineThickness":"' + kwargs['uterineThickness'] + '"}'
        else:
            sjzh += '"uterineThickness":"null"}'
        mas.spe_des_ute_con = sjzh
        #  卵巢超声情况
        if 'leftOvary' in kwargs:
            sjzh = '{"zclc":"' + kwargs['leftOvary'] + '",'
        else:
            sjzh = '{"zclc":"null",'
        if 'leftOvaryLength' in kwargs:
            sjzh += '"zclcc":"' + kwargs['leftOvaryLength'] + '",'
        else:
            sjzh += '"zclcc":"null",'
        if 'leftOvaryWidth' in kwargs:
            sjzh += '"zclck":"' + kwargs['leftOvaryWidth'] + '",'
        else:
            sjzh += '"zclck":"null",'
        if 'leftOvaryThickness' in kwargs:
            sjzh += '"zclcg":"' + kwargs['leftOvaryThickness'] + '",'
        else:
            sjzh += '"zclcg":"null",'
        if 'leftOvaryCyst' in kwargs:
            sjzh += '"zcnz":"' + kwargs['leftOvaryCyst'] + '",'
        else:
            sjzh += '"zcnz":"null",'
        if 'leftOvaryCystLength' in kwargs:
            sjzh += '"zcnzc":"' + kwargs['leftOvaryCystLength'] + '",'
        else:
            sjzh += '"zcnzc":"null",'
        if 'leftOvaryCystWidth' in kwargs:
            sjzh += '"zcnzk":"' + kwargs['leftOvaryCystWidth'] + '",'
        else:
            sjzh += '"zcnzk":"null",'
        if 'leftOvaryCystThickness' in kwargs:
            sjzh += '"zcnzg":"' + kwargs['leftOvaryCystThickness'] + '",'
        else:
            sjzh += '"zcnzg":"null",'
        if 'rightOvary' in kwargs:
            sjzh += '"yclc":"' + kwargs['rightOvary'] + '",'
        else:
            sjzh += '"yclc":"null",'
        if 'rightOvaryLength' in kwargs:
            sjzh += '"yclcc":"' + kwargs['rightOvaryLength'] + '",'
        else:
            sjzh += '"yclcc":"null",'
        if 'rightOvaryWidth' in kwargs:
            sjzh += '"yclck":"' + kwargs['rightOvaryWidth'] + '",'
        else:
            sjzh += '"yclck":"null",'
        if 'rightOvaryThickness' in kwargs:
            sjzh += '"yclcg":"' + kwargs['rightOvaryThickness'] + '",'
        else:
            sjzh += '"yclcg":"null",'
        if 'rightOvaryCyst' in kwargs:
            sjzh += '"ycnz":"' + kwargs['rightOvaryCyst'] + '",'
        else:
            sjzh += '"ycnz":"null",'
        if 'rightOvaryCystLength' in kwargs:
            sjzh += '"ycnzc":"' + kwargs['rightOvaryCystLength'] + '",'
        else:
            sjzh += '"ycnzc":"null",'
        if 'rightOvaryCystWidth' in kwargs:
            sjzh += '"ycnzk":"' + kwargs['rightOvaryCystWidth'] + '",'
        else:
            sjzh += '"ycnzk":"null",'
        if 'rightOvaryCystThickness' in kwargs:
            sjzh += '"ycnzg":"' + kwargs['rightOvaryCystThickness'] + '"}'
        else:
            sjzh += '"ycnzg":"null"}'
        mas.ova_ult_con = sjzh
        #  甲状腺B超情况
        if 'thyroidUlt' in kwargs:
            sjzh = '{"thyroidUlt":"' + kwargs['thyroidUlt'] + '",'
        else:
            sjzh = '{"thyroidUlt":"null",'
        if 'thyroidUltAbnormal' in kwargs:
            sjzh += '"thyroidUltAbnormal":"' + kwargs['thyroidUltAbnormal'] + '"}'
        else:
            sjzh += '"thyroidUltAbnormal":"null"}'
        mas.thy_ult_con = sjzh
        #  肾上腺B超情况
        if 'adrenalUlt' in kwargs:
            sjzh = '{"adrenalUlt":"' + kwargs['adrenalUlt'] + '",'
        else:
            sjzh = '{"adrenalUlt":"null",'
        if 'adrenalUltAbnormal' in kwargs:
            sjzh += '"adrenalUltAbnormal":"' + kwargs['adrenalUltAbnormal'] + '"}'
        else:
            sjzh += '"adrenalUltAbnormal":"null"}'
        mas.adr_ult_con = sjzh
        #  肾脏B超情况
        if 'renalUlt' in kwargs:
            sjzh = '{"renalUlt":"' + kwargs['renalUlt'] + '",'
        else:
            sjzh = '{"renalUlt":"null",'
        if 'renalUltAbnormal' in kwargs:
            sjzh += '"renalUltAbnormal":"' + kwargs['renalUltAbnormal'] + '"}'
        else:
            sjzh += '"renalUltAbnormal":"null"}'
        mas.ren_ult_con = sjzh
        #  病变骨骼X线片检查情况
        if 'boneX' in kwargs:
            sjzh = '{"boneX":"' + kwargs['boneX'] + '",'
        else:
            sjzh = '{"boneX":"null",'
        if 'boneXAbnormal' in kwargs:
            sjzh += '"boneXAbnormal":"' + kwargs['boneXAbnormal'] + '"}'
        else:
            sjzh += '"boneXAbnormal":"null"}'
        mas.X_exa_dis = sjzh
        #  头颅CT检查情况
        if 'placeCT' in kwargs:
            sjzh = '{"placeCT":"' + kwargs['placeCT'] + '",'
        else:
            sjzh = '{"placeCT":"null",'
        if 'typeCT' in kwargs:
            sjzh += '"typeCT":"' + kwargs['typeCT'] + '",'
        else:
            sjzh += '"typeCT":"null",'
        if 'CTdescription' in kwargs:
            sjzh += '"CTdescription":"' + kwargs['CTdescription'] + '"}'
        else:
            sjzh += '"CTdescription":"null"}'
        mas.hea_ct_exa = sjzh
        #  头颅MR检查情况
        if 'placeMR' in kwargs:
            sjzh = '{"placeMR":"' + kwargs['placeMR'] + '",'
        else:
            sjzh = '{"placeMR":"null",'
        if 'typeMR' in kwargs:
            sjzh += '"typeMR":"' + kwargs['typeMR'] + '",'
        else:
            sjzh += '"typeMR":"null",'
        if 'MRdescription' in kwargs:
            sjzh += '"MRdescription":"' + kwargs['MRdescription'] + '"}'
        else:
            sjzh += '"MRdescription":"null"}'
        mas.hea_mr_exa = sjzh
        #  全身骨扫描检查情况
        if 'bodyBoneScan' in kwargs:
            sjzh = '{"bodyBoneScan":"' + kwargs['bodyBoneScan'] + '",'
        else:
            sjzh = '{"bodyBoneScan":"null",'
        if 'bodyBoneScanAbnormal' in kwargs:
            sjzh += '"bodyBoneScanAbnormal":"' + kwargs['bodyBoneScanAbnormal'] + '"}'
        else:
            sjzh += '"bodyBoneScanAbnormal":"null"}'
        mas.foll_body_scan_exa = sjzh
        #  常规化验检查日期
        if 'commonCheckTime' in kwargs and kwargs['commonCheckTime'] is not None and len(kwargs['commonCheckTime']):
            time = kwargs['commonCheckTime'][0:10]
            mas.lab_exa = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  血常规
        if 'leukocyte' in kwargs:
            sjzh = '{"leukocyte":"' + kwargs['leukocyte'] + '",'
        else:
            sjzh = '{"leukocyte":"null",'
        if 'hemoglobin' in kwargs:
            sjzh += '"hemoglobin":"' + kwargs['hemoglobin'] + '",'
        else:
            sjzh += '"hemoglobin":"null",'
        if 'zxlxb' in kwargs:
            sjzh += '"zxlxb":"' + kwargs['zxlxb'] + '",'
        else:
            sjzh += '"zxlxb":"null",'
        if 'hxbjs' in kwargs:
            sjzh += '"hxbjs":"' + kwargs['hxbjs'] + '",'
        else:
            sjzh += '"hxbjs":"null",'
        if 'Neutrophils' in kwargs:
            sjzh += '"Neutrophils":"' + kwargs['Neutrophils'] + '",'
        else:
            sjzh += '"Neutrophils":"null",'
        if 'erythrocyteNum' in kwargs:
            sjzh += '"erythrocyteNum":"' + kwargs['erythrocyteNum'] + '",'
        else:
            sjzh += '"erythrocyteNum":"null",'
        if 'platelet' in kwargs:
            sjzh += '"platelet":"' + kwargs['platelet'] + '"}'
        else:
            sjzh += '"platelet":"null"}'
        mas.blo_rou = sjzh
        #  肝功能
        if 'ALT' in kwargs:
            sjzh = '{"ALT":"' + kwargs['ALT'] + '",'
        else:
            sjzh = '{"ALT":"null",'
        if 'AST' in kwargs:
            sjzh += '"AST":"' + kwargs['AST'] + '",'
        else:
            sjzh += '"AST":"null",'
        if 'LDH' in kwargs:
            sjzh += '"LDH":"' + kwargs['LDH'] + '",'
        else:
            sjzh += '"LDH":"null",'
        if 'gamaGT' in kwargs:
            sjzh += '"gamaGT":"' + kwargs['gamaGT'] + '",'
        else:
            sjzh += '"gamaGT":"null",'
        if 'totalBilirubin' in kwargs:
            sjzh += '"totalBilirubin":"' + kwargs['totalBilirubin'] + '",'
        else:
            sjzh += '"totalBilirubin":"null",'
        if 'directBilirubin' in kwargs:
            sjzh += '"directBilirubin":"' + kwargs['directBilirubin'] + '",'
        else:
            sjzh += '"directBilirubin":"null",'
        if 'indirectBilirubin' in kwargs:
            sjzh += '"indirectBilirubin":"' + kwargs['indirectBilirubin'] + '"}'
        else:
            sjzh += '"indirectBilirubin":"null"}'
        mas.liv_fun = sjzh
        #  肾功能
        if 'urea' in kwargs:
            sjzh = '{"urea":"' + kwargs['urea'] + '",'
        else:
            sjzh = '{"urea":"null",'
        if 'creatinine' in kwargs:
            sjzh += '"creatinine":"' + kwargs['creatinine'] + '",'
        else:
            sjzh += '"creatinine":"null",'
        if 'uricAcid' in kwargs:
            sjzh += '"uricAcid":"' + kwargs['uricAcid'] + '"}'
        else:
            sjzh += '"uricAcid":"null"}'
        mas.ren_fun = sjzh
        #  电解质
        if 'bloodK' in kwargs:
            sjzh = '{"bloodK":"' + kwargs['bloodK'] + '",'
        else:
            sjzh = '{"bloodK":"null",'
        if 'bloodNa' in kwargs:
            sjzh += '"bloodNa":"' + kwargs['bloodNa'] + '",'
        else:
            sjzh += '"bloodNa":"null",'
        if 'bloodCl' in kwargs:
            sjzh += '"bloodCl":"' + kwargs['bloodCl'] + '"}'
        else:
            sjzh += '"bloodCl":"null"}'
        mas.electrolyte = sjzh
        #  血脂
        if 'TC' in kwargs:
            sjzh = '{"TC":"' + kwargs['TC'] + '",'
        else:
            sjzh = '{"TC":"null",'
        if 'TG' in kwargs:
            sjzh += '"TG":"' + kwargs['TG'] + '",'
        else:
            sjzh += '"TG":"null",'
        if 'HDL' in kwargs:
            sjzh += '"HDL":"' + kwargs['HDL'] + '",'
        else:
            sjzh += '"HDL":"null",'
        if 'LDL' in kwargs:
            sjzh += '"LDL":"' + kwargs['LDL'] + '"}'
        else:
            sjzh += '"LDL":"null"}'
        mas.blood_fat = sjzh
        #  骨代谢检查
        if 'bloodCa' in kwargs:
            sjzh = '{"bloodCa":"' + kwargs['bloodCa'] + '",'
        else:
            sjzh = '{"bloodCa":"null",'
        if 'bloodP' in kwargs:
            sjzh += '"bloodP":"' + kwargs['bloodP'] + '",'
        else:
            sjzh += '"bloodP":"null",'
        if 'CTX' in kwargs:
            sjzh += '"CTX":"' + kwargs['CTX'] + '",'
        else:
            sjzh += '"CTX":"null",'
        if 'BGP' in kwargs:
            sjzh += '"BGP":"' + kwargs['BGP'] + '",'
        else:
            sjzh += '"BGP":"null",'
        if 'PINP' in kwargs:
            sjzh += '"PINP":"' + kwargs['PINP'] + '",'
        else:
            sjzh += '"PINP":"null",'
        if 'PTH' in kwargs:
            sjzh += '"PTH":"' + kwargs['PTH'] + '",'
        else:
            sjzh += '"PTH":"null",'
        if 'OHD25' in kwargs:
            sjzh += '"OHD25":"' + kwargs['OHD25'] + '",'
        else:
            sjzh += '"OHD25":"null",'
        if 'ALP' in kwargs:
            sjzh += '"ALP":"' + kwargs['ALP'] + '",'
        else:
            sjzh += '"ALP":"null",'
        if 'urineCa' in kwargs:
            sjzh += '"urineCa":"' + kwargs['urineCa'] + '",'
        else:
            sjzh += '"urineCa":"null",'
        if 'urineP' in kwargs:
            sjzh += '"urineP":"' + kwargs['urineP'] + '"}'
        else:
            sjzh += '"urineP":"null"}'
        mas.bone_met_exa = sjzh
        #  骨代谢检查日期
        if 'boneCheckTime' in kwargs and kwargs['boneCheckTime'] is not None and len(kwargs['boneCheckTime']):
            time = kwargs['boneCheckTime'][0:10]
            mas.bone_met_exa_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  性激素检查
        if 'LH' in kwargs:
            sjzh = '{"LH":"' + kwargs['LH'] + '",'
        else:
            sjzh = '{"LH":"null",'
        if 'FSH' in kwargs:
            sjzh += '"FSH":"' + kwargs['FSH'] + '",'
        else:
            sjzh += '"FSH":"null",'
        if 'E2' in kwargs:
            sjzh += '"E2":"' + kwargs['E2'] + '",'
        else:
            sjzh += '"E2":"null",'
        if 'T' in kwargs:
            sjzh += '"T":"' + kwargs['T'] + '",'
        else:
            sjzh += '"T":"null",'
        if 'PRL' in kwargs:
            sjzh += '"PRL":"' + kwargs['PRL'] + '"}'
        else:
            sjzh += '"PRL":"null"}'
        mas.sex_hor_exa = sjzh
        #  性激素检查日期
        if 'sexHormoneCheckTime' in kwargs and kwargs['sexHormoneCheckTime'] is not None and len(kwargs['sexHormoneCheckTime']):
            time = kwargs['sexHormoneCheckTime'][0:10]
            mas.sex_hor_exa_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  甲状腺功能及抗体检查
        if 'TT4' in kwargs:
            sjzh = '{"TT4":"' + kwargs['TT4'] + '",'
        else:
            sjzh = '{"TT4":"null",'
        if 'TT3' in kwargs:
            sjzh += '"TT3":"' + kwargs['TT3'] + '",'
        else:
            sjzh += '"TT3":"null",'
        if 'TSH' in kwargs:
            sjzh += '"TSH":"' + kwargs['TSH'] + '",'
        else:
            sjzh += '"TSH":"null",'
        if 'FT4' in kwargs:
            sjzh += '"FT4":"' + kwargs['FT4'] + '",'
        else:
            sjzh += '"FT4":"null",'
        if 'FT3' in kwargs:
            sjzh += '"FT3":"' + kwargs['FT3'] + '",'
        else:
            sjzh += '"FT3":"null",'
        if 'TPOAb' in kwargs:
            sjzh += '"TPOAb":"' + kwargs['TPOAb'] + '",'
        else:
            sjzh += '"TPOAb":"null",'
        if 'TGAb' in kwargs:
            sjzh += '"TGAb":"' + kwargs['TGAb'] + '"}'
        else:
            sjzh += '"TGAb":"null"}'
        mas.thy_fun_ant_exa = sjzh
        #  甲状腺功能及抗体检查日期
        if 'thyroidFunctionCheckTime' in kwargs and kwargs['thyroidFunctionCheckTime'] is not None and len(kwargs['thyroidFunctionCheckTime']):
            time = kwargs['thyroidFunctionCheckTime'][0:10]
            mas.thy_fun_ant_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  肾上腺功能检查
        if 'ACTH' in kwargs:
            sjzh = '{"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            sjzh = '{"ACTH":"null",'
        if 'ACTH8' in kwargs:
            sjzh += '"ACTH8":"' + kwargs['ACTH8'] + '",'
        else:
            sjzh += '"ACTH8":"null",'
        if 'ACTH4' in kwargs:
            sjzh += '"ACTH4":"' + kwargs['ACTH4'] + '",'
        else:
            sjzh += '"ACTH4":"null",'
        if 'AM8' in kwargs:
            sjzh += '"AM8":"' + kwargs['AM8'] + '",'
        else:
            sjzh += '"AM8":"null",'
        if 'PM4' in kwargs:
            sjzh += '"PM4":"' + kwargs['PM4'] + '",'
        else:
            sjzh += '"PM4":"null",'
        if 'UFC' in kwargs:
            sjzh += '"UFC":"' + kwargs['UFC'] + '"}'
        else:
            sjzh += '"UFC":"null"}'
        mas.adr_fun_exa = sjzh
        #  肾上腺功能检查日期
        if 'adrenalFunctionCheckTime' in kwargs and kwargs['adrenalFunctionCheckTime'] is not None and len(kwargs['adrenalFunctionCheckTime']):
            time = kwargs['adrenalFunctionCheckTime'][0:10]
            mas.adr_fun_exa_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  生长激素分泌功能检查
        if 'GH' in kwargs:
            sjzh = '{"GH":"' + kwargs['GH'] + '",'
        else:
            sjzh = '{"GH":"null",'
        if 'IGF1' in kwargs:
            sjzh += '"IGF1":"' + kwargs['IGF1'] + '",'
        else:
            sjzh += '"IGF1":"null",'
        if 'IGFBP3' in kwargs:
            sjzh += '"IGFBP3":"' + kwargs['IGFBP3'] + '"}'
        else:
            sjzh += '"IGFBP3":"null"}'
        mas.phy_exa = sjzh
        #  生长激素分泌功能检查时间
        if 'somatotropinFunctionCheckTime' in kwargs and kwargs['somatotropinFunctionCheckTime'] is not None and len(kwargs['somatotropinFunctionCheckTime']):
            time = kwargs['somatotropinFunctionCheckTime'][0:10]
            mas.gro_hor_exa = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  糖代谢情况
        if 'FBS' in kwargs:
            sjzh = '{"FBS":"' + kwargs['FBS'] + '",'
        else:
            sjzh = '{"FBS":"null",'
        if 'FINS' in kwargs:
            sjzh += '"FINS":"' + kwargs['FINS'] + '",'
        else:
            sjzh += '"FINS":"null",'
        if 'FCP' in kwargs:
            sjzh += '"FCP":"' + kwargs['FCP'] + '",'
        else:
            sjzh += '"FCP":"null",'
        if 'HbA1c' in kwargs:
            sjzh += '"HbA1c":"' + kwargs['HbA1c'] + '"}'
        else:
            sjzh += '"HbA1c":"null"}'
        mas.glu_met = sjzh
        #  糖代谢情况时间
        if 'glycometabolismFunctionCheckTime' in kwargs and kwargs['glycometabolismFunctionCheckTime'] is not None and len(kwargs['glycometabolismFunctionCheckTime']):
            time = kwargs['glycometabolismFunctionCheckTime'][0:10]
            mas.glu_met_date = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  心电图检查
        if 'ecgExamination' in kwargs:
            sjzh = '{"ecgExamination":"' + kwargs['ecgExamination'] + '",'
        else:
            sjzh = '{"ecgExamination":"null",'
        if 'ecgExaminationAbnormal' in kwargs:
            sjzh += '"ecgExaminationAbnormal":"' + kwargs['ecgExaminationAbnormal'] + '"}'
        else:
            sjzh += '"ecgExaminationAbnormal":"null"}'
        mas.ecg_exa = sjzh
        #  X线骨龄检查
        if 'XboneAge' in kwargs:
            mas.x_bone_exa = kwargs['XboneAge']
        #  垂体MR检查
        if 'pituitaryMR' in kwargs:
            sjzh = '{"pituitaryMR":"' + kwargs['pituitaryMR'] + '",'
        else:
            sjzh = '{"pituitaryMR":"null",'
        if 'pituitaryMRAbnormal' in kwargs:
            sjzh += '"pituitaryMRAbnormal":"' + kwargs['pituitaryMRAbnormal'] + '"}'
        else:
            sjzh += '"pituitaryMRAbnormal":"null"}'
        mas.pit_exa = sjzh
        #  GNAS基因测定是否检查
        if 'GNAS' in kwargs:
            mas.GNAS = kwargs['GNAS']
        #  标本采样类型或部位
        if 'GNASSampling' in kwargs:
            sjzh = '{"GNASSampling":"' + kwargs['GNASSampling'] + '",'
        else:
            sjzh = '{"GNASSampling":"null",'
        if 'gnasSamplingPosition' in kwargs:
            sjzh += '"gnasSamplingPosition":"' + kwargs['gnasSamplingPosition'] + '"}'
        else:
            sjzh += '"gnasSamplingPosition":"null"}'
        mas.GNAS_sam_loc = sjzh

        #  遗传学检测方法
        if 'GNASSMethods' in kwargs:
            ycff = '{"GNASSMethods":"' + kwargs['GNASSMethods'] + '",'
        else:
            ycff = '{"GNASSMethods":"null",'
        if 'OtherAssays' in kwargs:
            ycff += '"OtherAssays":"' + kwargs['OtherAssays'] + '"}'
        else:
            ycff += '"OtherAssays":"null"}'
        mas.gen_tes_met = ycff
        #  检测结果
        if 'TestResults' in kwargs:
            mas.det_res = kwargs['TestResults']
        #  检测版本
        if 'DetectVersion' in kwargs:
            mas.det_ver = kwargs['DetectVersion']
        #  突变位点
        if 'MutationLocation' in kwargs:
            mas.mut_sit = kwargs['MutationLocation']

        #  病理活检是否检查
        if 'pathBiopsyExamination' in kwargs:
            mas.pat_exa = kwargs['pathBiopsyExamination']
        #  标本采样类型或部位
        if 'pathBiopsyPosition' in kwargs:
            mas.pat_sam_loc = kwargs['pathBiopsyPosition']
        #  是否行GnRH激发试验
        if 'GnRH' in kwargs:
            mas.GnRH = kwargs['GnRH']
        #  评估指标
        if 'GnRHDrugName' in kwargs:
            sjzh = '{"GnRHDrugName":"' + kwargs['GnRHDrugName'] + '",'
        else:
            sjzh = '{"GnRHDrugName":"null",'
        if 'GnRHDrugDosage' in kwargs:
            sjzh += '"GnRHDrugDosage":"' + kwargs['GnRHDrugDosage'] + '",'
        else:
            sjzh += '"GnRHDrugDosage":"null",'
        if 'GnRHUsageTime' in kwargs:
            sjzh += '"GnRHUsageTime":"' + kwargs['GnRHUsageTime'] + '",'
        else:
            sjzh += '"GnRHUsageTime":"null",'
        if 'LFMax' in kwargs:
            sjzh += '"LFMax":"' + kwargs['LFMax'] + '",'
        else:
            sjzh += '"LFMax":"null",'
        if 'FSHMax' in kwargs:
            sjzh += '"FSHMax":"' + kwargs['FSHMax'] + '",'
        else:
            sjzh += '"FSHMax":"null",'
        if 'LFRatio' in kwargs:
            sjzh += '"LFRatio":"' + kwargs['LFRatio'] + '"}'
        else:
            sjzh += '"LFRatio":"null"}'
        mas.GnRH_eva = sjzh
        #  是否行小剂量地塞米松抑制试验
        if 'LDDST' in kwargs:
            mas.low_dose = kwargs['LDDST']
        #  评估指标
        if 'LDDSTDrugName' in kwargs:
            sjzh = '{"LDDSTDrugName":"' + kwargs['LDDSTDrugName'] + '",'
        else:
            sjzh = '{"LDDSTDrugName":"null",'
        if 'LDDSTDrugDosage' in kwargs:
            sjzh += '"LDDSTDrugDosage":"' + kwargs['LDDSTDrugDosage'] + '",'
        else:
            sjzh += '"LDDSTDrugDosage":"null",'
        if 'LDDSTUsageTime' in kwargs:
            sjzh += '"LDDSTUsageTime":"' + kwargs['LDDSTUsageTime'] + '",'
        else:
            sjzh += '"LDDSTUsageTime":"null",'
        if 'ACTHAfter' in kwargs:
            sjzh += '"ACTHAfter":"' + kwargs['ACTHAfter'] + '",'
        else:
            sjzh += '"ACTHAfter":"null",'
        if 'ACTHBefore' in kwargs:
            sjzh += '"ACTHBefore":"' + kwargs['ACTHBefore'] + '",'
        else:
            sjzh += '"ACTHBefore":"null",'
        if 'cortisolAfter' in kwargs:
            sjzh += '"cortisolAfter":"' + kwargs['cortisolAfter'] + '",'
        else:
            sjzh += '"cortisolAfter":"null",'
        if 'cortisolBefore' in kwargs:
            sjzh += '"cortisolBefore":"' + kwargs['cortisolBefore'] + '",'
        else:
            sjzh += '"cortisolBefore":"null",'
        if 'UFFAfter' in kwargs:
            sjzh += '"UFFAfter":"' + kwargs['UFFAfter'] + '",'
        else:
            sjzh += '"UFFAfter":"null",'
        if 'UFFBefore' in kwargs:
            sjzh += '"UFFBefore":"' + kwargs['UFFBefore'] + '"}'
        else:
            sjzh += '"UFFBefore":"null"}'
        mas.low_dose_eva = sjzh
        #  是否行生长激素-葡萄糖抑制试验
        if 'GHGIT' in kwargs:
            mas.gro_glu = kwargs['GHGIT']
        #  评估指标
        if 'GHGITDrugName' in kwargs:
            sjzh = '{"GHGITDrugName":"' + kwargs['GHGITDrugName'] + '",'
        else:
            sjzh = '{"GHGITDrugName":"null",'
        if 'GHGITDrugDosage' in kwargs:
            sjzh += '"GHGITDrugDosage":"' + kwargs['GHGITDrugDosage'] + '",'
        else:
            sjzh += '"GHGITDrugDosage":"null",'
        if 'GHGITUsageTime' in kwargs:
            sjzh += '"GHGITUsageTime":"' + kwargs['GHGITUsageTime'] + '",'
        else:
            sjzh += '"GHGITUsageTime":"null",'
        if 'ACTHAfter' in kwargs:
            sjzh += '"ACTHAfter":"' + kwargs['ACTHAfter'] + '",'
        else:
            sjzh += '"ACTHAfter":"null",'
        if 'GH0' in kwargs:
            sjzh += '"GH0":"' + kwargs['GH0'] + '",'
        else:
            sjzh += '"GH0":"null",'
        if 'GH3' in kwargs:
            sjzh += '"GH3":"' + kwargs['GH3'] + '",'
        else:
            sjzh += '"GH3":"null",'
        if 'GH6' in kwargs:
            sjzh += '"GH6":"' + kwargs['GH6'] + '",'
        else:
            sjzh += '"GH6":"null",'
        if 'GH9' in kwargs:
            sjzh += '"GH9":"' + kwargs['GH9'] + '",'
        else:
            sjzh += '"GH9":"null",'
        if 'XTZ0' in kwargs:
            sjzh += '"XTZ0":"' + kwargs['XTZ0'] + '",'
        else:
            sjzh += '"XTZ0":"null",'
        if 'XTZ3' in kwargs:
            sjzh += '"XTZ3":"' + kwargs['XTZ3'] + '",'
        else:
            sjzh += '"XTZ3":"null",'
        if 'XTZ6' in kwargs:
            sjzh += '"XTZ6":"' + kwargs['XTZ6'] + '",'
        else:
            sjzh += '"XTZ6":"null",'
        if 'XTZ9' in kwargs:
            sjzh += '"XTZ9":"' + kwargs['XTZ9'] + '",'
        else:
            sjzh += '"XTZ9":"null",'
        if 'XTZ12' in kwargs:
            sjzh += '"XTZ12":"' + kwargs['XTZ12'] + '",'
        else:
            sjzh += '"XTZ12":"null",'
        if 'GH12' in kwargs:
            sjzh += '"GH12":"' + kwargs['GH12'] + '"}'
        else:
            sjzh += '"GH12":"null"}'
        mas.gro_glu_eva = sjzh
        #  是否存在性早熟
        if 'isSexualPrecocity' in kwargs:
            mas.sex_pre = kwargs['isSexualPrecocity']
        #  是否存在甲状腺功能亢进
        if 'isHyperthyroidism' in kwargs:
            mas.hyper = kwargs['isHyperthyroidism']
        #  是否存在甲状腺功能亢进
        if 'isGrowthHormone' in kwargs:
            mas.is_gro_hor = kwargs['isGrowthHormone']
        #  是否存在皮质醇增多症
        if 'isIncreasedCortisol' in kwargs:
            mas.is_inc_cor = kwargs['isIncreasedCortisol']
        #  填表日期
        mas.fill_date = datetime.datetime.now()
        #  填表医生
        mas.fill_doctor = "1"
        #  填表医生手机号码
        mas.fill_doctor_phone = "18069815584"
        mas.del_flg = "1"
        result = mas
    except:
        result = False
    return result

# 修改或添加 MasFoll
def modifyorAddMasFoll(casePk=0, kwargs=0):
    try:
        if casePk and casePk > 0:
            try:
                masfoll = models.MasFoll.objects.get(mas__pk = casePk)
            except:
                masfoll = models.MasFoll()
        # 是否达终身高
        if 'isFinalHeight' in kwargs:
            sjzh = '{"isFinalHeight":"' + kwargs['isFinalHeight'] + '",'
        else:
            sjzh = '{"isFinalHeight":"null",'
        if 'finalHeight' in kwargs:
            sjzh += '"finalHeight":"' + kwargs['finalHeight'] + '"}'
        else:
            sjzh += '"finalHeight":"null"}'
        masfoll.is_finalhei = sjzh
        masfoll.mas_id = casePk
        #  有无对外周性性早熟
        if 'isPPP' in kwargs:
            tgjc = '{"isPPP":"' + kwargs['isPPP'] + '",'
        else:
            tgjc = '{"isPPP":"null",'
        if 'isPrecociousPuberty' in kwargs:
            tgjc += '"isPrecociousPuberty":"' + kwargs['isPrecociousPuberty'] + '",'
        else:
            tgjc += '"isPrecociousPuberty":"null",'
        if 'treatmentCyclePPP' in kwargs:
            tgjc += '"treatmentCyclePPP":' + json.dumps(kwargs['treatmentCyclePPP']) + '}'
        else:
            tgjc += '"treatmentCyclePPP":"null"}'
        masfoll.is_per_pre = tgjc
        #  随访情况
        if 'precocityData' in kwargs:
            masfoll.per_pre_sf = json.dumps(kwargs['precocityData'])
        #  有无对甲状腺功能亢进进行治疗
        if 'isHyperthyreosis' in kwargs:
            tgjc = '{"isHyperthyreosis":"' + kwargs['isHyperthyreosis'] + '",'
        else:
            tgjc = '{"isHyperthyreosis":"null",'
        if 'isThyroidFunction' in kwargs:
            tgjc += '"isThyroidFunction":"' + kwargs['isThyroidFunction'] + '",'
        else:
            tgjc += '"isThyroidFunction":"null",'
        if 'treatmentCycleHyper' in kwargs:
            tgjc += '"treatmentCycleHyper":' + json.dumps(kwargs['treatmentCycleHyper']) + '}'
        else:
            tgjc += '"treatmentCycleHyper":"null"}'
        masfoll.is_hyper = tgjc
        #  随访情况
        if 'hyperData' in kwargs:
            masfoll.hyper_sf = json.dumps(kwargs['hyperData'])
        # #  监测指标
        # if 'morbidAge' in kwargs:
        #     masfoll.hyper_jc = kwargs['morbidAge']
        #  有无对生长激素分泌过多进行治疗
        if 'isGrowth' in kwargs:
            tgjc = '{"isGrowth":"' + kwargs['isGrowth'] + '",'
        else:
            tgjc = '{"isGrowth":"null",'
        if 'isGrowthHormonePlethora' in kwargs:
            tgjc += '"isGrowthHormonePlethora":"' + kwargs['isGrowthHormonePlethora'] + '",'
        else:
            tgjc += '"isGrowthHormonePlethora":"null",'
        if 'treatmentCycleGrowth' in kwargs:
            tgjc += '"treatmentCycleGrowth":' + json.dumps(kwargs['treatmentCycleGrowth']) + '}'
        else:
            tgjc += '"treatmentCycleGrowth":"null"}'
        masfoll.is_gro_hor = tgjc
        #  随访情况
        if 'growthData' in kwargs:
            masfoll.gro_hor_sf = json.dumps(kwargs['growthData'])
        #  有无对高泌乳素血症进行治疗
        if 'isHPRL' in kwargs:
            tgjc = '{"isHPRL":"' + kwargs['isHPRL'] + '",'
        else:
            tgjc = '{"isHPRL":"null",'
        if 'isHyperprolactinemia' in kwargs:
            tgjc += '"isHyperprolactinemia":"' + kwargs['isHyperprolactinemia'] + '",'
        else:
            tgjc += '"isHyperprolactinemia":"null",'
        if 'treatmentCycleHPRL' in kwargs:
            tgjc += '"treatmentCycleHPRL":' + json.dumps(kwargs['treatmentCycleHPRL']) + '}'
        else:
            tgjc += '"treatmentCycleHPRL":"null"}'
        masfoll.is_tre_hpy = tgjc
        #  随访情况
        if 'HPRLData' in kwargs:
            masfoll.tre_hpy_sf = json.dumps(kwargs['HPRLData'])
        #  有无对皮质醇增多症进行治疗
        if 'isCortisol' in kwargs:
            tgjc = '{"isCortisol":"' + kwargs['isCortisol'] + '",'
        else:
            tgjc = '{"isCortisol":"null",'
        if 'isHypercortisolism' in kwargs:
            tgjc += '"isHypercortisolism":"' + kwargs['isHypercortisolism'] + '",'
        else:
            tgjc += '"isHypercortisolism":"null",'
        if 'treatmentCycleCortisol' in kwargs:
            tgjc += '"treatmentCycleCortisol":' + json.dumps(kwargs['treatmentCycleCortisol']) + '}'
        else:
            tgjc += '"treatmentCycleCortisol":"null"}'
        masfoll.is_inc_cor = tgjc
        #  随访情况
        if 'cortisolData' in kwargs:
            masfoll.inc_cor_sf = json.dumps(kwargs['cortisolData'])
        # #  监测指标
        # if 'morbidAge' in kwargs:
        #     masfoll.inc_cor_jc = kwargs['morbidAge']
        #  是否行颅内手术
        if 'isIntracranialSurgery' in kwargs:
            masfoll.is_int_sur = kwargs['isIntracranialSurgery']
        #  是否行双侧肾上腺切除术
        if 'isBilateralAdrenalectomy' in kwargs:
            masfoll.is_bil_adr = kwargs['isBilateralAdrenalectomy']
        #  是否对骨痛进行治疗
        if 'isOstealgia' in kwargs:
            tgjc = '{"isOstealgia":"' + kwargs['isOstealgia'] + '",'
        else:
            tgjc = '{"isOstealgia":"null",'
        if 'isTreatBonePain' in kwargs:
            tgjc += '"isTreatBonePain":' + json.dumps(kwargs['isTreatBonePain']) + '}'
        else:
            tgjc += '"isTreatBonePain":"null"}'
        masfoll.is_bon_pai = tgjc
        #  随访情况
        if 'bonePainData' in kwargs:
            masfoll.bon_pai_sf = json.dumps(kwargs['bonePainData'])
        # #  监测指标
        # if 'morbidAge' in kwargs:
        #     masfoll.bon_pai_jc = kwargs['morbidAge']
        #  是否对低磷酸盐血症进行治疗
        if 'isHaveHypophosphatemia' in kwargs:
            tgjc = '{"isHaveHypophosphatemia":"' + kwargs['isHaveHypophosphatemia'] + '",'
        else:
            tgjc = '{"isHaveHypophosphatemia":"null",'
        if 'isHypophosphatemia' in kwargs:
            tgjc += '"isHypophosphatemia":' + json.dumps(kwargs['isHypophosphatemia']) + '}'
        else:
            tgjc += '"isHypophosphatemia":"null"}'
        masfoll.hypop = tgjc
        #  随访情况
        if 'hypophosphatemiaData' in kwargs:
            masfoll.hypop_sf = json.dumps(kwargs['hypophosphatemiaData'])
        #  是否行骨骼外科手术
        if 'isSkeletalSurgery' in kwargs:
            tgjc = '{"isSkeletalSurgery":"' + kwargs['isSkeletalSurgery'] + '",'
        else:
            tgjc = '{"isSkeletalSurgery":"null",'
        if 'isHaveSkeletalSurgery' in kwargs:
            tgjc += '"isHaveSkeletalSurgery":"' + kwargs['isHaveSkeletalSurgery'] + '",'
        else:
            tgjc += '"isHaveSkeletalSurgery":"null",'
        if 'surgicalPurpose' in kwargs:
            tgjc += '"surgicalPurpose":"' + kwargs['surgicalPurpose'] + '"}'
        else:
            tgjc += '"surgicalPurpose":"null"}'
        masfoll.is_ske_sur = tgjc
        #  是否对牛奶咖啡斑进行激光治疗
        if 'isLaserTherapy' in kwargs:
            masfoll.is_cafe_spot = kwargs['isLaserTherapy']
        #  是否进形心理疏导
        if 'isPsychologicalCounseling' in kwargs:
            masfoll.is_psy_cou = kwargs['isPsychologicalCounseling']
        #  生存状态
        if 'isSurvivalState' in kwargs:
            tgjc = '{"isSurvivalState":"' + kwargs['isSurvivalState'] + '",'
        else:
            tgjc = '{"isSurvivalState":"null",'
        if 'CauseOfDeath' in kwargs:
            tgjc += '"CauseOfDeath":"' + kwargs['CauseOfDeath'] + '"}'
        else:
            tgjc += '"CauseOfDeath":"null"}'
        masfoll.sur_sta = tgjc
        masfoll.del_flg = "1"
        result = masfoll
    except:
        result = False
    return result
        

# 修改或添加E路童萌
def modifyorAddEltm(casePk=0, kwargs=0):
    try:
        if 'queryId' in kwargs and len(kwargs['queryId']) > 0:
            eltm = models.SzfyEltm.objects.get(patient__pk=decode_id(kwargs['queryId']))
        else:
            eltm = models.SzfyEltm()
        # Tanner分期
        if 'Tanner' in kwargs:
            eltm.tanner = kwargs['Tanner']
        eltm.patient_id = casePk
        # 发生时间
        if 'eventTime' in kwargs:
            eltm.star_time = kwargs['eventTime']
        # 结束时间
        if 'end_time' in kwargs:
            eltm.endTime = kwargs['end_time']
        # 是否为严重不良事件
        if 'isAdEvent' in kwargs:
            eltm.is_adv_eve = kwargs['isAdEvent']
        # 与研究药物的关系(LA-rhGH)
        if 'larhGH' in kwargs:
            eltm.la_rhGH = kwargs['larhGH']
        # 是否调整剂量
        if 'isAdjust' in kwargs:
            eltm.is_adjust = kwargs['isAdjust']
        # 与研究药物的关系(rhGH)
        if 'isRhGH' in kwargs:
            eltm.rhGH = kwargs['isRhGH']
        # 不良事件的转归
        if 'outcome' in kwargs:
            eltm.outcome = kwargs['outcome']
        # 药物名称
        if 'medicationName' in kwargs:
            eltm.med_name = kwargs['medicationName']
        # 单次剂量
        if 'dose' in kwargs:
            eltm.dose = kwargs['dose']
        # 用药天数
        if 'days' in kwargs:
            eltm.days = kwargs['days']
        # 是否停药
        if 'stopMedication' in kwargs:
            eltm.stop_med = kwargs['stopMedication']
        # 停药原因
        if 'stopReason' in kwargs:
            eltm.stop_rea = kwargs['stopReason']
        # 记录日期
        if 'recordDate' in kwargs:
            eltm.rec_date = kwargs['recordDate']
        # 有无既往用药史
        if 'hasHistory' in kwargs:
            eltm.is_has_his = kwargs['hasHistory']
        # 用药史
        if 'sampleBankMed' in kwargs:
            eltm.has_his = json.dumps(kwargs['sampleBankMed'])
        # 基因检测方法
        if 'geneMethod' in kwargs:
            eltm.gene_method = kwargs['geneMethod']
        # 基因结果
        if 'geneRes' in kwargs:
            eltm.gene_res = kwargs['geneRes']
        # 基因名称
        if 'geneName' in kwargs:
            eltm.gene_name = kwargs['geneName']
        # 突变位点
        if 'genePoint' in kwargs:
            eltm.gene_point = kwargs['genePoint']
        # 突变类型
        if 'geneType' in kwargs:
            eltm.gene_type = kwargs['geneType']
        # 遗传模式
        if 'geneMode' in kwargs:
            eltm.gene_mode = kwargs['geneMode']
        # 染色体核型
        if 'chromosom' in kwargs:
            eltm.chrom = kwargs['chromosom']
        # 其它异常核型
        if 'chromosomOther' in kwargs:
            eltm.chrom_other = kwargs['chromosomOther']
        # 一般症状
        if 'generalSymptoms' in kwargs:
            eltm.gen_sym = kwargs['generalSymptoms']
        # 代谢相关症状
        if 'metabolicSymptoms' in kwargs:
            eltm.met_sym = kwargs['metabolicSymptoms']
        # 骨骼和肌肉症状
        if 'boneMuscleSymptoms' in kwargs:
            eltm.bone_sym = kwargs['boneMuscleSymptoms']
        # 内分泌症状
        if 'endocrineSymptoms' in kwargs:
            eltm.endo_sym = kwargs['endocrineSymptoms']
        # 其他症状
        if 'otherSymptoms' in kwargs:
            eltm.other_sym = kwargs['otherSymptoms']
        result = eltm
    except:
        result = False
    return result

class ImageView(FormattedView):
    @require_arguments(['queryId', 'organ', 'path'])
    def get(self, request, *args, **kwargs):
        """
        用于图片下载，判断登录权限并获取到必要参数后交由Nignx处理
        """

        queryId = decode_id(kwargs['queryId'])
        organ = kwargs['organ']
        path = os.path.join(organ, str(queryId % 64), str(queryId), kwargs['path'])
        path = path.replace('-', '/')

        # 转发给Nginx处理
        url = '/protected_files/{}'.format(path)

        # print('image:', url)
        response = HttpResponse('')
        response['X-Accel-Redirect'] = url.encode()

        if 'type' not in kwargs:
            response['Content-Type'] = 'application/octet-stream'
        elif kwargs['type'] == 'mp4':
            response['Content-Type'] = 'video/mp4'

        return response

    @require_arguments(['queryId', 'organ', 'path'], 'body')
    def post(self, request, *args, **kwargs):
        """
        图片上传
        """

        queryId = decode_id(kwargs['queryId'])
        data = request.FILES.get('package', None)
        if data is None:
            return self.make_response(None, code=Code.MISSING_REQUIRED_ARGUMENTS)
        # ret = save_img(queryId, kwargs['organ'], kwargs['path'], data, kwargs)
        ret = save_img(queryId, kwargs['organ'], kwargs['path'], data)
        if ret:
            return self.make_response(ret)
        else:
            return self.make_response(None, code=Code.FAIL_SAVIMG)

    @require_arguments(['queryId', 'organ', 'path'])
    def delete(self, request, *args, **kwargs):
        """
        图片删除
        """

        queryId = decode_id(kwargs['queryId'])
        save_img(queryId, kwargs['organ'], kwargs['path'], None)

        return self.make_response('ok')

class PatientNumView(FormattedView):
    extractor = extractors.AllExtractor()
    # 查询患者详细数据根据患者编号模糊查询
    @require_arguments(['userNum'], 'url')
    def get(self, request, *args, **kwargs):
        try:
            patient = models.Patient.objects.filter(user_num__contains=kwargs['userNum'], dis_class=kwargs['disClass']).values('user_num','name','sex','birth_time')
            return self.make_response(list(patient))
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

class PatientNumJqView(FormattedView):
    extractor = extractors.AllExtractor()
    # 查询患者详细数据根据患者编号精确查询
    @require_arguments(['userNum'], 'url')
    def get(self, request, *args, **kwargs):
        try:
            patient = models.Patient.objects.filter(user_num=kwargs['userNum'])[0]
            return self.make_response(patient)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

class FollowListView(FormattedView):
    extractor = extractors.AllExtractor()

    # 查询随访记录
    @parse_arguments('url')
    def get(self, request, *args, **kwargs):
        patFoll = models.PatFoll.objects.filter(patient__pk=decode_id(kwargs['queryId']), del_flg='1').all()
        limit = kwargs['limit']
        paginator = Paginator(patFoll, limit)  # 每页显示10条
        page = kwargs['currPage']
        if page == '0':
            page = '1'
        pagedata = {}  # 获取分页信息
        pagedata['count'] = paginator.count
        pagedata['num_pages'] = paginator.num_pages
        pagedata['per_page'] = limit
        pagedata['current'] = page
        context = {}
        list = paginator.page(page).object_list
        contacts = self.extractor.extract(list)
        context['contacts'] = contacts
        context['pagedata'] = pagedata
        return self.make_response(context)

class FollowView(FormattedView):
    extractor = extractors.AllExtractor()
    # 查询随访详细数据
    @require_arguments(['queryPId'], 'url')
    def get(self, request, *args, **kwargs):
        # 主键id
        if len(kwargs['queryPId']) > 0:
            patFoll_id = decode_id(kwargs['queryPId'])
        else:
            return self.make_response(None, Code.MAIN_NULL)
        try:
            patFoll = models.PatFoll.objects.get(pk=patFoll_id)
            return self.make_response(patFoll)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

    # 添加随访记录
    @require_arguments(['queryId'], 'body')
    def put(self, request, *args, **kwargs):
        try:
            patient = models.Patient.objects.get(pk=decode_id(kwargs['queryId']))
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)
        if 'queryPId' in kwargs and len(kwargs['queryPId']) > 0:
            patFoll = models.PatFoll.objects.get(pk=decode_id(kwargs['queryPId']))
        else:
            patFoll = models.PatFoll()
        # 病例主表id
        patFoll.patient_id = patient.pk
        # 随访日期
        if 'followTime' in kwargs and kwargs['followTime'] is not None and len(kwargs['followTime']):
            time = kwargs['followTime'][0:10]
            patFoll.foll_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        #  现身高
        if 'Ht' in kwargs:
            patFoll.Ht = kwargs['Ht']
        #  年龄
        if 'age' in kwargs:
            patFoll.age = kwargs['age']
        #  现体重
        if 'Wt' in kwargs:
            patFoll.Wt = kwargs['Wt']
        #  是否行为发育评估
        if 'isFYassess' in kwargs:
            patFoll.beh_dev_ass = kwargs['isFYassess']
        #  Peabody运动发育评估
        if 'Peabody' in kwargs:
            patFoll.ped_mot_dev_ass = kwargs['Peabody']
        #  粗大运动
        if 'GriffithsA' in kwargs:
            patFoll.gro_mot = kwargs['GriffithsA']
        #  个人社会
        if 'GriffithsB' in kwargs:
            patFoll.ind_soc = kwargs['GriffithsB']
        #  听力语言
        if 'GriffithsC' in kwargs:
            patFoll.lis_lan = kwargs['GriffithsC']
        #  手眼协调
        if 'GriffithsD' in kwargs:
            patFoll.han_eye_coo = kwargs['GriffithsD']
        #  视觉表现
        if 'GriffithsE' in kwargs:
            patFoll.vis_rep = kwargs['GriffithsE']
        #  实际推理
        if 'GriffithsF' in kwargs:
            patFoll.pra_rea = kwargs['GriffithsF']
        #  韦氏智力量表
        if 'wszlb' in kwargs:
            patFoll.wec_sca = kwargs['wszlb']
        #  双乳生殖器分期
        if 'breastDev' in kwargs:
            patFoll.gen_stag = kwargs['breastDev']
        #  外生殖器分期
        if 'genStag' in kwargs:
            patFoll.gen_stag = kwargs['genStag']
        #  阴毛分期
        if 'pubStag' in kwargs:
            patFoll.pub_stag = kwargs['pubStag']
        #  IGF-1（ng/ml）
        if 'IGF1' in kwargs:
            patFoll.IGF1 = kwargs['IGF1']
        #  IGFBP-3（ug/ml）
        if 'IGFBP3' in kwargs:
            patFoll.IGFBP3 = kwargs['IGFBP3']
        #  甲功
        if 'Jiagong' in kwargs:
            jg = '{"Jiagong":"' + kwargs['Jiagong'] + '",'
        else:
            jg = '{"Jiagong":"null",'
        if 'JiagongDes' in kwargs:
            jg += '"JiagongDes":"' + kwargs['JiagongDes'] + '"}'
        else:
            jg += '"JiagongDes":"null"}'
        patFoll.Jiagong = jg
        #  空腹血糖
        if 'fasBloodGlu' in kwargs:
            patFoll.fas_blood_glu = kwargs['fasBloodGlu']
        #  空腹胰岛素
        if 'fasInsulin' in kwargs:
            patFoll.fas_insulin = kwargs['fasInsulin']
        #  肝肾脂电解质
        if 'livKidLip' in kwargs:
            gz = '{"livKidLip":"' + kwargs['livKidLip'] + '",'
        else:
            gz = '{"livKidLip":"null",'
        if 'LAKLEdes' in kwargs:
            gz += '"LAKLEdes":"' + kwargs['LAKLEdes'] + '"}'
        else:
            gz += '"LAKLEdes":"null"}'
        patFoll.liv_kid_lip = gz
        #  糖化血红蛋白
        if 'glyHem' in kwargs:
            patFoll.gly_hem = kwargs['glyHem']
        #  LH
        if 'LH' in kwargs:
            patFoll.LH = kwargs['LH']
        #  FSH
        if 'FSH' in kwargs:
            patFoll.FSH = kwargs['FSH']
        #  E2
        if 'E2' in kwargs:
            patFoll.E2 = kwargs['E2']
        #  T
        if 'T' in kwargs:
            patFoll.T = kwargs['T']
        #  DHT
        if 'DHT' in kwargs:
            patFoll.DHT = kwargs['DHT']
        #  游离睾酮
        if 'FT' in kwargs:
            patFoll.yltg = kwargs['FT']
        #  SHBG
        if 'SHBG' in kwargs:
            patFoll.SHBG = kwargs['SHBG']
        #  性腺B超
        if 'uterusOne' in kwargs:
            xxbc = '{"uterusOne":"' + kwargs['uterusOne'] + '",'
        else:
            xxbc = '{"uterusOne":"null",'
        if 'uterusTwo' in kwargs:
            xxbc += '"uterusTwo":"' + kwargs['uterusTwo'] + '",'
        else:
            xxbc += '"uterusTwo":"null",'
        if 'uterusThr' in kwargs:
            xxbc += '"uterusThr":"' + kwargs['uterusThr'] + '",'
        else:
            xxbc += '"uterusThr":"null",'
        if 'cervixLong' in kwargs:
            xxbc += '"cervixLong":"' + kwargs['cervixLong'] + '",'
        else:
            xxbc += '"cervixLong":"null",'
        if 'intima' in kwargs:
            xxbc += '"intima":"' + kwargs['intima'] + '",'
        else:
            xxbc += '"intima":"null",'
        if 'ovaLeftOne' in kwargs:
            xxbc += '"ovaLeftOne":"' + kwargs['ovaLeftOne'] + '",'
        else:
            xxbc += '"ovaLeftOne":"null",'
        if 'ovaLeftTwo' in kwargs:
            xxbc += '"ovaLeftTwo":"' + kwargs['ovaLeftTwo'] + '",'
        else:
            xxbc += '"ovaLeftTwo":"null",'
        if 'ovaLeftThr' in kwargs:
            xxbc += '"ovaLeftThr":"' + kwargs['ovaLeftThr'] + '",'
        else:
            xxbc += '"ovaLeftThr":"null",'
        if 'ovaRightOne' in kwargs:
            xxbc += '"ovaRightOne":"' + kwargs['ovaRightOne'] + '",'
        else:
            xxbc += '"ovaRightOne":"null",'
        if 'ovaRightTwo' in kwargs:
            xxbc += '"ovaRightTwo":"' + kwargs['ovaRightTwo'] + '",'
        else:
            xxbc += '"ovaRightTwo":"null",'
        if 'ovaRightThr' in kwargs:
            xxbc += '"ovaRightThr":"' + kwargs['ovaRightThr'] + '",'
        else:
            xxbc += '"ovaRightThr":"null",'
        if 'follDiameter' in kwargs:
            xxbc += '"follDiameter":"' + kwargs['follDiameter'] + '",'
        else:
            xxbc += '"follDiameter":"null",'
        if 'isCyst' in kwargs:
            xxbc += '"isCyst":"' + kwargs['isCyst'] + '",'
        else:
            xxbc += '"isCyst":"null",'
        if 'cyst' in kwargs:
            xxbc += '"cyst":"' + kwargs['cyst'] + '",'
        else:
            xxbc += '"cyst":"null",'
        if 'cystOne' in kwargs:
            xxbc += '"cystOne":"' + kwargs['cystOne'] + '",'
        else:
            xxbc += '"cystOne":"null",'
        if 'cystTwo' in kwargs:
            xxbc += '"cystTwo":"' + kwargs['cystTwo'] + '",'
        else:
            xxbc += '"cystTwo":"null",'
        if 'cystThr' in kwargs:
            xxbc += '"cystThr":"' + kwargs['cystThr'] + '",'
        else:
            xxbc += '"cystThr":"null",'
        if 'cystDescribe' in kwargs:
            xxbc += '"cystDescribe":"' + kwargs['cystDescribe'] + '",'
        else:
            xxbc += '"cystDescribe":"null",'
        if 'testisLeftOne' in kwargs:
            xxbc += '"testisLeftOne":"' + kwargs['testisLeftOne'] + '",'
        else:
            xxbc += '"testisLeftOne":"null",'
        if 'testisLeftTwo' in kwargs:
            xxbc += '"testisLeftTwo":"' + kwargs['testisLeftTwo'] + '",'
        else:
            xxbc += '"testisLeftTwo":"null",'
        if 'testisLeftThr' in kwargs:
            xxbc += '"testisLeftThr":"' + kwargs['testisLeftThr'] + '",'
        else:
            xxbc += '"testisLeftThr":"null",'
        if 'testisLeftLon' in kwargs:
            xxbc += '"testisLeftLon":"' + kwargs['testisLeftLon'] + '",'
        else:
            xxbc += '"testisLeftLon":"null",'
        if 'testisRightOne' in kwargs:
            xxbc += '"testisRightOne":"' + kwargs['testisRightOne'] + '",'
        else:
            xxbc += '"testisRightOne":"null",'
        if 'testisRightTwo' in kwargs:
            xxbc += '"testisRightTwo":"' + kwargs['testisRightTwo'] + '",'
        else:
            xxbc += '"testisRightTwo":"null",'
        if 'testisRightThr' in kwargs:
            xxbc += '"testisRightThr":"' + kwargs['testisRightThr'] + '",'
        else:
            xxbc += '"testisRightThr":"null",'
        if 'MRI' in kwargs:
            xxbc += '"MRI":"' + kwargs['MRI'] + '",'
        else:
            xxbc += '"MRI":"null",'
        if 'mriDescribe' in kwargs:
            xxbc += '"mriDescribe":"' + kwargs['mriDescribe'] + '",'
        else:
            xxbc += '"mriDescribe":"null",'
        if 'testisRightLon' in kwargs:
            xxbc += '"testisRightLon":"' + kwargs['testisRightLon'] + '"}'
        else:
            xxbc += '"testisRightLon":"null"}'
        patFoll.gon_B_ult = xxbc
        #  诊疗方案
        if 'diaPlan' in kwargs:
            zlfa = '{"diaPlan":"' + kwargs['diaPlan'] + '",'
        else:
            zlfa = '{"diaPlan":"null",'
        if 'rhGH' in kwargs:
            zlfa += '"rhGH":"' + kwargs['rhGH'] + '",'
        else:
            zlfa += '"rhGH":"null",'
        if 'rhCustomizationDiaPlan' in kwargs:
            zlfa += '"rhCustomizationDiaPlan":"' + kwargs['rhCustomizationDiaPlan'] + '",'
        else:
            zlfa += '"rhCustomizationDiaPlan":"null",'
        if 'rhCustomizationPrompt' in kwargs:
            zlfa += '"rhCustomizationPrompt":"' + kwargs['rhCustomizationPrompt'] + '",'
        else:
            zlfa += '"rhCustomizationPrompt":"null",'
        if 'rhUnitedCustomization' in kwargs:
            zlfa += '"rhUnitedCustomization":"' + kwargs['rhUnitedCustomization'] + '",'
        else:
            zlfa += '"rhUnitedCustomization":"null",'
        if 'rhUnitedDose' in kwargs:
            zlfa += '"rhUnitedDose":"' + kwargs['rhUnitedDose'] + '",'
        else:
            zlfa += '"rhUnitedDose":"null",'
        if 'genData' in kwargs:
            zlfa += '"genData":"' + json.dumps(kwargs['genData']) + '",'
        else:
            zlfa += '"genData":"null",'
        if 'otherMedicine' in kwargs:
            zlfa += '"otherMedicine":"' + json.dumps(kwargs['otherMedicine']) + '",'
        else:
            zlfa += '"otherMedicine":"null",'
        if 'rhGHdoseKG' in kwargs:
            zlfa += '"rhGHdoseKG":"' +  kwargs['rhGHdoseKG'] + '",'
        else:
            zlfa += '"rhGHdoseKG":"null",'
        if 'PEGrhGHdose' in kwargs:
            zlfa += '"PEGrhGHdose":"' +  kwargs['PEGrhGHdose'] + '",'
        else:
            zlfa += '"PEGrhGHdose":"null",'
        if 'rhCustomizationDiaPlan' in kwargs:
            zlfa += '"rhCustomizationDiaPlan":"' +  kwargs['rhCustomizationDiaPlan'] + '",'
        else:
            zlfa += '"rhCustomizationDiaPlan":"null",'
        if 'rhCustomizationPrompt' in kwargs:
            zlfa += '"rhCustomizationPrompt":"' +  kwargs['rhCustomizationPrompt'] + '",'
        else:
            zlfa += '"rhCustomizationPrompt":"null",'
        if 'PEGrhCustomizationPrompt' in kwargs:
            zlfa += '"PEGrhCustomizationPrompt":"' +  kwargs['PEGrhCustomizationPrompt'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPrompt":"null",'
        if 'rhCustomizationPromptKG' in kwargs:
            zlfa += '"rhCustomizationPromptKG":"' +  kwargs['rhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"rhCustomizationPromptKG":"null",'
        if 'PEGrhCustomizationPromptKG' in kwargs:
            zlfa += '"PEGrhCustomizationPromptKG":"' +  kwargs['PEGrhCustomizationPromptKG'] + '",'
        else:
            zlfa += '"PEGrhCustomizationPromptKG":"null",'
        if 'PEGrhGHdoseKG' in kwargs:
            zlfa += '"PEGrhGHdoseKG":"' +  kwargs['PEGrhGHdoseKG'] + '",'
        else:
            zlfa += '"PEGrhGHdoseKG":"null",'
        if 'planData' in kwargs:
            zlfa += '"planData":"' +  json.dumps(kwargs['planData']) + '",'
        else:
            zlfa += '"planData":"null",'
        if 'rhGHdose' in kwargs:
            zlfa += '"rhGHdose":"' + kwargs['rhGHdose'] + '"}'
        else:
            zlfa += '"rhGHdose":"null"}'
        patFoll.dia_trea_plan = zlfa
        # 其他
        if 'other' in kwargs:
            patFoll.other = kwargs['other']
        #  实验室检查以上没有字段都在这
        if 'LHFSHTime' in kwargs:
            labexaother = '{"LHFSHTime":"' + kwargs['LHFSHTime'] + '",'
        else:
            labexaother = '{"LHFSHTime":"null",'
        if 'E2Time' in kwargs:
            labexaother += '"E2Time":"' + kwargs['E2Time'] + '",'
        else:
            labexaother += '"E2Time":"null",'
        if 'TTime' in kwargs:
            labexaother += '"TTime":"' + kwargs['TTime'] + '",'
        else:
            labexaother += '"TTime":"null",'
        if 'PRL' in kwargs:
            labexaother += '"PRL":"' + kwargs['PRL'] + '",'
        else:
            labexaother += '"PRL":"null",'
        if 'PRLTime' in kwargs:
            labexaother += '"PRLTime":"' + kwargs['PRLTime'] + '",'
        else:
            labexaother += '"PRLTime":"null",'
        if 'IGFBPTime' in kwargs:
            labexaother += '"IGFBPTime":"' + kwargs['IGFBPTime'] + '",'
        else:
            labexaother += '"IGFBPTime":"null",'
        if 'thyroidTime' in kwargs:
            labexaother += '"thyroidTime":"' + kwargs['thyroidTime'] + '",'
        else:
            labexaother += '"thyroidTime":"null",'
        if 'ACTH' in kwargs:
            labexaother += '"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            labexaother += '"ACTH":"null",'
        if 'ACTHTime' in kwargs:
            labexaother += '"ACTHTime":"' + kwargs['ACTHTime'] + '",'
        else:
            labexaother += '"ACTHTime":"null",'
        if 'cortisol' in kwargs:
            labexaother += '"cortisol":"' + kwargs['cortisol'] + '",'
        else:
            labexaother += '"cortisol":"null",'
        if 'cortisolTime' in kwargs:
            labexaother += '"cortisolTime":"' + kwargs['cortisolTime'] + '",'
        else:
            labexaother += '"cortisolTime":"null",'
        if 'DHEAS' in kwargs:
            labexaother += '"DHEAS":"' + kwargs['DHEAS'] + '",'
        else:
            labexaother += '"DHEAS":"null",'
        if 'DHEATime' in kwargs:
            labexaother += '"DHEATime":"' + kwargs['DHEATime'] + '",'
        else:
            labexaother += '"DHEATime":"null",'
        if 'OHP' in kwargs:
            labexaother += '"OHP":"' + kwargs['OHP'] + '",'
        else:
            labexaother += '"OHP":"null",'
        if 'OHPTime' in kwargs:
            labexaother += '"OHPTime":"' + kwargs['OHPTime'] + '",'
        else:
            labexaother += '"OHPTime":"null",'
        if 'blood' in kwargs:
            labexaother += '"blood":"' + kwargs['blood'] + '",'
        else:
            labexaother += '"blood":"null",'
        if 'bloodDescribe' in kwargs:
            labexaother += '"bloodDescribe":"' + kwargs['bloodDescribe'] + '",'
        else:
            labexaother += '"bloodDescribe":"null",'
        if 'bloodTime' in kwargs:
            labexaother += '"bloodTime":"' + kwargs['bloodTime'] + '",'
        else:
            labexaother += '"bloodTime":"null",'
        if 'urinalysis' in kwargs:
            labexaother += '"urinalysis":"' + kwargs['urinalysis'] + '",'
        else:
            labexaother += '"urinalysis":"null",'
        if 'urinalysisDescribe' in kwargs:
            labexaother += '"urinalysisDescribe":"' + kwargs['urinalysisDescribe'] + '",'
        else:
            labexaother += '"urinalysisDescribe":"null",'
        if 'urinalysisTime' in kwargs:
            labexaother += '"urinalysisTime":"' + kwargs['urinalysisTime'] + '",'
        else:
            labexaother += '"urinalysisTime":"null",'
        if 'LAKLGETime' in kwargs:
            labexaother += '"LAKLGETime":"' + kwargs['LAKLGETime'] + '",'
        else:
            labexaother += '"LAKLGETime":"null",'
        if 'HBs' in kwargs:
            labexaother += '"HBs":"' + kwargs['HBs'] + '",'
        else:
            labexaother += '"HBs":"null",'
        if 'HBsTime' in kwargs:
            labexaother += '"HBsTime":"' + kwargs['HBsTime'] + '",'
        else:
            labexaother += '"HBsTime":"null",'
        if 'gh' in kwargs:
            labexaother += '"gh":"' + kwargs['gh'] + '",'
        else:
            labexaother += '"gh":"null",'
        if 'glyHemA' in kwargs:
            labexaother += '"glyHemA":"' + kwargs['glyHemA'] + '",'
        else:
            labexaother += '"glyHemA":"null",'
        if 'glyHemATime' in kwargs:
            labexaother += '"glyHemATime":"' + kwargs['glyHemATime'] + '",'
        else:
            labexaother += '"glyHemATime":"null",'
        if 'fasBloodGluTime' in kwargs:
            labexaother += '"fasBloodGluTime":"' + kwargs['fasBloodGluTime'] + '",'
        else:
            labexaother += '"fasBloodGluTime":"null",'
        if 'fasInsulinTime' in kwargs:
            labexaother += '"fasInsulinTime":"' + kwargs['fasInsulinTime'] + '",'
        else:
            labexaother += '"fasInsulinTime":"null",'
        if 'glyHemTime' in kwargs:
            labexaother += '"glyHemTime":"' + kwargs['glyHemTime'] + '",'
        else:
            labexaother += '"glyHemTime":"null",'
        if 'ghTime' in kwargs:
            labexaother += '"ghTime":"' + kwargs['ghTime'] + '"}'
        else:
            labexaother += '"ghTime":"null"}'
        patFoll.lab_exa_other = labexaother
        # 疾病
        if 'disease' in kwargs:
            patFoll.disease = kwargs['disease']
        # 地舒单抗
        if 'denosumab' in kwargs:
            patFoll.dsdk = kwargs['denosumab']
        # 唑来膦酸
        if 'zoledronate' in kwargs:
            patFoll.clls = kwargs['zoledronate']
        # 其他用量
        if 'otherUsages' in kwargs:
            patFoll.qtyl = kwargs['otherUsages']
        # 其他检查
        if 'other' in kwargs:
            patFoll.other_exam = kwargs['other']
        # EOS
        if 'EOS' in kwargs:
            patFoll.eos = kwargs['EOS']
        # 骨密度
        if 'boneDensity' in kwargs:
            patFoll.bon_min_den = kwargs['boneDensity']
        #  实验室检查mas
        if 'ACTH' in kwargs:
            labexamas = '{"ACTH":"' + kwargs['ACTH'] + '",'
        else:
            labexamas = '{"ACTH":"null",'
        if 'prealbumin' in kwargs:
            labexamas += '"prealbumin":"' + kwargs['prealbumin'] + '",'
        else:
            labexamas += '"prealbumin":"null",'
        if 'ALT' in kwargs:
            labexamas += '"ALT":"' + kwargs['ALT'] + '",'
        else:
            labexamas += '"ALT":"null",'
        if 'asparagine' in kwargs:
            labexamas += '"asparagine":"' + kwargs['asparagine'] + '",'
        else:
            labexamas += '"asparagine":"null",'
        if 'ALKP' in kwargs:
            labexamas += '"ALKP":"' + kwargs['ALKP'] + '",'
        else:
            labexamas += '"ALKP":"null",'
        if 'glutamyl' in kwargs:
            labexamas += '"glutamyl":"' + kwargs['glutamyl'] + '",'
        else:
            labexamas += '"glutamyl":"null",'
        if 'TBIL' in kwargs:
            labexamas += '"TBIL":"' + kwargs['TBIL'] + '",'
        else:
            labexamas += '"TBIL":"null",'
        if 'DBIL' in kwargs:
            labexamas += '"DBIL":"' + kwargs['DBIL'] + '",'
        else:
            labexamas += '"DBIL":"null",'
        if 'TP' in kwargs:
            labexamas += '"TP":"' + kwargs['TP'] + '",'
        else:
            labexamas += '"TP":"null",'
        if 'ricim' in kwargs:
            labexamas += '"ricim":"' + kwargs['ricim'] + '",'
        else:
            labexamas += '"ricim":"null",'
        if 'WBCPer' in kwargs:
            labexamas += '"WBCPer":"' + kwargs['WBCPer'] + '",'
        else:
            labexamas += '"WBCPer":"null",'
        if 'bileAcid' in kwargs:
            labexamas += '"bileAcid":"' + kwargs['bileAcid'] + '",'
        else:
            labexamas += '"bileAcid":"null",'
        if 'urea' in kwargs:
            labexamas += '"urea":"' + kwargs['urea'] + '",'
        else:
            labexamas += '"urea":"null",'
        if 'creatinine' in kwargs:
            labexamas += '"creatinine":"' + kwargs['creatinine'] + '",'
        else:
            labexamas += '"creatinine":"null",'
        if 'uricAcid' in kwargs:
            labexamas += '"uricAcid":"' + kwargs['uricAcid'] + '",'
        else:
            labexamas += '"uricAcid":"null",'
        if 'Na' in kwargs:
            labexamas += '"Na":"' + kwargs['Na'] + '",'
        else:
            labexamas += '"Na":"null",'
        if 'potassium' in kwargs:
            labexamas += '"potassium":"' + kwargs['potassium'] + '",'
        else:
            labexamas += '"potassium":"null",'
        if 'CI' in kwargs:
            labexamas += '"CI":"' + kwargs['CI'] + '",'
        else:
            labexamas += '"CI":"null",'
        if 'carbonDioxide' in kwargs:
            labexamas += '"carbonDioxide":"' + kwargs['carbonDioxide'] + '",'
        else:
            labexamas += '"carbonDioxide":"null",'
        if 'Ca' in kwargs:
            labexamas += '"Ca":"' + kwargs['Ca'] + '",'
        else:
            labexamas += '"Ca":"null",'
        if 'phosphorus' in kwargs:
            labexamas += '"phosphorus":"' + kwargs['phosphorus'] + '",'
        else:
            labexamas += '"phosphorus":"null",'
        if 'serumMa' in kwargs:
            labexamas += '"serumMa":"' + kwargs['serumMa'] + '",'
        else:
            labexamas += '"serumMa":"null",'
        if 'CyC' in kwargs:
            labexamas += '"CyC":"' + kwargs['CyC'] + '",'
        else:
            labexamas += '"CyC":"null",'
        if 'egFiltration' in kwargs:
            labexamas += '"egFiltration":"' + kwargs['egFiltration'] + '",'
        else:
            labexamas += '"egFiltration":"null",'
        if 'ExPeptide' in kwargs:
            labexamas += '"ExPeptide":"' + kwargs['ExPeptide'] + '",'
        else:
            labexamas += '"ExPeptide":"null",'
        if 'ColOrder' in kwargs:
            labexamas += '"ColOrder":"' + kwargs['ColOrder'] + '",'
        else:
            labexamas += '"ColOrder":"null",'
        if 'osteocalcin' in kwargs:
            labexamas += '"osteocalcin":"' + kwargs['osteocalcin'] + '",'
        else:
            labexamas += '"osteocalcin":"null",'
        if 'insulinGF' in kwargs:
            labexamas += '"insulinGF":"' + kwargs['insulinGF'] + '",'
        else:
            labexamas += '"insulinGF":"null",'
        if 'insulinGFBProtein' in kwargs:
            labexamas += '"insulinGFBProtein":"' + kwargs['insulinGFBProtein'] + '",'
        else:
            labexamas += '"insulinGFBProtein":"null",'
        if 'FreeFT3' in kwargs:
            labexamas += '"FreeFT3":"' + kwargs['FreeFT3'] + '",'
        else:
            labexamas += '"FreeFT3":"null",'
        if 'FreeFT4' in kwargs:
            labexamas += '"FreeFT4":"' + kwargs['FreeFT4'] + '",'
        else:
            labexamas += '"FreeFT4":"null",'
        if 'ThyroidTSH' in kwargs:
            labexamas += '"ThyroidTSH":"' + kwargs['ThyroidTSH'] + '",'
        else:
            labexamas += '"ThyroidTSH":"null",'
        if 'gonadotropin' in kwargs:
            labexamas += '"gonadotropin":"' + kwargs['gonadotropin'] + '",'
        else:
            labexamas += '"gonadotropin":"null",'
        if 'hydroxylD' in kwargs:
            labexamas += '"hydroxylD":"' + kwargs['hydroxylD'] + '",'
        else:
            labexamas += '"hydroxylD":"null",'
        if 'parathyroidPTH' in kwargs:
            labexamas += '"parathyroidPTH":"' + kwargs['parathyroidPTH'] + '",'
        else:
            labexamas += '"parathyroidPTH":"null",'
        if 'bloodCortisol' in kwargs:
            labexamas += '"bloodCortisol":"' + kwargs['bloodCortisol'] + '",'
        else:
            labexamas += '"bloodCortisol":"null",'
        if 'TRAb' in kwargs:
            labexamas += '"TRAb":"' + kwargs['TRAb'] + '",'
        else:
            labexamas += '"TRAb":"null",'
        if 'TPOAb' in kwargs:
            labexamas += '"TPOAb":"' + kwargs['TPOAb'] + '",'
        else:
            labexamas += '"TPOAb":"null",'
        if 'calcitoninCT' in kwargs:
            labexamas += '"calcitoninCT":"' + kwargs['calcitoninCT'] + '",'
        else:
            labexamas += '"calcitoninCT":"null",'
        if 'thyrotropinTSI' in kwargs:
            labexamas += '"thyrotropinTSI":"' + kwargs['thyrotropinTSI'] + '",'
        else:
            labexamas += '"thyrotropinTSI":"null",'
        if 'thyroidTBG' in kwargs:
            labexamas += '"thyroidTBG":"' + kwargs['thyroidTBG'] + '",'
        else:
            labexamas += '"thyroidTBG":"null",'
        if 'LH' in kwargs:
            labexamas += '"LH":"' + kwargs['LH'] + '",'
        else:
            labexamas += '"LH":"null",'
        if 'FSH' in kwargs:
            labexamas += '"FSH":"' + kwargs['FSH'] + '",'
        else:
            labexamas += '"FSH":"null",'
        if 'PRL' in kwargs:
            labexamas += '"PRL":"' + kwargs['PRL'] + '",'
        else:
            labexamas += '"PRL":"null",'
        if 'E2' in kwargs:
            labexamas += '"E2":"' + kwargs['E2'] + '",'
        else:
            labexamas += '"E2":"null",'
        if 'P' in kwargs:
            labexamas += '"P":"' + kwargs['P'] + '",'
        else:
            labexamas += '"P":"null",'
        if 'T' in kwargs:
            labexamas += '"T":"' + kwargs['T'] + '",'
        else:
            labexamas += '"T":"null",'
        if 'sexHormone' in kwargs:
            labexamas += '"sexHormone":"' + kwargs['sexHormone'] + '"}'
        else:
            labexamas += '"sexHormone":"null"}'
        patFoll.lab_exa_mas = labexamas
        #  是否终身高
        if 'isFinalHeight' in kwargs:
            patFoll.is_finalhei = kwargs['isFinalHeight']
        #  诊疗方案其他字段
        if 'otherMedicine' in kwargs:
            patFoll.otherMedicine = kwargs['otherMedicine']
        # 其他图片名称
        if 'otherImageNames' in kwargs:
            patFoll.other_ima_name = kwargs['otherImageNames']
        try:
            patFoll.save()
            return self.make_response(patFoll, extractor=self.extractor)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

    # 删除病例详细数据
    @require_arguments(['queryPId'], 'url')
    def delete(self, request, *args, **kwargs):
        # 主键id
        if len(kwargs['queryPId']) > 0:
            case_id = decode_id(kwargs['queryPId'])
        else:
            return self.make_response(None, Code.MAIN_NULL)
        try:
            patFoll = models.PatFoll.objects.get(pk=case_id, del_flg='1')
            patFoll.del_flg = '0'
            patFoll.save()
            return self.make_response(None)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

class MasFollowView(FormattedView):
    extractor = extractors.AllExtractor()
    # 查询Mas随访详细数据
    @require_arguments(['queryMId'], 'url')
    def get(self, request, *args, **kwargs):
        # 主键id
        if len(kwargs['queryMId']) > 0:
            mas_id = decode_id(kwargs['queryMId'])
        else:
            return self.make_response(None, Code.MAIN_NULL)
        try:
            masFoll = models.MasFoll.objects.get(mas__pk = mas_id)
            return self.make_response(masFoll)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

class FollowListNoView(FormattedView):
    extractor = extractors.AllExtractor()

    # 查询随访记录不分页
    @parse_arguments('url')
    def get(self, request, *args, **kwargs):
        patFoll = models.PatFoll.objects.filter(patient__pk=decode_id(kwargs['queryId']), del_flg='1').all().order_by('foll_time')
        return self.make_response(patFoll)

# # 下载文件
# class loadFile(FormattedView):
#     # 请求导出
#     @parse_arguments('body')
#     def put(self, request, *args, **kwargs):
#
#         file_info = self.getExcel(kwargs)
#         if file_info:
#             return self.make_response(file_info)
#         else:
#             return self.make_response(None, Code.DEFULT_SAVE_ZIP)
#
#     # 下载Excel
#     def get(self, request, *args, **kwargs):
#
#         file_path = kwargs['organ'] + '/' + str(int(kwargs['queryId']) % 64) + "/" + kwargs['queryId'] + '/' + kwargs['path']+ '/' + kwargs['filename']
#         url = '/protected_files/{}'.format(file_path.replace(',', '/'))
#         response = HttpResponse('')
#         response['X-Accel-Redirect'] = url.encode()
#         response['Content-Type'] = 'application/octet-stream'
#         return response
#
#     def getExcel(self, kwargs):
#
#         queryId = str(decode_id(kwargs['caseId']))
#         path = kwargs['path'].split('-')[0]
#         dirname = kwargs['path'].split('-')[1]
#         if dirname:
#             return {
#                 'organ': kwargs['organ'],
#                 'queryId': queryId,
#                 'path': path,
#                 'filename': dirname
#             }
#         else:
#             return None

# 获取百分位数标准差
class heightPercentSD(FormattedView):
    extractor = extractors.AllExtractor()

    @parse_arguments('url')
    def get(self, request, *args, **kwargs):

        a = float(kwargs['age'])
        s = int(kwargs['sex'])
        trueHeight = float(kwargs['height'])
        avgHeight = get_height(a, 50, s)
        high = get_height(a, 25, s)
        SD = (trueHeight - avgHeight) / (avgHeight - high)
        result = {
            "SD": SD,
            "avgHeight": avgHeight
        }
        return self.make_response(result)

# 统计每家医院的上传数量
class StatisticPosi(FormattedView):
    extractor = extractors.AllExtractor()
    @parse_arguments()
    def get(self, request, *args, **kwargs):
        # 查询数据
        row = getStatisticPosi(request, kwargs)

        limit = kwargs['limit']
        paginator = Paginator(row, limit)  # 每页显示10条
        page = kwargs['currPage']
        if page == '0':
            page = '1'
        pagedata = {}  # 获取分页信息
        pagedata['count'] = paginator.count
        pagedata['num_pages'] = paginator.num_pages
        pagedata['per_page'] = limit
        pagedata['current'] = page
        context = {}
        objectlist = list(paginator.page(page).object_list)
        contacts = self.extractor.extract(objectlist)
        context['contacts'] = contacts
        context['pagedata'] = pagedata

        return self.make_response(context)

def getStatisticPosi(request, kwargs):
    user = request.user
    # 组合sql语句
    sql1 = "SELECT u.unit_name,"
    sql = "count(c.dis_class) sums," \
          "sum(YEARWEEK( date_format(  c.c_time,'%Y-%m-%d' ) ) = YEARWEEK( now() )) AS benz," \
          "sum(DATE_FORMAT( c.c_time, '%Y%m' ) = DATE_FORMAT( CURDATE() ,'%Y%m' )) AS beny " \
          "FROM datamain_patient as c,login_unit as u where c.del_flg='1' and u.id = c.up_mec"
    if 'organId' in kwargs and len(kwargs['organId']) > 0:
        sql += " and c.dis_class = '" + kwargs['organId'] + "'"
    if 'upunit' in kwargs and len(kwargs['upunit']) > 0:
        sql += " and u.unit_name like '%" + kwargs['upunit'] + "%'"
    else:
        sql += " and u.unit_name like '%%'"
    if 'createDateRange' in kwargs and ',' in kwargs['createDateRange']:
        items = kwargs['createDateRange'].split(',')
        ctimegte = items[0]
        ctimelte = items[1]
        sql += " and c.c_time between '" + ctimegte + "' and  '" + ctimelte + "'"
    sql2 = sql1 + sql + " group by c.up_mec"
    sql = "select '合计' as up_mec ," + sql + " UNION ALL " + sql2
    if 'sortby' in kwargs and kwargs['sortby']:
        if 'order' in kwargs and kwargs['order'] == 'desc':
            sql = sql + " order by "+kwargs['sortby']+" DESC"
        else:
            sql = sql + " order by "+kwargs['sortby']
    else:
        sql = sql + " order by sums DESC"
    cursor = connection.cursor()
    cursor.execute(sql)
    row = cursor.fetchall()
    return row

# 统计每病例的上传数量
class StaBl(FormattedView):
    extractor = extractors.AllExtractor()
    @parse_arguments()
    def get(self, request, *args, **kwargs):
        # 查询数据
        row = getStaBl(request, kwargs)
        data = {'hj': 0, 'xzs': 0, 'ggbb': 0, 'lnkfb': 0, 'jzxgnkj': 0, 'zdfd': 0, 'slxj': 0, 'xjjt': 0, 'tlxj': 0, 'qt': 0}
        for num in row:
            if num[0] == "1":
                data['xzs'] = data['xzs']+num[1]
            elif num[0] == "2":
                data['ggbb'] = data['ggbb']+num[1]
            elif num[0] == "3":
                data['lnkfb'] = data['lnkfb']+num[1]
            elif num[0] == "4":
                data['jzxgnkj'] = data['jzxgnkj']+num[1]
            elif num[0] == "5":
                data['zdfd'] = data['zdfd']+num[1]
            elif num[0] == "6":
                data['slxj'] = data['slxj']+num[1]
            elif num[0] == "7":
                data['xjjt'] = data['xjjt']+num[1]
            elif num[0] == "8":
                data['tlxj'] = data['tlxj']+num[1]
            else:
                data['qt'] = data['qt']+num[1]
            data['hj'] = data['xzs']+data['ggbb']+data['lnkfb']+data['jzxgnkj']+data['zdfd']+data['slxj']+data['xjjt']+data['tlxj']+data['qt']
        dateone = [
            {"value":data['xzs'], "name":"性早熟"},
            {"value":data['ggbb'], "name":"骨骼病变"},
            {"value":data['lnkfb'], "name":"牛奶咖啡斑"},
            {"value":data['jzxgnkj'], "name":"甲状腺功能亢进"},
            {"value":data['zdfd'], "name":"肢端肥大/生长过速"},
            {"value":data['slxj'], "name":"视力下降"},
            {"value":data['xjjt'], "name":"嗅觉减退"},
            {"value":data['tlxj'], "name":"听力下降"},
            {"value":data['qt'], "name":"其他"}
        ]
        context = {}
        context['data'] = data
        context['data1'] = dateone
        return self.make_response(context)

def getStaBl(request, kwargs):
    user = request.user
    # 组合sql语句
    sql = "select ma.ini_per,count(*) from datamain_mas as ma,datamain_patient as ca,login_unit as u " \
           "where ma.patient_id = ca.id and ca.del_flg=1 and u.id = ca.up_mec "
    if 'upunit' in kwargs and len(kwargs['upunit']) > 0:
        sql += " and u.unit_name like '%" + kwargs['upunit'] + "%'"
    else:
        sql += " and u.unit_name like '%%'"
    if 'createDateRange' in kwargs and ',' in kwargs['createDateRange']:
        items = kwargs['createDateRange'].split(',')
        ctimegte = items[0]
        ctimelte = items[1]
        sql += " and ca.c_time between '" + ctimegte + "' and  '" + ctimelte + "'"
    sql = sql + " group by ma.ini_per"
    cursor = connection.cursor()
    cursor.execute(sql)
    row = cursor.fetchall()
    return row

# 小程序随访记录(不需要登录)
class XFollowView(FormattedView):
    extractor = extractors.AllExtractor()
    # 不需要登录
    loginRequired = False

    # 添加随访记录
    @parse_arguments()
    # @require_arguments(['queryId'], 'body')
    def get(self, request, *args, **kwargs):
        #  手机号
        try:
            patient = models.Patient.objects.get(contacts_num=kwargs['mobilephone'])
            # 出生日期
            if 'date_birth' in kwargs and kwargs['date_birth'] is not None and len(kwargs['date_birth']):
                time = kwargs['date_birth'][0:10]
                patient.birth_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
        except:
            patient = models.Patient()
            patient.dis_class = '10000001'
            patient.name = kwargs['patient_name']
            patient.contacts_num = kwargs['mobilephone']
            patient.fir_vis_time = datetime.datetime.now
            patient.imp_per = '1'
            patient.c_time = datetime.datetime.now
            patient.del_flg = '1'
            patient.up_mec = 10
            qianzui = 'US-Xcx' + str(
                timezone.now().year * 10000 + timezone.now().month * 100 + timezone.now().day)
            num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(num=Count('case_num'))
            nums = num['num']
            caseNum = qianzui + str(nums + 1)
            case_num = caseNum
            patient.case_num = case_num
            patient.save()
            # 出生日期
            if 'date_birth' in kwargs and kwargs['date_birth'] is not None and len(kwargs['date_birth']):
                time = kwargs['date_birth'][0:10]
                patient.birth_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")

        try:
            with transaction.atomic():
                patient.save()
                # 判断检查部位，新增附表
                result = modifyorAddCase(patient.pk, kwargs)
                result.save()
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)


        if 'queryPId' in kwargs and len(kwargs['queryPId']) > 0:
            patFoll = models.PatFoll.objects.get(pk=decode_id(kwargs['queryPId']))
        else:
            patFoll = models.PatFoll()
        #  身高
        if 'height' in kwargs:
            patFoll.Ht = kwargs['height']
        #  体重
        if 'weight' in kwargs:
            patFoll.Wt = kwargs['weight']
        #  年龄
        BirthDate = patient.birth_time
        Today = datetime.datetime.now()
        interval = Today - BirthDate
        interval = interval.days
        age = str(interval/365).split('.')
        year = age[0]
        month = int(round(float("0."+age[1])*12, 0))
        patFoll.age = str(year) + "岁" + str(month) + "个月"
        # 病例主表id
        patFoll.patient_id = patient.pk
        # 删除标志
        patFoll.del_flg = '1'

        try:
            patFoll.save()
            return self.make_response(patFoll)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

# 小程序根据手机号查询随访记录(不需要登录)
class CFollowView(FormattedView):
    extractor = extractors.AllExtractor()
    # 不需要登录
    loginRequired = False

    # 查询随访记录
    @parse_arguments()
    def get(self, request, *args, **kwargs):
        try:
            context = {}
            patient = models.Patient.objects.get(contacts_num=kwargs['mobilephone'])
            patFolllist = models.PatFoll.objects.filter(patient=patient.pk)
            # 性别
            sex = patient.sex
            # 年龄身高
            ageh = []
            # 年龄体重
            agew = []
            for patFoll in patFolllist:
                if patFoll.Ht and patFoll.Wt and patFoll.age:
                    age = patFoll.age.replace("个月", "").split("岁")
                    age = round(int(age[0])+int(age[1])/12, 2)
                    dgageh = []
                    dgagew = []
                    dgageh.append(age)
                    dgageh.append(float(patFoll.Ht))
                    ageh.append(dgageh)
                    dgagew.append(age)
                    dgagew.append(float(patFoll.Wt))
                    agew.append(dgagew)
            context['sex'] = sex
            context['ageh'] = ageh
            context['agew'] = agew
            return self.make_response(context)
        except:
            return self.make_response(None, Code.DATA_PARSE_FAILED)

# E路陪伴添加记录(不需要登录)
class EPatientView(FormattedView):
    extractor = extractors.AllExtractor()
    # 不需要登录
    loginRequired = False

    # 添加病例主表信息
    # @parse_arguments('body')
    def post(self, request, *args, **kwargs):
        code = 0
        try:
            data = json.loads(request.POST.get("data"))
            context = {}
            followlist = data['followList']
            patient = data['patientInfo']
            print(followlist)
            print(patient)
            if patient['dis_class'] == '100000010':
                # 天元公学项目
                # 根据手机号，姓名查询是否存在该学生
                phone = patient['contacts_num']
                name = patient['name']
                if 'single1756696675045' in followlist[0]:
                    try:
                        student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                    except:
                        student = schoolmoddel.Student()
                    try:
                        # 基本信息
                        student.phone = phone
                        student.name = name
                        student.sex = patient['sex']
                        student.up_mec = patient['doctor_unit']
                        student.doctor = patient['doctor_name']
                        student.count = json.dumps(followlist[0])
                        student.save()
                    except Exception as e:
                        print(e)        
                # 长处和困难问卷
                elif 'single5935289439709' in followlist[0]:
                    # 根据主表主键查找是否已填写
                    try:
                        student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                        try:
                            cchkn = schoolmoddel.Cchkn.objects.get(student=student.pk)
                        except:
                            cchkn = schoolmoddel.Cchkn()
                            cchkn.student = student
                        cchkn.count = json.dumps(followlist[0])
                        cchkn.save()
                    except Exception as e:
                        print(e)   
                # 儿童气质问卷（CBQ-SF）
                elif 'single1756731698125' in followlist[0]:
                    # 根据主表主键查找是否已填写
                    student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                    try:
                        cbq = schoolmoddel.Cbq.objects.get(student=student.pk)
                    except:
                        cbq = schoolmoddel.Cbq()
                        cbq.student = student
                    cbq.count = json.dumps(followlist[0])
                    cbq.save()
                # 母亲照养方式问卷
                elif 'single1756730173727' in followlist[0]:
                    # 根据主表主键查找是否已填写
                    student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                    try:
                        mqzyfs = schoolmoddel.Mqzyfs.objects.get(student=student.pk)
                    except:
                        mqzyfs = schoolmoddel.Mqzyfs()
                        mqzyfs.student = student
                    mqzyfs.count = json.dumps(followlist[0])
                    mqzyfs.save()
                # 亲子活动
                elif 'single1756729275470' in followlist[0]:
                    # 根据主表主键查找是否已填写
                    student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                    try:
                        qzhd = schoolmoddel.Qzhd.objects.get(student=student.pk)
                    except:
                        qzhd = schoolmoddel.Qzhd()
                        qzhd.student = student
                    qzhd.count = json.dumps(followlist[0])
                    qzhd.save()
                # 屏幕暴露
                elif 'input1756728983808' in followlist[0]:
                    # 根据主表主键查找是否已填写
                    student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                    try:
                        pmbl = schoolmoddel.Pmbl.objects.get(student=student.pk)
                    except:
                        pmbl = schoolmoddel.Pmbl()
                        pmbl.student = student
                    pmbl.count = json.dumps(followlist[0])
                    pmbl.save()
                # 身体活动
                elif 'number1756728845587' in followlist[0]:
                    try:
                        # 根据主表主键查找是否已填写
                        student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                        try:
                            sthd = schoolmoddel.Sthd.objects.get(student=student.pk)
                        except:
                            sthd = schoolmoddel.Sthd()
                            sthd.student = student
                        sthd.count = json.dumps(followlist[0])
                        sthd.save()
                    except Exception as e:
                        print(e)   
                # 儿童睡眠习惯问卷
                elif 'single9117268486568' in followlist[0]:
                    # 根据主表主键查找是否已填写
                    student = schoolmoddel.Student.objects.get(phone=phone,name=name)
                    try:
                        smxg = schoolmoddel.Smxg.objects.get(student=student.pk)
                    except:
                        smxg = schoolmoddel.Smxg()
                        smxg.student = student
                    smxg.count = json.dumps(followlist[0])
                    smxg.save()
                # datamap ={
                #     # 基本信息
                #     'single1756696675045':'填表人和孩子的关系',
                #     'input1756696665371':'编号',
                #     'input1756696732840':'班级',
                #     'input1756696803524':'姓名',
                #     'single1756696772937':'性别',
                #     'date1756696944829':'出生日期',
                #     'number1756696965952':'当前身高',
                #     'number1756696967346':'当前体重',
                #     'single1756707724097':'母亲受教育程度',
                #     'single1756707744295':'父亲受教育程度',
                #     'input1756707761018':'家庭年收入',
                #     'input1756707779473':'孩子的主要照护人',
                #     'single1756707756181':'主要照护人受教育程度',
                #     'single1756707801503':'是否有兄弟姐妹',
                #     'single1756708417969':'医生诊断为妊娠期糖尿病',
                #     'single1756708418487':'医生诊断为妊娠期高血压',
                #     'single1756708418676':'精神压力大或者情绪问题且需要专业人员帮助',
                #     'single1756708418856':'医生诊断为营养不良',
                #     'single1756708875449':'分娩方式',
                #     'number1756708895187':'出生体重',
                #     'number1756708910323':'出生孕周',
                #     'single1756708921047':'出生时是否发生窒息或抢救',
                #     'single1756709150840':'出生后喂养方式',
                #     'input1756709169802':'断母乳时间',
                #     'input1756709180243':'添加辅食时间',
                #     # 长处和困难问卷
                #     'single5935289439709':'能体谅到别人的感受',
                #     'single4555877475803':'不安定、过分活跃、不能长久安静',
                #     'single1328234131362':'经常抱怨头痛、肚子痛或身体不舒服',
                #     'single8238193464660':'很乐意与别的小孩分享东西，比如糖果、玩具、铅笔等等',
                #     'single2652969504432':'经常发脾气或大吵大闹',
                #     'single0282801624200':'比较孤独，喜欢自己一个人玩',
                #     'single8712646846608':'一般来说，比较顺从，通常是大人要求做的都肯做',
                #     'single7679296064974':'有很多担忧，经常表现出忧虑',
                #     'single1566635330819':'如果有人受伤，不舒服或是生病，都很乐意提供帮助',
                #     'single2665050173852':'经常的坐立不安或躁动',
                #     'single6550612352425':'有一个或一个以上的好朋友',
                #     'single6021318142047':'经常与别的小孩吵架或欺负其他小孩子',
                #     'single1350396599925':'经常不高兴、情绪低落或哭泣',
                #     'single9557716661777':'一般来说，受别的小孩所喜欢',
                #     'single3662841468580':'容易分心，注意力不集中',
                #     'single3631574605284':'在新环境下，会紧张或粘住大人，容易失去信心',
                #     'single1892615498655':'爱对年纪小的儿童和善',
                #     'single2355561277734':'经常撒谎或欺骗',
                #     'single8460428279830':'受别的小孩捉弄或欺负',
                #     'single9934054573311':'经常自愿的帮助别人(父母、老师或其他小孩）',
                #     'single4319205653350':'做事前会想清楚',
                #     'single3825572213238':'会从事家里、学校或其他地方偷东西',
                #     'single0072231055600':'跟大人相处比跟小孩子相处融洽',
                #     'single2781431451836':'对很多事情容易感到害怕，容易受惊吓',
                #     'single3050243727447':'做事情能做到底，注意力持久',
                #     # 儿童气质问卷（CBQ-SF）
                #     'single1756731698125':'似乎总是匆匆忙忙地从一个地方到另一个地方',
                #     'single1756731742920':'当被阻止做想做的事时，会变得非常失落',
                #     'single1756731743090':'在书上画图或涂色时表现得非常专注',
                #     'single1756731743291':'喜欢滑高的滑梯或其它冒险性活动',
                #     'single1756731743492':'会因为很小的切伤或擦伤而非常不安',
                #     'single1756731743739':'能为旅行或外出准备他/她需要的东西',
                #     'single1756731743959':'经常贸然进入新情境(比如，鲁莽地到没去过的地方去玩儿，或者仓促地参加新的游戏)',
                #     'single1756731744211':'如果家庭计划没有兑现（比如，外出旅行没有按计划实施)',
                #     'single1756731744500':'喜欢别人对他唱歌',
                #     'single1756731744692':'好像对任何人都不拘束',
                #     'single1756731744929':'害怕夜贼或者“大灰狼”',
                #     'single1756731745161':'当父母穿了新衣服时，能注意到',
                #     'single1756731883290':'相对于活跃性游戏，更喜欢安静的活动',
                #     'single1756732329380':'对某事生气时，往往要持续十分钟或更长时间',
                #     'single1756731745428':'当搭建或者拼凑某些东西时，能够非常投入且坚持很长时间',
                #     'single1756731848262':'在荡秋千时喜欢又高又快',
                #     'single1756731848503':'在不能完成某些任务时似乎很沮丧',
                #     'single1756731848765':'善于按照要求行动',
                #     'single1756731924715':'需要花很长的时间适应新的环境',
                #     'single1756731925283':'在感冒时很少抱怨',
                #     'single1756732009604':'喜欢歌曲，比如童谣',
                #     'single1756732023916':'即使在认识了很长时间的人面前，有时也会害羞',
                #     'single1756732024111':'在烦躁时，很容易被安抚下来',
                #     'single1756732024322':'能很快注意到客厅里的新东西',
                #     'single1756732024533':'即使在晚上，也精力充沛',
                #     'single1756732028918':'不害怕黑夜',
                #     'single1756732030803':'有时会专注于图画书很长时间',
                #     'single1756732032230':'不喜欢粗野的游戏',
                #     'single1756732041398':'对轻微切伤或擦伤并不十分心烦',
                #     'single1756732044034':'到听说有危险的地方时会小心翼翼',
                #     'single1756732044941':'会缓慢而不匆忙地决定接下来要做的事',
                #     'single1756732129180':'当不能找到他/她想玩的东西时会生气',
                #     'single1756732129355':'喜欢柔和有节拍的活动，比如摇摆',
                #     'single1756732129631':'有时会对新认识的人害羞地转过脸去',
                #     'single1756732130270':'当喜欢的亲戚或朋友在来访后准备离开时，变得烦躁',
                #     'single1756732130480':'会对父母外表的变化做评价',
                #     # 母亲照养方式问卷
                #     'single1756730173727':'只要孩子不高兴，犯了错误，也不批评',
                #     'single1756730267576':'孩子要什么就给什么',
                #     'single1756730267999':'孩子是否服从自己无所谓',
                #     'single1756730268451':'只要孩子高兴，可以不惜一切',
                #     'single1756730268857':'孩子不服从家长时打骂孩子',
                #     'single1756730269250':'对孩子犯了错误并不在乎',
                #     'single1756730269707':'对孩子一点点异常过分着急',
                #     'single1756730270137':'鼓励孩子做他会做的事',
                #     'single1756730270613':'孩子做不好的事情替他做',
                #     'single1756730271081':'根据孩子本人的兴趣培养他的特长',
                #     'single1756730321182':'对孩子的哭闹，有时查问清楚，有时拒绝',
                #     'single1756730321463':'在和孩子谈话时允许孩子插话提问',
                #     'single1756730321722':'要求孩子做什么事都必须报告家长',
                #     'single1756730321972':'当孩子做错时问明原因再批评',
                #     'single1756730322214':'孩子缠着问这问那不耐烦',
                #     'single1756730322456':'不向孩子做任何承诺',
                #     'single1756730322727':'吩咐孩子做事时让孩子明白为什么或怎么做',
                #     'single1756730322995':'对孩子的学习、生活有时关心，有时不关心',
                #     'single1756730323282':'不注意孩子在做什么或怎么做',
                #     'single1756730323889':'自己忙的时候不理睬孩子的提问',
                #     'single1756730384728':'孩子想怎么样就怎么样',
                #     'single1756730384909':'不切实际地表扬孩子',
                #     'single1756730385104':'不了解孩子不和父母在一起时具体做什么',
                #     'single1756730385274':'以适当的方式表扬或奖励孩子',
                #     'single1756730385442':'对孩子无理要求，有时满足，有时拒绝',
                #     'single1756730385636':'对孩子没有惩罚或奖励',
                #     'single1756730385941':'在孩子学习或做其他事遇到困难时帮助他解决',
                #     'single1756730386004':'看着孩子做事情并随时指点',
                #     'single1756730386207':'对孩子提出的问题予以认真解答',
                #     'single1756730386392':'同样一件事情，有时允许，有时拒绝',
                #     'single1756730440679':'对孩子不讲是非',
                #     'single1756730440867':'孩子和谁在一起经过家长同意',
                #     'single1756730441053':'同孩子一起消遣、游戏',
                #     'single1756730441235':'要求孩子做什么事必讲明原因或怎么做',
                #     'single1756730441417':'孩子做了错事，有时批评，有时无所谓',
                #     'single1756730441587':'不关心孩子的生活小事',
                #     'single1756730508832':'通过说理使孩子服从',
                #     'single1756730441765':'有时说服孩子，有时强制孩子',
                #     'single1756730441935':'孩子在家里随便做自己的事情，家长没有具体要求',
                #     'single1756730442126':'培养孩子哪方面特长由家长决定',
                #     # 亲子活动
                #     'single1756729275470':'与孩子一起阅读、看图画书',
                #     'single1756729370162':'在生活中教孩子数的概念',
                #     'single1756729383638':'涂涂画画',
                #     'single1756729383834':'跟孩子一起玩开发智力的游戏(如搭积木、图形配对、过家家等)',
                #     'single1756729384068':'结合日常生活与孩子一起识字',
                #     'single1756729384270':'一起听唱歌曲、诗歌、童谣',
                #     'single1756729384503':'讲故事(家长讲，孩子听）',
                #     'single1756729384701':'做手工',
                #     'single1756729384937':'做运动',
                #     'single1756729385138':'教孩子生活自理技能，如吃饭、穿衣等',
                #     'single1756729385392':'与孩子谈论周围发生的一些事',
                #     'single1756729385604':'与孩子一起认识大自然的动植物',
                #     # 屏幕暴露
                #     'input1756728983808':'第一次接触电子屏幕的月龄',
                #     'number1756729001329':'平均每天接触电子屏幕时间',
                #     'matrix_input1756729018269':'矩阵填写',
                #     'single1756729052454':'您的孩子观看电视时，您或者其他照样人陪同观看的时间',
                #     'single1756729160013':'您的孩子观看电视时，您或者其他照样人与其交流电视内容的时间',
                #     # 身体活动
                #     'number1756728845587':'参加的“中强度身体活动”的频率（每周几天）',
                #     'number1756728884518':'在参加中高强度身体活动的那几天，您孩子通常每天花多少时间来做中高强度身体活动？每次至少持续10分钟以上',
                #     'number1756728899960':'参加的“低强度身体活动”的频率（每周几天）',
                #     'number1756728901468':'在参加低强度身体活动的那几天，您孩子通常每天花多少时间来做低强度身体活动？每次至少持续10分钟以上',
                #     'number1756728917338':'孩子静坐的频率（每周几天）',
                #     'number1756728918541':'每天静坐的时间',
                #     'number1756728926713':'孩子非看屏幕的静坐频率（每周几天）',
                #     'number1756728927915':'每天非看屏幕的静坐时间如躺在垫子上，坐在高脚椅上、婴儿车或手推车中而几乎不动，坐着看书或坐着玩游戏',
                #     'number1756728935555':'孩子屏幕前静坐的频率（每周几天）',
                #     'number1756728936756':'每天屏幕前静坐的时间如被动地观看屏幕娱乐节目（电视、计算机、移动设备）的时间,不包括需要进行身体活动或运动的积极屏幕游戏',
                #     # 儿童睡眠习惯问卷
                #     'date3424648648781':'孩子晚上就寝/上床时间：平时（周一至周五）',
                #     'date0347156049369':'孩子晚上就寝/上床时间：周末',
                #     'date1756713401175':'孩子晚上睡着时间（通常晚于就寝时间）：平时（周一至周五）',
                #     'date1756713401358':'孩子晚上睡着时间（通常晚于就寝时间）：周末',
                #     'single9117268486568':'孩子晚上在固定时间上床睡觉',
                #     'single3074236108195':'孩子上床后在20分钟内入睡',
                #     'single6845340622502':'孩子在自己床上独自入睡',
                #     'single5965729156379':'孩子在他人（父母或兄弟姐妹）床上入睡',
                #     'single4585907475779':'孩子入睡时出现摇摆或节律性动作',
                #     'single1179448680117':'孩子需要特定物品入睡（如玩偶、特定的毛毯等）',
                #     'single7023247818938':'孩子需要家长在房间陪伴才能入睡',
                #     'single2296349923815':'到了就寝时间，孩子会准备好去睡觉',
                #     'single2758246626579':'到了就寝时间，孩子会抗拒去睡觉',
                #     'single9761184994125':'到了就寝时间，孩子会挣扎（如哭闹、拒绝待在床上等）',
                #     'single3840259063071':'孩子害怕在黑暗中睡觉',
                #     'single1242217990682':'孩子害怕独自一个人睡觉',
                #     'date1756799182359':'通常孩子每天的睡眠：（包括夜间睡眠和日间小睡时间） ',
                #     'single3296537974217':'孩子睡得太少',
                #     'single5282752881329':'孩子睡得太多',
                #     'single4332688179714':'孩子的睡眠适量',
                #     'single1572624835778':'孩子每天的睡眠量都一样',
                #     'single5824145347676':'孩子晚上会尿床',
                #     'single3109259599021':'孩子睡眠中会说梦话',
                #     'single0518882235023':'孩子睡眠中不安稳，常动来动去',
                #     'single6628688222336':'孩子夜间会梦游（睡眠过程中行走）',
                #     'single8627876835523':'孩子夜间会移动到他人（如父母、兄弟姐妹等）的床上',
                #     'single7968351340792':'孩子反映睡眠中身体疼痛。如果有，说明哪里痛',
                #     'input5099542814368':'如果有，在何部位',
                #     'single9690438596368':'孩子睡眠中有磨牙现象（牙医可能告诉过您）',
                #     'single8345927733618':'孩子睡眠中打鼾/呼噜很响',
                #     'single8145811901085':'孩子睡眠中出现呼吸暂停',
                #     'single6482231189336':'孩子睡眠中鼻息重或气急',
                #     'single2583367066271':'孩子不在家（如到亲戚家或去旅行）睡觉时有问题',
                #     'single0913054362107':'孩子抱怨睡眠问题',
                #     'single5349607208939':'孩子夜间醒来尖叫、出汗且无法安抚',
                #     'single3480759321391':'孩子被噩梦惊醒',
                #     'date1756799208708':'夜间醒来一般总共持续：平时（周一至周五）',
                #     'date1756799244875':'夜间醒来一般总共持续：周末',
                #     'single1846808792850':'孩子夜间会醒来一次',
                #     'single6875194117042':'孩子夜间会醒来一次以上',
                #     'date1756714509878':'孩子早晨醒来的时间：平时',
                #     'date1756714483389':'孩子早晨醒来的时间：周末',
                #     'date1756714509574':'孩子早晨起床的时间（一般晚于醒来的时间）：平时',
                #     'date1756714509768':'孩子早晨起床的时间（一般晚于醒来的时间）：周末',
                #     'single3786727293210':'孩子早晨自己醒来',
                #     'single6798271528604':'孩子早晨由闹钟叫醒',
                #     'single1918146210811':'孩子醒来后情绪不佳',
                #     'single7459564628797':'孩子早晨由他人（如家长或兄弟姐妹）叫醒',
                #     'single3340495030646':'孩子早晨起床困难',
                #     'single8665467198880':'孩子早晨需要很长时间才能清醒',
                #     'single5307508515278':'孩子早晨醒来很早',
                #     'single4166842310706':'孩子早晨胃口很好',
                #     'single3988919934371':'孩子日间会小睡',
                #     'single1678343519059':'孩子在兴奋活动中突然睡着了',
                #     'single1807864306925':'孩子看起来很疲倦',
                #     'single8339065459929':'独自玩耍',
                #     'single1737636405692':'看电视',
                #     'single3746312232276':'坐车',
                #     'single3899616922374':'吃饭',
                # }
                return self.make_response(None)
            else:
                # 添加基本信息
                try:
                    patientk = models.Patient.objects.filter(name=patient['name'],contacts_num = patient['contacts_num'])[0]
                except:
                    patientk = models.Patient()
                    patientk.dis_class = patient['dis_class']
                    patientk.name = patient['name']
                    patientk.imp_per = '2'
                    patientk.up_mec = 2
                    patientk.c_time = datetime.datetime.now
                    patientk.del_flg = '1'
                    disclassmap = {
                        '10000001':'',
                        '10000002':'',
                        '10000003':'',
                        '10000004':'',
                        '10000005':'',
                        '10000006':'',
                        '10000007':'',
                        '10000008':'',
                        '10000009':'',
                        '10000010':'天元公学',
                    }
                    qianzui = 'US-Test' + str(
                        timezone.now().year * 10000 + timezone.now().month * 100 + timezone.now().day)
                    num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(
                        num=Count('case_num'))
                    nums = num['num']
                    caseNum = qianzui + str(nums + 1)
                    case_num = caseNum
                    patientk.case_num = case_num
                    patientk.doctor_name = "严乾超测试"
                    patientk.hospital_name = "严乾超测试"
                    patientk.tags = "严乾超测试"
                    patientk.sex = patient['sex']
                    patientk.gonadal_sex = patient['sex']
                    patientk.medrec_num = case_num
                # 随访日期
                if 'followTime' in followlist and followlist['followTime'] is not None and len(followlist['followTime']):
                    time = followlist['followTime']
                    patientk.fir_vis_time = datetime.datetime.strptime(time, "%Y-%m-%d")
                if 'birth_time' in patient and patient['birth_time'] is not None and len(patient['birth_time']):
                    time = patient['birth_time']
                    patientk.birth_time = datetime.datetime.strptime(time, "%Y-%m-%d")
                # if 'FHt' in patient:
                #     patientk.FHt = patient['FHt']
                # if 'MHt' in patient:
                #     patientk.MHt = patient['MHt']
                try:
                    with transaction.atomic():
                        patientk.save()
                        # 判断检查部位，新增附表
                        result = modifyorAddShort(patientk.pk, kwargs)
                        result.save()
                        # 添加随访信息
                        for patFolls in followlist:
                            patFoll = models.PatFoll()
                            # 病例主表id
                            patFoll.patient_id = patientk.pk
                            # 随访日期
                            if 'followTime' in patFolls and patFolls['followTime'] is not None and len(patFolls['followTime']):
                                time = patFolls['followTime'][0:10]
                                patFoll.foll_time = datetime.datetime.strptime(time + ' 00:00:00', "%Y-%m-%d %H:%M:%S")
                            #  现身高
                            if 'Ht' in patFolls:
                                patFoll.Ht = patFolls['Ht']
                            #  年龄
                            if 'age' in patFolls:
                                patFoll.age = patFolls['age']
                            #  现体重
                            if 'Wt' in patFolls:
                                patFoll.Wt = patFolls['Wt']
                            # #  是否行为发育评估
                            # if 'isFYassess' in patFolls:
                            #     patFoll.beh_dev_ass = patFolls['isFYassess']
                            # #  Peabody运动发育评估
                            # if 'Peabody' in patFolls:
                            #     patFoll.ped_mot_dev_ass = patFolls['Peabody']
                            # #  粗大运动
                            # if 'GriffithsA' in patFolls:
                            #     patFoll.gro_mot = patFolls['GriffithsA']
                            # #  个人社会
                            # if 'GriffithsB' in patFolls:
                            #     patFoll.ind_soc = patFolls['GriffithsB']
                            # #  听力语言
                            # if 'GriffithsC' in patFolls:
                            #     patFoll.lis_lan = patFolls['GriffithsC']
                            # #  手眼协调
                            # if 'GriffithsD' in patFolls:
                            #     patFoll.han_eye_coo = patFolls['GriffithsD']
                            # #  视觉表现
                            # if 'GriffithsE' in patFolls:
                            #     patFoll.vis_rep = patFolls['GriffithsE']
                            # #  实际推理
                            # if 'GriffithsF' in patFolls:
                            #     patFoll.pra_rea = patFolls['GriffithsF']
                            # #  韦氏智力量表
                            # if 'wszlb' in patFolls:
                            #     patFoll.wec_sca = patFolls['wszlb']
                            #  双乳生殖器分期
                            if 'breastDev' in patFolls:
                                patFoll.gen_stag = patFolls['breastDev']
                            #  外生殖器分期
                            if 'genStag' in patFolls:
                                patFoll.gen_stag = patFolls['genStag']
                            #  阴毛分期
                            if 'pubStag' in patFolls:
                                patFoll.pub_stag = patFolls['pubStag']
                            #  IGF-1（ng/ml）
                            if 'IGF1' in patFolls:
                                patFoll.IGF1 = patFolls['IGF1']
                            #  IGFBP-3（ug/ml）
                            if 'IGFBP3' in patFolls:
                                patFoll.IGFBP3 = patFolls['IGFBP3']
                            #  甲功
                            if 'Jiagong' in patFolls:
                                jg = '{"Jiagong":"' + patFolls['Jiagong'] + '",'
                            else:
                                jg = '{"Jiagong":"null",'
                            if 'JiagongDes' in patFolls:
                                jg += '"JiagongDes":"' + patFolls['JiagongDes'] + '"}'
                            else:
                                jg += '"JiagongDes":"null"}'
                            patFoll.Jiagong = jg
                            #  空腹血糖
                            if 'fasBloodGlu' in patFolls:
                                patFoll.fas_blood_glu = patFolls['fasBloodGlu']
                            #  空腹胰岛素
                            if 'fasInsulin' in patFolls:
                                patFoll.fas_insulin = patFolls['fasInsulin']
                            #  肝肾脂电解质
                            if 'livKidLip' in patFolls:
                                gz = '{"livKidLip":"' + patFolls['livKidLip'] + '",'
                            else:
                                gz = '{"livKidLip":"null",'
                            if 'LAKLEdes' in patFolls:
                                gz += '"LAKLEdes":"' + patFolls['LAKLEdes'] + '"}'
                            else:
                                gz += '"LAKLEdes":"null"}'
                            patFoll.liv_kid_lip = gz
                            #  糖化血红蛋白
                            if 'glyHem' in patFolls:
                                patFoll.gly_hem = patFolls['glyHem']
                            #  LH
                            if 'LH' in patFolls:
                                patFoll.LH = patFolls['LH']
                            #  FSH
                            if 'FSH' in patFolls:
                                patFoll.FSH = patFolls['FSH']
                            #  E2
                            if 'E2' in patFolls:
                                patFoll.E2 = patFolls['E2']
                            #  T
                            if 'T' in patFolls:
                                patFoll.T = patFolls['T']
                            #  DHT
                            if 'DHT' in patFolls:
                                patFoll.DHT = patFolls['DHT']
                            #  游离睾酮
                            if 'FT' in patFolls:
                                patFoll.yltg = patFolls['FT']
                            #  SHBG
                            if 'SHBG' in patFolls:
                                patFoll.SHBG = patFolls['SHBG']
                            #  性腺B超
                            if 'uterusOne' in patFolls:
                                xxbc = '{"uterusOne":"' + patFolls['uterusOne'] + '",'
                            else:
                                xxbc = '{"uterusOne":"null",'
                            if 'uterusTwo' in patFolls:
                                xxbc += '"uterusTwo":"' + patFolls['uterusTwo'] + '",'
                            else:
                                xxbc += '"uterusTwo":"null",'
                            if 'uterusThr' in patFolls:
                                xxbc += '"uterusThr":"' + patFolls['uterusThr'] + '",'
                            else:
                                xxbc += '"uterusThr":"null",'
                            if 'cervixLong' in patFolls:
                                xxbc += '"cervixLong":"' + patFolls['cervixLong'] + '",'
                            else:
                                xxbc += '"cervixLong":"null",'
                            if 'intima' in patFolls:
                                xxbc += '"intima":"' + patFolls['intima'] + '",'
                            else:
                                xxbc += '"intima":"null",'
                            if 'ovaLeftOne' in patFolls:
                                xxbc += '"ovaLeftOne":"' + patFolls['ovaLeftOne'] + '",'
                            else:
                                xxbc += '"ovaLeftOne":"null",'
                            if 'ovaLeftTwo' in patFolls:
                                xxbc += '"ovaLeftTwo":"' + patFolls['ovaLeftTwo'] + '",'
                            else:
                                xxbc += '"ovaLeftTwo":"null",'
                            if 'ovaLeftThr' in patFolls:
                                xxbc += '"ovaLeftThr":"' + patFolls['ovaLeftThr'] + '",'
                            else:
                                xxbc += '"ovaLeftThr":"null",'
                            if 'ovaRightOne' in patFolls:
                                xxbc += '"ovaRightOne":"' + patFolls['ovaRightOne'] + '",'
                            else:
                                xxbc += '"ovaRightOne":"null",'
                            if 'ovaRightTwo' in patFolls:
                                xxbc += '"ovaRightTwo":"' + patFolls['ovaRightTwo'] + '",'
                            else:
                                xxbc += '"ovaRightTwo":"null",'
                            if 'ovaRightThr' in patFolls:
                                xxbc += '"ovaRightThr":"' + patFolls['ovaRightThr'] + '",'
                            else:
                                xxbc += '"ovaRightThr":"null",'
                            if 'follDiameter' in patFolls:
                                xxbc += '"follDiameter":"' + patFolls['follDiameter'] + '",'
                            else:
                                xxbc += '"follDiameter":"null",'
                            if 'isCyst' in patFolls:
                                xxbc += '"isCyst":"' + patFolls['isCyst'] + '",'
                            else:
                                xxbc += '"isCyst":"null",'
                            if 'cyst' in patFolls:
                                xxbc += '"cyst":"' + patFolls['cyst'] + '",'
                            else:
                                xxbc += '"cyst":"null",'
                            if 'cystOne' in patFolls:
                                xxbc += '"cystOne":"' + patFolls['cystOne'] + '",'
                            else:
                                xxbc += '"cystOne":"null",'
                            if 'cystTwo' in patFolls:
                                xxbc += '"cystTwo":"' + patFolls['cystTwo'] + '",'
                            else:
                                xxbc += '"cystTwo":"null",'
                            if 'cystThr' in patFolls:
                                xxbc += '"cystThr":"' + patFolls['cystThr'] + '",'
                            else:
                                xxbc += '"cystThr":"null",'
                            if 'cystDescribe' in patFolls:
                                xxbc += '"cystDescribe":"' + patFolls['cystDescribe'] + '",'
                            else:
                                xxbc += '"cystDescribe":"null",'
                            if 'testisLeftOne' in patFolls:
                                xxbc += '"testisLeftOne":"' + patFolls['testisLeftOne'] + '",'
                            else:
                                xxbc += '"testisLeftOne":"null",'
                            if 'testisLeftTwo' in patFolls:
                                xxbc += '"testisLeftTwo":"' + patFolls['testisLeftTwo'] + '",'
                            else:
                                xxbc += '"testisLeftTwo":"null",'
                            if 'testisLeftThr' in patFolls:
                                xxbc += '"testisLeftThr":"' + patFolls['testisLeftThr'] + '",'
                            else:
                                xxbc += '"testisLeftThr":"null",'
                            if 'testisLeftLon' in patFolls:
                                xxbc += '"testisLeftLon":"' + patFolls['testisLeftLon'] + '",'
                            else:
                                xxbc += '"testisLeftLon":"null",'
                            if 'testisRightOne' in patFolls:
                                xxbc += '"testisRightOne":"' + patFolls['testisRightOne'] + '",'
                            else:
                                xxbc += '"testisRightOne":"null",'
                            if 'testisRightTwo' in patFolls:
                                xxbc += '"testisRightTwo":"' + patFolls['testisRightTwo'] + '",'
                            else:
                                xxbc += '"testisRightTwo":"null",'
                            if 'testisRightThr' in patFolls:
                                xxbc += '"testisRightThr":"' + patFolls['testisRightThr'] + '",'
                            else:
                                xxbc += '"testisRightThr":"null",'
                            if 'MRI' in patFolls:
                                xxbc += '"MRI":"' + patFolls['MRI'] + '",'
                            else:
                                xxbc += '"MRI":"null",'
                            if 'mriDescribe' in patFolls:
                                xxbc += '"mriDescribe":"' + patFolls['mriDescribe'] + '",'
                            else:
                                xxbc += '"mriDescribe":"null",'
                            if 'testisRightLon' in patFolls:
                                xxbc += '"testisRightLon":"' + patFolls['testisRightLon'] + '"}'
                            else:
                                xxbc += '"testisRightLon":"null"}'
                            patFoll.gon_B_ult = xxbc
                            #  诊疗方案
                            if 'diaPlan' in patFolls:
                                zlfa = '{"diaPlan":"' + patFolls['diaPlan'] + '",'
                            else:
                                zlfa = '{"diaPlan":"null",'
                            if 'rhGH' in patFolls:
                                zlfa += '"rhGH":"' + patFolls['rhGH'] + '",'
                            else:
                                zlfa += '"rhGH":"null",'
                            if 'rhCustomizationDiaPlan' in patFolls:
                                zlfa += '"rhCustomizationDiaPlan":"' + patFolls['rhCustomizationDiaPlan'] + '",'
                            else:
                                zlfa += '"rhCustomizationDiaPlan":"null",'
                            if 'rhCustomizationPrompt' in patFolls:
                                zlfa += '"rhCustomizationPrompt":"' + patFolls['rhCustomizationPrompt'] + '",'
                            else:
                                zlfa += '"rhCustomizationPrompt":"null",'
                            if 'rhUnitedCustomization' in patFolls:
                                zlfa += '"rhUnitedCustomization":"' + patFolls['rhUnitedCustomization'] + '",'
                            else:
                                zlfa += '"rhUnitedCustomization":"null",'
                            if 'rhUnitedDose' in patFolls:
                                zlfa += '"rhUnitedDose":"' + patFolls['rhUnitedDose'] + '",'
                            else:
                                zlfa += '"rhUnitedDose":"null",'
                            if 'rhGHdose' in patFolls:
                                zlfa += '"rhGHdose":"' + patFolls['rhGHdose'] + '"}'
                            else:
                                zlfa += '"rhGHdose":"null"}'
                            patFoll.dia_trea_plan = zlfa
                            # 其他
                            if 'other' in patFolls:
                                patFoll.other = patFolls['other']
                            # 删除标志
                            patFoll.del_flg = '1'
                            # print(patFoll)
                            patFoll.save()
                        return self.make_response(None)
                except:
                    print(patient['name'] + "----33333--" + patient['contacts_num'])
                    return self.make_response(None, Code.DATA_PARSE_FAILED)
                return self.make_response(None)
        except:
            print(patient['name'] + "----4444--" + patient['contacts_num'])
            return self.make_response(None, Code.DATA_PARSE_FAILED)

# 小程序添加记录新(不需要登录)
class EPatientNewView(FormattedView):
    extractor = extractors.AllExtractor()
    # 不需要登录
    loginRequired = False

    # 添加病例主表信息
    @parse_arguments('body')
    def post(self, request, *args, **kwargs):
        code = 0
        try:
            context = {}
            followlist = kwargs['followList']
            patient = kwargs['patientInfo']
            # 添加基本信息
            try:
                patientk = models.Patient.objects.get(name=patient['name'], contacts_num=patient['contacts_num'])
            except:
                patientk = models.Patient()
                # 身份证
                patientk.idcard = patient['idcard']
                # 患者姓名
                patientk.name = patient['name']
                # 混淆姓名
                patientk.confuse_name = patient['confuseName']
                # 姓名大写
                patientk.upper_case = patient['upperCase']
                # 本人电话
                patientk.self_tel = patient['selfTel']
                # 联系电话
                patientk.contacts_num = patient['tel']
                # 联系人是患者的谁
                patientk.contacts_num = patient['relation']
                # 性别
                if patient['sex'] == '男':
                    patientk.sex = 1
                    patientk.gonadal_sex = 1
                else:
                    patientk.sex = 2
                    patientk.gonadal_sex = 2
                # 出生日期
                if 'birthday' in patient and patient['birthday'] is not None and len(patient['birthday']):
                    time = patient['birthday'][0:19]
                    patientk.birth_time = datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                # 年龄
                patientk.age = patient['age']
                # 患者照片
                patientk.photo = patient['photo']
                # 网格地址
                patientk.address = patient['address']
                # 标签
                patientk.tags = patient['tags']
                # 所患疾病
                patientk.dis_class = patient['categoryList']
                # 疾病描述
                patientk.category_describe = patient['categoryDescribe']
                # 身高(cm)
                patientk.height = patient['height']
                # 体重(kg)
                patientk.weight = patient['weight']
                # bmi值
                patientk.bmi = patient['bmi']

                patientk.imp_per = '2'
                patientk.c_time = datetime.datetime.now
                patientk.del_flg = '1'
                patientk.up_mec = 2
                qianzui = 'US-ElpbNew' + str(
                    timezone.now().year * 10000 + timezone.now().month * 100 + timezone.now().day)
                num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(
                    num=Count('case_num'))
                nums = num['num']
                caseNum = qianzui + str(nums + 1)
                case_num = caseNum
                patientk.case_num = case_num
                patientk.doctor_name = patient['doctor_name']
                patientk.hospital_name = patient['hospital_name']
            # # 随访日期
            # if 'fir_vis_time' in patient and patient['fir_vis_time'] is not None and len(patient['fir_vis_time']):
            #     time = patient['fir_vis_time'][0:19]
            #     patientk.fir_vis_time = datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")

            try:
                with transaction.atomic():
                    patientk.save()
                    # 判断检查部位，新增附表
                    result = modifyorAddCase(patientk.pk, kwargs)
                    result.save()
                    # 添加随访信息
                    for patFolls in followlist:
                        patFoll = models.PatFoll()
                        #  身高
                        patFoll.Ht = patFolls['ht']
                        #  体重
                        patFoll.Wt = patFolls['wt']
                        #  年龄
                        patFoll.age = patFolls['age']
                        # R系列骨龄
                        patFoll.rboneAge = patFolls['rboneAge']
                        # C系列骨龄
                        patFoll.cboneAge = patFolls['cboneAge']
                        # 体脂( %)
                        patFoll.bodyFat = patFolls['bodyFat']
                        # bmi值
                        patFoll.bmi = patFolls['bmi']
                        # 腰围(cm)
                        patFoll.waistline = patFolls['waistline']
                        # 臀围(cm)
                        patFoll.hips = patFolls['hips']
                        # 腰臀比
                        patFoll.waistToHipRatio = patFolls['waistToHipRatio']

                        # 病例主表id
                        patFoll.patient = patientk
                        # 随访时间
                        if 'foll_time' in patient and patient['foll_time'] is not None and len(patient['foll_time']):
                            time = patient['foll_time'][0:19]
                            patientk.foll_time = datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                        # 删除标志
                        patFoll.del_flg = '1'
                        patFoll.save()
                    context['code'] = code
                    return self.make_response(context)
            except:
                print(patient['name'] + "----33333--" + patient['contacts_num'])
                return self.make_response(None, Code.DATA_PARSE_FAILED)
        except:
            print(patient['name'] + "----4444--" + patient['contacts_num'])
            return self.make_response(None, Code.DATA_PARSE_FAILED)

class DownZipPl(FormattedView):
    # 请求压缩包
    @require_arguments(['selectIds'], 'body')
    def put(self, request, *args, **kwargs):
        user = request.user
        caseids = kwargs['selectIds']
        # 普通用户、审核人员
        if user.level != 1:
            return self.make_response(None, Code.PERMISSION_DENIED)
        # 管理员
        file_info = self.getZip(request, caseids)
        if file_info:
            return self.make_response(file_info)
        else:
            return self.make_response(None, Code.DEFULT_SAVE_ZIP)

    # 下载压缩包
    def get(self, request, *args, **kwargs):
        user = request.user
        # 管理员
        if user.level == 1:
            file_path = '111.zip'
            url = '/protected_files/{}'.format(file_path.replace(',', '/'))
            response = HttpResponse('')
            response['X-Accel-Redirect'] = url.encode()
            response['Content-Type'] = 'application/octet-stream'
            return response
        # 普通用户、审核人员
        else:
            return self.make_response(None, Code.PERMISSION_DENIED)

    def getZip(self, request, case_id):
        if case_id and len(case_id) > 0:
            try:
                # 添加压缩包
                file_path = write_zippl(case_id)
                ids = ''
                # 下载记录
                for caseid in case_id:
                    caseid = decode_id(caseid)
                    ids = ids+','+str(caseid)
                operStep = 10050016
                # log.saveLog(request, ids, operStep)
                return {
                    'filename': '111.zip'
                }
            except:
                return None

# 下载文件
class loadFile(FormattedView):
    # 请求导出
    @parse_arguments('body')
    def put(self, request, *args, **kwargs):

        file_info = self.getExcel(kwargs)
        if file_info:
            return self.make_response(file_info)
        else:
            return self.make_response(None, Code.DEFULT_SAVE_ZIP)

    # 下载Excel
    def get(self, request, *args, **kwargs):
        user = request.user
        # 管理员
        if user.level == 1:
            file_path = os.path.join('oneExcel.xls')
            url = '/protected_files/{}'.format(file_path.replace(',', '/'))
            response = HttpResponse('')
            response['X-Accel-Redirect'] = url.encode()
            response['Content-Type'] = 'application/octet-stream'
            return response
        # 普通用户、审核人员
        else:
            return self.make_response(None, Code.PERMISSION_DENIED)


    def getExcel(self, kwargs):

        # 查询数据
        # row = getPatiExcel(kwargs)
        patient = models.Patient.objects.filter(del_flg=1).all()

        if 'caseNum' in kwargs and kwargs['caseNum']:
            patient.filter(case_num__contains=kwargs['caseNum'])

        if 'gender' in kwargs and kwargs['gender']:
            patient.filter(sex__contains=kwargs['gender'])

        if 'disclass' in kwargs and kwargs['disclass']:
            patient.filter(dis_class=kwargs['disclass'])

        if 'name' in kwargs and kwargs['name']:
            patient.filter(name__contains=kwargs['name'])

        if 'userNum' in kwargs and kwargs['userNum']:
            patient.filter(user_num__contains=kwargs['userNum'])

        if 'createDateRange' in kwargs and ',' in kwargs['createDateRange']:
            items = kwargs['createDateRange'].split(',')
            patient.filter(c_time__gte=items[0])
            patient.filter(c_time__lte=items[1])

        patient = patient.exclude(case_num__contains='Elpb')
        patient = patient.exclude(case_num__contains='Xcx')



        # 导出Excel
        file_path = ExcelFile.imp_case_excel_one(patient)
        if file_path:
            return {
                'organ': file_path.split('/')[0],
                'filename': 'oneExcel.xls'
            }
        else:
            return None
# 下载文件
class loadFilemas(FormattedView):
    # 请求导出
    @parse_arguments('body')
    def put(self, request, *args, **kwargs):

        file_info = self.getExcel(kwargs)
        if file_info:
            return self.make_response(file_info)
        else:
            return self.make_response(None, Code.DEFULT_SAVE_ZIP)

    # 下载Excel
    def get(self, request, *args, **kwargs):
        user = request.user
        # 管理员
        if user.level == 1:
            file_path = os.path.join('oneExcel.xls')
            url = '/protected_files/{}'.format(file_path.replace(',', '/'))
            response = HttpResponse('')
            response['X-Accel-Redirect'] = url.encode()
            response['Content-Type'] = 'application/octet-stream'
            return response
        # 普通用户、审核人员
        else:
            return self.make_response(None, Code.PERMISSION_DENIED)


    def getExcel(self, kwargs):

        # 查询数据
        # row = getPatiExcel(kwargs)
        patient = models.Patient.objects.filter(del_flg=1,dis_class='10000004').all()

        if 'caseNum' in kwargs and kwargs['caseNum']:
            patient.filter(case_num__contains=kwargs['caseNum'])

        if 'gender' in kwargs and kwargs['gender']:
            patient.filter(sex__contains=kwargs['gender'])

        if 'disclass' in kwargs and kwargs['disclass']:
            patient.filter(dis_class=kwargs['disclass'])

        if 'name' in kwargs and kwargs['name']:
            patient.filter(name__contains=kwargs['name'])

        if 'userNum' in kwargs and kwargs['userNum']:
            patient.filter(user_num__contains=kwargs['userNum'])

        if 'createDateRange' in kwargs and ',' in kwargs['createDateRange']:
            items = kwargs['createDateRange'].split(',')
            patient.filter(c_time__gte=items[0])
            patient.filter(c_time__lte=items[1])

        patient = patient.exclude(case_num__contains='Elpb')
        patient = patient.exclude(case_num__contains='Xcx')



        # 导出Excel
        file_path = ExcelFile.imp_case_excel_mas(patient)
        if file_path:
            return {
                'organ': file_path.split('/')[0],
                'filename': 'oneExcel.xls'
            }
        else:
            return None

# 添加修改日志
def addModifylod(request, caseid,operStep):
    try:
        # 添加修改日志
        modifyLog = models.OperLog()
        modifyLog.oper_per_id = request.user.id
        modifyLog.oper_case_id = caseid
        modifyLog.oper_step = operStep
        modifyLog.oper_data = datetime.datetime.now()
        modifyLog.del_flg = "1"
        modifyLog.save()
        result = True
    except:
        result = False
    return result

class ModifylodListView(FormattedView):
    extractor = extractors.AllExtractor()

    # 查询修改日志列表
    @parse_arguments('url')
    def get(self, request, *args, **kwargs):
        filters = self.get_filters(request.user, kwargs)
        if filters is None:
            return self.make_response(None, Code.PERMISSION_DENIED)
        else:
            modifyLog = models.OperLog.objects.filter(**filters).all()

            if 'name' in kwargs and kwargs['name']:
                userlist = loginmoddel.User.objects.filter(name__contains = kwargs['name']).values_list('id')
                modifyLog = modifyLog.filter(oper_per_id__in=userlist)
            # if 'sortby' in kwargs and kwargs['sortby']:
            #     sortby_map = {
            #         'disClass': 'dis_class',
            #         'caseNum': 'case_num',
            #         'sex': 'sex',
            #         'cTime': 'c_time',
            #         'name': 'name',
            #     }
            #     if 'order' in kwargs and kwargs['order'] == 'desc':
            #         patient = patient.order_by('-' + sortby_map[kwargs['sortby']])
            #     else:
            #         patient = patient.order_by(sortby_map[kwargs['sortby']])
            # else:
            #     patient = patient.order_by('-modify_time')
            limit = kwargs['limit']
            paginator = Paginator(modifyLog, limit)  # 每页显示10条
            page = kwargs['currPage']
            if page == '0':
                page = '1'
            pagedata = {}  # 获取分页信息
            pagedata['count'] = paginator.count
            pagedata['num_pages'] = paginator.num_pages
            pagedata['per_page'] = limit
            pagedata['current'] = page
            context = {}
            try:
                list = paginator.page(page).object_list
            except:
                list = paginator.page('1').object_list
            for item in list:
                if item.oper_per_id and len(item.oper_per_id) > 0:
                    item.oper_per_id = loginView.getNameById(item.oper_per_id)
            contacts = self.extractor.extract(list)
            context['contacts'] = contacts
            context['pagedata'] = pagedata
            return self.make_response(context)

    def get_filters(self, user, source):
        """
        获取查询条件（将请求参数转换为数据库表字段）
        为安全考虑，请求传过来的参数名称尽量不要和数据库中的字段同名
        """

        filters = {}

        if 'createDateRange' in source and ',' in source['createDateRange']:
            items = source['createDateRange'].split(',')
            filters['c_time__gte'] = items[0]
            filters['c_time__lte'] = items[1]

        filters['del_flg'] = '1'

        return filters

# 同步数据库内容的方法，使用之后删除
class ModifyDbView(FormattedView):
    loginRequired = False
    @parse_arguments()
    def get(self, request, *args, **kwargs):
        modifyxfyyc()
        # tongbujzxax()
        # tiaoshi()
        return self.make_response(None)
    
from openpyxl import load_workbook
def modifyxfyyc():
    wb = load_workbook('/opt/eksjk/111.xlsx')
    # 获取活动的工作表或者通过名字获取特定的工作表
    ws = wb.active  # 或者 ws = wb['Sheet1']
    foll_time = '2021-06-15 00:00:00.000000'
    time = foll_time[0:19]
    # 读取数据
    for row in ws.iter_rows(values_only=True):
        patientk = models.Patient()
        patientk.dis_class = '10000007'
        patientk.name = row[0]
        patientk.imp_per = '2'
        patientk.up_mec = 2
        patientk.c_time =  datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
        patientk.del_flg = '1'
        qianzui = 'US-Szfy' + str(timezone.now().year * 10000 + timezone.now().month * 100 + timezone.now().day)
        num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(num=Count('case_num'))
        nums = num['num']
        caseNum = qianzui + str(nums + 1)
        case_num = caseNum
        patientk.case_num = case_num
        patientk.idcard = row[1]
        patientk.birth_time =  datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
        patientk.save()
        # 判断检查部位，新增附表
        szfy = models.SzfyEltm()
        szfy.patient = patientk
        szfy.save()
        # print(row)

def tiaoshi():
    a =""
    CqSchoolvalue.cqvalyue(a)


def tongbujzxax():
    shortlist = models.Short.objects.all()
    for row in shortlist:
        if row.fam_his:
            jslist = json.loads(row.fam_his)
            fahight = 180
            mahight = 180
            for js in jslist:
                if js['relation'] == '父亲' and js['height'].isdigit():
                    fahight = int(js['height'])
                if js['relation'] == '母亲' and js['height'].isdigit():
                    mahight = int(js['height'])
            if fahight < 160 or mahight < 150:
                patient = models.Patient.objects.get(pk = row.patient_id)
                patient.dis_class = "10000006"
                patient.case_num = patient.case_num.replace("Short","JzxShort")
                patient.save()
                jzxshaot = models.JzxShort()
                jzxshaot.id = row.id           
                jzxshaot.patient = row.patient           
                jzxshaot.user_num = row.user_num           
                jzxshaot.fam_his = row.fam_his           
                jzxshaot.mot_dev_back = row.mot_dev_back           
                jzxshaot.lan_dev_back = row.lan_dev_back           
                jzxshaot.int_dev_back = row.int_dev_back           
                jzxshaot.abn_hear = row.abn_hear           
                jzxshaot.rec_inf_his = row.rec_inf_his           
                jzxshaot.con_his = row.con_his           
                jzxshaot.past_other = row.past_other           
                jzxshaot.med_his = row.med_his           
                jzxshaot.phy_exa = row.phy_exa           
                jzxshaot.lab_exa = row.lab_exa           
                jzxshaot.electr = row.electr           
                jzxshaot.gon_B_ult = row.gon_B_ult           
                jzxshaot.dia_trea_plan = row.dia_trea_plan           
                jzxshaot.bio_sam_bank = row.bio_sam_bank           
                jzxshaot.main_dia = row.main_dia           
                jzxshaot.sec_dia = row.sec_dia           
                jzxshaot.follow_up = row.follow_up           
                jzxshaot.B_ult_image = row.B_ult_image           
                jzxshaot.spe_kar = row.spe_kar           
                jzxshaot.SRY = row.SRY           
                jzxshaot.gen_mut_name = row.gen_mut_name           
                jzxshaot.mut_kind = row.mut_kind           
                jzxshaot.sour_mut = row.sour_mut           
                jzxshaot.base_mut = row.base_mut           
                jzxshaot.baami_aci_mutse_mut = row.ami_aci_mut           
                jzxshaot.f_bio_sam_bank = row.f_bio_sam_bank           
                jzxshaot.m_bio_sam_bank = row.m_bio_sam_bank   
                jzxshaot.save()    


# 每天24点定时任务
# 手动同步当天数据按钮人物
class TbsffyEltmView(FormattedView):
    extractor = extractors.AllExtractor()
    # 不需要登录
    loginRequired = False

    # 同步当天数据szfy
    @parse_arguments('url')
    def get(self, request, *args, **kwargs):
        create_fake_data()
        context = {}
        return self.make_response(context)



# from apscheduler.jobstores.base import ConflictingIdError
# # 使用它可以使你的定时任务在后台运行
# from apscheduler.schedulers.background import BackgroundScheduler
# from django_apscheduler.jobstores import DjangoJobStore, register_job
import requests
# 创建虚拟数据
# scheduler = BackgroundScheduler()                       # 创建调度器对象
# scheduler.add_jobstore(DjangoJobStore(), "default")     # 调度器使用DjangoJobStore()

# # 检查任务是否已存在
# job_id = 'create_fake_data'  # 任务ID

def create_fake_data():
    # 查询今天所有记录
    response = requests.get('http://36.26.56.94:8026/datamain/getpatient')
    data = response.json()
    datalist = data['data']['data']
    age = ""
    sex =""
    birthday = ""
    idCard = ""
    name = ""
    phone= ""
    for info in datalist:
        try:    
            patient = models.Patient.objects.get(eltm_id = info[1])
            szfyEltm = models.SzfyEltm.objects.get(patient_id= patient.pk)
        except:
            patient = models.Patient()
            szfyEltm = models.SzfyEltm()
            patient.c_time = info[7]
            qianzui = 'US-Eltm'+str(timezone.now().year*10000+timezone.now().month*100+timezone.now().day)
            num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(num=Count('case_num'))
            nums = num['num']
            caseNum = qianzui+str(nums+1)
            case_num = caseNum
            patient.case_num = case_num
            patient.eltm_id = info[1]
            if info[6]:
                patient.modify_per = info[6]
                patient.imp_per = info[6]
            else:
                patient.modify_per = 1
                patient.imp_per = 1
        if info[3] == 'age':
            age = info[4]
        elif info[3] == 'sex':
            sex = info[4]
        elif info[3] == 'birthday':
            birthday = info[4]
        elif info[3] == 'idCard':
            idCard = info[4]
        elif info[3] == 'name':
            name = info[4]
        elif info[3] == 'phone':
            phone = info[4]
        # 疾病分类
        patient.dis_class = '10000007'
        # 患者姓名
        patient.name = name
        # 年龄
        patient.age = age
        # 性别
        patient.sex = sex
        # 手机
        patient.self_tel = phone
        patient.contacts_num = phone
        # 出生日期先按基础表来，后续填了就更新
        if birthday:
            patient.birth_time = birthday
        else:
            patient.birth_time = datetime.datetime.now()
        # 患者身份证号码
        if idCard and len(idCard)<19:
            patient.card = idCard
            patient.idcard = idCard
        try:    
            with transaction.atomic():
                patient.save()
                szfyEltm.patient_id = patient.pk
                szfyEltm.save()
        except:
            print("保存数据报错")
    # # 循环当前数据222987612160
    # for lbdata in datalist:
    #     # 判断随访计划是否是E路童萌,E路童萌的planid = 3059
    #     if lbdata[2] == 3059:
    #         # 判断是否在库中，根据患者E路童萌中的id
    #         eltmid = lbdata[20]
    #         txdata = json.loads(lbdata[4])
    #         # 根据患者id查询患者数据去E路童萌中查询
    #         response = requests.get('http://36.26.56.94:8026/datamain/getpatibyid?queryId='+str(eltmid))
    #         data = response.json() 
    #         try:    
    #             patient = models.Patient.objects.get(eltm_id = eltmid)
    #             szfyEltm = models.SzfyEltm.objects.get(patient_id= patient.pk)
    #         except:
    #             patient = models.Patient()
    #             szfyEltm = models.SzfyEltm()
    #             patient.c_time = lbdata[22]
    #             qianzui = 'US-Eltm'+str(timezone.now().year*10000+timezone.now().month*100+timezone.now().day)
    #             num = models.Patient.objects.filter(case_num__contains=qianzui).values().aggregate(num=Count('case_num'))
    #             nums = num['num']
    #             caseNum = qianzui+str(nums+1)
    #             case_num = caseNum
    #             patient.case_num = case_num
    #             patient.modify_per = lbdata[23]
    #             patient.eltm_id = eltmid
    #         patient.imp_per = lbdata[23]
    #         # patient.up_mec = data['data']['data3'][0][1]
    #         # patient.check_hospital = data['data']['data3'][0][1]
    #         patienteltm = data['data']['data']
    #         age = ""
    #         sex =""
    #         birthday = ""
    #         idCard = ""
    #         name = ""
    #         phone= ""
    #         for info in patienteltm:
    #             if info[3] == 'age':
    #                 age = info[4]
    #             elif info[3] == 'sex':
    #                 sex = info[4]
    #             elif info[3] == 'birthday':
    #                 birthday = info[4]
    #             elif info[3] == 'idCard':
    #                 idCard = info[4]
    #             elif info[3] == 'name':
    #                 name = info[4]
    #             elif info[3] == 'phone':
    #                 phone = info[4]
    #         # 疾病分类
    #         patient.dis_class = '10000007'
    #         # 患者姓名
    #         patient.name = name
    #         # 年龄
    #         patient.age = age
    #         # 性别
    #         patient.sex = sex
    #         # 手机
    #         patient.self_tel = phone
    #         patient.contacts_num = phone
    #         # 民族
    #         if 'input1757343120305' in txdata:
    #             patient.ethnic = txdata['input1757343120305']
    #         # 出生日期先按基础表来，后续填了就更新
    #         patient.birth_time = birthday
    #         if 'date1757343132923' in txdata:
    #             patient.birth_time = txdata['date1757343132923']
    #         # 患者身份证号码
    #         patient.card = idCard
    #         patient.idcard = idCard
    #         # 籍贯
    #         if 'input1757343149933' in txdata:
    #             patient.nat_pla = txdata['input1757343149933']
    #         # 家庭住址
    #         if 'input1757343154236' in txdata:
    #             patient.fam_adr = txdata['input1757343154236']
    #         # 现病史
    #         if 'input1757343186390' in txdata:
    #             patient.category_describe = txdata['input1757343186390']
    #         # 家族史
    #         if 'input1757343197188' in txdata:
    #             patient.family_his = txdata['input1757343197188']
    #         # 出生身长
    #         if 'number1757343231377' in txdata:
    #             patient.BL = txdata['number1757343231377']
    #         # 出生体重
    #         if 'number1757343249458' in txdata:
    #             patient.BWt = txdata['number1757343249458']
    #         # 出生胎龄
    #         if 'number1757343266169' in txdata:
    #             patient.ges_week = txdata['number1757343266169']
    #         # 出生胎次
    #         if 'number1757343276810' in txdata:
    #             patient.parity = txdata['number1757343276810']
    #         # 出生产次
    #         if 'number1757343290418' in txdata:
    #             patient.pronum = txdata['number1757343290418']
    #         # 一般症状
    #         if 'input1757343324611' in txdata:
    #             szfyEltm.gen_sym = txdata['input1757343324611']
    #         # 代谢相关症状
    #         if 'input1757343336681' in txdata:
    #             szfyEltm.met_sym = txdata['input1757343336681']
    #         # 骨骼和肌肉症状
    #         if 'input1757343361394' in txdata:
    #             szfyEltm.bone_sym = txdata['input1757343361394']
    #         # 内分泌症状
    #         if 'input1757343370934' in txdata:
    #             szfyEltm.endo_sym = txdata['input1757343370934']
    #         # 其他症状
    #         if 'input1757343393317' in txdata:
    #             szfyEltm.other_sym = txdata['input1757343393317']
    #         # 当前身高
    #         if 'number1757343431046' in txdata:
    #             patient.height = txdata['number1757343431046']
    #         # 当前体重
    #         if 'number1757343459325' in txdata:
    #             patient.weight = txdata['number1757343459325']
    #         # BMI
    #         if 'function_calc1757343540850' in txdata:
    #             patient.bmi = txdata['function_calc1757343540850']
    #         # tanner
    #         if 'select1757343473775label' in txdata:
    #             szfyEltm.tanner = txdata['select1757343473775label']
    #         # 染色体核型
    #         if 'mul_choice1757344138695' in txdata:
    #             chrom = str(txdata['mul_choice1757344138695']['value']).replace('0971136a-ee1d-4d0f-8dbb-cca3565b374f','其他异常核型')
    #             chrom = chrom.replace('14b0f9e2-b2e6-49ad-b97b-34d3f63ec1bf','正常核型')
    #             chrom = chrom.replace('17328f52-5090-4dce-9aab-0dfac86547da','染色体平衡易位')
    #             chrom = chrom.replace('64398b35-4eae-4a33-b51a-8030dd984df5','染色体嵌合体')
    #             chrom = chrom.replace('a3407b7f-ec6c-4531-81ca-5e5b45fc2890','21三体综合征')
    #             chrom = chrom.replace('b5564f31-5508-4168-b287-b082d76a352c','克氏综合征')
    #             chrom = chrom.replace('e5cd20a8-93b5-436a-a16b-98fba2a52050','特纳综合征')
    #             szfyEltm.chrom = chrom
    #         # 其它异常核型0971136a-ee1d-4d0f-8dbb-cca3565b374f
    #         if 'mul_choice1757344138695' in txdata:
    #             szfyEltm.chrom_other = txdata['mul_choice1757344138695']['0971136a-ee1d-4d0f-8dbb-cca3565b374f']
    #         # 基因检测方法
    #         if 'input1757344213776' in txdata:
    #             szfyEltm.gene_method = txdata['input1757344213776']
    #         # 基因结果
    #         if 'single1757344227841' in txdata:
    #             szfyEltm.gene_res = txdata['single1757344227841']['labelList'][txdata['single1757344227841']['value']]
    #         # 基因名称
    #         if 'input1757344256695' in txdata:
    #             szfyEltm.gene_name = txdata['input1757344256695']
    #         # 突变位点
    #         if 'input1757344259204' in txdata:
    #             szfyEltm.gene_point = txdata['input1757344259204']
    #         # 突变类型
    #         if 'input1757344266107' in txdata:
    #             szfyEltm.gene_type = txdata['input1757344266107']
    #         # 遗传模式
    #         if 'input1757344273374' in txdata:
    #             szfyEltm.gene_mode = txdata['input1757344273374']
    #         # 记录日期
    #         if 'date2368659643971' in txdata:
    #             szfyEltm.rec_date = txdata['date2368659643971']
    #         # 有无既往用药史
    #         if 'single8309584011394' in txdata:
    #             szfyEltm.is_has_his = txdata['single8309584011394']['labelList'][txdata['single8309584011394']['value']]
    #         # 用药史
    #         if 'sub_form5565411096460' in txdata:
    #             szfyEltm.has_his = txdata['sub_form5565411096460']
    #         # 药物名称
    #         if 'single0745187438197' in txdata:
    #             szfyEltm.med_name = txdata['single0745187438197']['labelList'][txdata['single0745187438197']['value']]
    #         # 单次剂量
    #         if 'input5804973587911' in txdata:
    #             szfyEltm.dose = txdata['input5804973587911']
    #         # 用药天数
    #         if 'input9327175187008' in txdata:
    #             szfyEltm.days = txdata['input9327175187008']
    #         # 是否停药
    #         if 'single2686629089228' in txdata:
    #             szfyEltm.stop_med = txdata['single2686629089228']['labelList'][txdata['single2686629089228']['value']]
    #         # 停药原因
    #         if 'single2686629089228' in txdata:
    #             szfyEltm.stop_rea = txdata['single2686629089228']['057dd072-873a-45cc-82d2-58236a37c0b2']
    #         # 发生时间
    #         if 'date1757344915422' in txdata:
    #             szfyEltm.star_time = txdata['date1757344915422']
    #         # 结束时间
    #         if 'date1757344919623' in txdata:
    #             szfyEltm.end_time = txdata['date1757344919623']
    #         # 是否为严重不良事件
    #         if 'single1757344943445' in txdata:
    #             szfyEltm.is_adv_eve = txdata['single1757344943445']
    #         # 与研究药物的关系(LA-rhGH)
    #         if 'single1757344975348' in txdata:
    #             szfyEltm.la_rhGH = txdata['single1757344975348']
    #         # 是否调整剂量
    #         if 'single1757345015384' in txdata:
    #             szfyEltm.is_adjust = txdata['single1757345015384']
    #         # 与研究药物的关系(rhGH)
    #         if 'single1757345001172' in txdata:
    #             szfyEltm.rhGH = txdata['single1757345001172']
    #         # 不良事件的转归
    #         if 'single1757345055036' in txdata:
    #             szfyEltm.outcome = txdata['single1757345055036']
    #         try:
    #             with transaction.atomic():
    #                 patient.save()
    #                 szfyEltm.patient_id = patient.pk
    #                 szfyEltm.save()
    #         except:
    #             print("保存数据报错")
    #     else:
    #         print("此条数据非E路童萌数据")



# # 尝试添加任务，如果任务已存在则跳过 【如果要彻x底关闭：删除数据库相应任务，然后注释此代码】
# try:
#     # 添加任务 【'cron:执行方式' month='月', day='天', hour='小时',  minute='分', second='秒', # max_instances=2  [最大同时运行]】
#     scheduler.add_job(create_fake_data, 'cron', month='1-12', day='10', hour='23',  minute='57', second='0',  id=job_id, replace_existing=True, )
#     # 每十秒执行一次
#     # scheduler.add_job(create_fake_data, 'interval', seconds=10, id=job_id, replace_existing=True)
#     # 启动定时任务
#     scheduler.start()
# except ConflictingIdError:
#     print(f"定时任务 '{job_id}' 已存在.")    

